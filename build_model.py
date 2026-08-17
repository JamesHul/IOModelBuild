"""
Generate the IO impact model workbook from build/sources.pkl.

    python scripts/build_model.py

STATUS: this is the proven v0.2 generator (25 tabs, 28,406 formulas, recalculates
clean). It does NOT yet comply with the verbatim-RAW rule in CLAUDE.md. Four
things to fix when the real nine-region data is loaded:

  1. raw_matrix() writes only the 115 industry columns and Q1-Q7. It must write
     the source block whole, including Total Industry Uses, Final Uses and Total
     Supply, and the re-exports row.
  2. Table 5 and Table 8 lose their primary-input rows (T1, P1-P6, Australian
     Production, Value Added). check_sources.py needs those rows for the output
     denominator, so they must survive into RAW.
  3. The state block is expanded from 114 to 115 codes inside the RAW writer.
     Move that to a MAP_Spine tab of formulas.
  4. Multiplier 'n.a.' cells are written as blanks and 6700 is filled by copying
     6701. Both belong in MAP_Multipliers, not in RAW.

The fix in outline: RAW_* writers take whatever the loader gives them and write
it unchanged; add MAP_Spine (code -> row/col position in each RAW block),
MAP_Multipliers (bridge + text handling) and MAP_StripData (column-group
aggregation) as formula layers; CALC_StripData then reads MAP, not RAW.

After every build, recalculate and scan for errors before declaring success.
"""

import pickle, json
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SOURCES = ROOT / 'build' / 'sources.pkl'
OUTFILE = ROOT / 'output' / 'IO_Impact_Model_v0-3.xlsx'

# ---------------------------------------------------------------------------
# INPUT CONTRACT. load_sources.py must provide these. See CLAUDE.md.
#   R['CODES']  list of 115 IOIG codes as text, in ABS order
#   R['T5'] R['T8']            115 x 122 arrays, Australia
#   R['S5'] R['S8']            115 x 122 arrays, selected state
#   R['M'][margin_name]        115 x 122 array per ABS margin table
#   R['TAX']                   115 x 122 array, ABS Table 35
#   R['NAMES']                 {code: industry name}
#   MU['blocks'] MU['effects'] MU['data']  the supplied multiplier set
# ---------------------------------------------------------------------------
_src = pickle.load(open(SOURCES, 'rb'))
R = _src['build_arrays']          # TODO: load_sources.py must populate this
MU = _src['multipliers']
NAMES = R['NAMES']
CODES = R['CODES']; N = len(CODES)
COLG = ['Intermediate', 'HFCE', 'GGFCE', 'GFCF_Priv', 'GFCF_PubCorp', 'GFCF_GG', 'Inventories', 'Exports']
QLAB = ['Households FCE', 'General Government FCE', 'Private GFCF', 'Public Corporations GFCF',
        'General Government GFCF', 'Changes in Inventories', 'Exports of Goods and Services']
NYR = 8; NLINE = 40
MARGIND = [('3301', 'Wholesale trade', ['Wholesale']), ('3901', 'Retail trade', ['Retail']),
           ('4501', 'Food and beverage services', ['RestHotelClub']), ('4601', 'Road transport', ['Road']),
           ('4701', 'Rail transport', ['Rail']), ('4801', 'Water, pipeline and other transport', ['Pipeline', 'Water']),
           ('4901', 'Air and space transport', ['Air']), ('5201', 'Transport support services and storage', ['PortHandling']),
           ('6301', 'Insurance and superannuation funds', ['MarineIns']), ('2701', 'Gas supply', ['Gas']),
           ('2605', 'Electricity transmission and distribution', ['Electricity'])]
OTHER = ['RestHotelClub', 'Road', 'Rail', 'Pipeline', 'Water', 'Air', 'PortHandling', 'MarineIns', 'Gas', 'Electricity']
OTHER_TBL = {'RestHotelClub': 'Table 25', 'Road': 'Table 26', 'Rail': 'Table 27', 'Pipeline': 'Table 28',
             'Water': 'Table 29', 'Air': 'Table 30', 'PortHandling': 'Table 31', 'MarineIns': 'Table 32',
             'Gas': 'Table 33', 'Electricity': 'Table 34'}
MEAS = [(1, 'Output', '$m'), (2, 'Income (wages)', '$m'), (4, 'Value added (basic prices)', '$m'),
        (5, 'Value added (market prices)', '$m'), (6, 'Employment', 'persons')]
EFF = {'Initial': 1, 'FirstRound': 2, 'Simple': 3, 'IndSupport': 4, 'ProdInduced': 5, 'Total': 8, 'ConsInduced': 9}
REGIONS = ['Aus', 'NSW', 'Vic', 'QLD', 'SA', 'WA', 'Tas', 'NT', 'ACT']
LOADED = ['Aus', 'NSW', 'Vic', 'SA']

A_ = 'Arial'
T = Font(name=A_, size=14, bold=True, color='1F4E79'); H2 = Font(name=A_, size=11, bold=True, color='1F4E79')
HD = Font(name=A_, size=9, bold=True, color='FFFFFF'); HF = PatternFill('solid', fgColor='1F4E79')
SUBF = PatternFill('solid', fgColor='DDEBF7'); ORNG = PatternFill('solid', fgColor='FCE4D6')
BLUE = Font(name=A_, size=10, color='0000FF'); BLK = Font(name=A_, size=10)
GRN = Font(name=A_, size=10, color='008000'); RED = Font(name=A_, size=10, bold=True, color='C00000')
NOTE = Font(name=A_, size=9, italic=True, color='595959'); YEL = PatternFill('solid', fgColor='FFFF00')
thin = Side(style='thin', color='BFBFBF'); BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM = '#,##0.0;(#,##0.0);-'; PCT = '0.0%'

wb = openpyxl.Workbook()


def title(ws, txt, sub=None, warn=None):
    ws['A1'] = txt; ws['A1'].font = T
    r = 2
    if warn:
        ws.cell(row=r, column=1, value=warn).font = RED
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12); r += 1
    if sub:
        c = ws.cell(row=r, column=1, value=sub); c.font = NOTE
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 32


def hdr(ws, row, labels, widths=None, start=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=lab)
        c.font = HD; c.fill = HF; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[CL(start + i)].width = w


def raw_matrix(ws, arr, src, note, warn=None, rowoff=0):
    """Write a 115 x 122 ABS matrix. Rows 6.. = products, cols C..DM = 115 industries, DN..DT = Q1..Q7."""
    title(ws, src, note, warn)
    ws.cell(row=4, column=1, value='IOIG').font = HD; ws.cell(row=4, column=1).fill = HF
    ws.cell(row=4, column=2, value='Product / industry').font = HD; ws.cell(row=4, column=2).fill = HF
    ws.column_dimensions['A'].width = 9; ws.column_dimensions['B'].width = 38
    for j, c in enumerate(CODES):
        cell = ws.cell(row=4, column=3 + j, value=c); cell.font = Font(name=A_, size=8, bold=True); cell.fill = SUBF
        ws.cell(row=5, column=3 + j, value=NAMES[c]).font = Font(name=A_, size=7)
        ws.column_dimensions[CL(3 + j)].width = 10
    for j, q in enumerate(QLAB):
        cell = ws.cell(row=4, column=118 + j, value='Q%d' % (j + 1))
        cell.font = Font(name=A_, size=8, bold=True); cell.fill = ORNG
        ws.cell(row=5, column=118 + j, value=q).font = Font(name=A_, size=7)
        ws.column_dimensions[CL(118 + j)].width = 12
    for i, c in enumerate(CODES):
        r = 6 + i + rowoff
        ws.cell(row=r, column=1, value=c).font = BLK
        ws.cell(row=r, column=2, value=NAMES[c]).font = BLK
        for j in range(122):
            v = float(arr[i, j])
            if v != 0.0:
                cell = ws.cell(row=r, column=3 + j, value=round(v, 4))
                cell.font = BLK; cell.number_format = NUM
    ws.freeze_panes = 'C6'
    return 6 + rowoff, 5 + N + rowoff


# ===================================================================== README
ws = wb.active; ws.title = 'README'
title(ws, 'Australian IO Impact Model — v0.2')
rows = [
    ('', ''),
    ('WHAT CHANGED IN v0.2', 'The ABS source tables now live inside the workbook and every strip rate is computed '
                             'from them by formula. Nothing in the calculation chain is a pasted result.'),
    ('', ''),
    ('THE CALCULATION CHAIN', ''),
    ('  1  RAW_T5_Aus / RAW_T8_Aus', 'ABS Table 5 (domestic only) and Table 8 (imports embedded), 2023-24, 115 x 122'),
    ('  2  RAW_T5_State / RAW_T8_State', 'PLACEHOLDER state tables — see the red banner on those tabs'),
    ('  3  RAW_T23 / RAW_T24 / RAW_T25_34 / RAW_T35', 'ABS margin and tax matrices, 2023-24'),
    ('  4  CALC_StripData', 'Aggregates rows 1-3 to 8 purchasing-column groups. Imports = Table 8 less Table 5'),
    ('  5  CALC_PP', 'Purchasers price = domestic + imports + taxes + margins'),
    ('  6  CALC_Strip', 'Each spend line split into its components using the rates from 4 and 5'),
    ('  7  CALC_Margins / CALC_Vector', 'Margins reallocated to the industries that earn them; the direct domestic vector'),
    ('  8  RAW_Multipliers -> ENG_SelectedMult', 'The supplied multiplier set, filtered to the selected region'),
    ('  9  CALC_Impacts -> OUT_Summary', 'Vector x multipliers'),
    (' 10  QA_Checks', '20 gates. Read this before any result'),
    ('', ''),
    ('PLACEHOLDER DATA — READ THIS', ''),
    ('  State flow tables', 'RAW_T5_State and RAW_T8_State hold the NSW 2021-22 block from Master_IO_model_v3, '
                            'expanded onto the 115-code spine with 6700 set to zero. They stand in for EVERY '
                            'non-Australia region until the real nine-region set is loaded. Any state result is '
                            'therefore NSW structure wearing another state name.'),
    ('  Multipliers', 'Aus, NSW, Vic and SA are real. QLD, WA, Tas, NT and ACT are EMPTY, not placeholders — the model '
                      'will refuse to produce a result for them rather than quietly return national numbers.'),
    ('  6700 Imputed rent', 'Has no row in the supplied multiplier set. Bridged to 6701 and flagged.'),
    ('  Margins and taxes by region', 'National only. The ABS publishes no state margin matrices. State studies use '
                                      'national margin and tax rates with state import shares.'),
    ('  Vintage', 'Multipliers 2021-22, ABS tables 2023-24. Unresolved — QA gate 19.'),
    ('', ''),
    ('CONTROL TOTALS (2023-24, 115-code spine, $m)', ''),
    ('  All margins, Tables 23-34', '421,410 — the published 422,034 less the re-exports row, which is not on the spine'),
    ('  Net taxes on products, Table 35', '168,673'),
    ('  Table 5 total intermediate use', '2,136,804'),
    ('  Table 8 less Table 5 = competing imports', '589,795 across all use categories'),
    ('', ''),
    ('HOW TO RUN A STUDY', ''),
    ('  1', 'Save a copy. Never edit this file in place.'),
    ('  2', 'Settings: region, start year, study type, headline basis.'),
    ('  3', 'IN_Shock: enter the spending lines. Read the guidance at the top of that tab first.'),
    ('  4', 'QA_Checks: every gate must read OK.'),
    ('  5', 'OUT_Summary for results, Assumptions to attach to the report.'),
    ('', ''),
    ('CHANGE LOG', ''),
    ('  v0.1', '13 Aug 2026 — strip engine, multiplier lookup, QA gates. Strip data pasted as values.'),
    ('  v0.2', '14 Aug 2026 — ABS source tables embedded; strip data now fully formula-driven; state placeholder '
               'tables added; control-total gates added.'),
]
r = 3
for a, b in rows:
    ws.cell(row=r, column=1, value=a).font = H2 if (a and a.strip().isupper()) else BLK
    c = ws.cell(row=r, column=2, value=b); c.font = BLK
    c.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
ws.column_dimensions['A'].width = 40; ws.column_dimensions['B'].width = 100

# ===================================================================== Settings
ws = wb.create_sheet('Settings')
title(ws, 'Settings', 'Blue cells on a yellow fill are the only ones you change here.')
sets = [('Region', 'Aus', 'Aus uses the real 2023-24 flow tables. Any other region uses the PLACEHOLDER state tables'),
        ('Start year (FY ending)', 2026, 'First year of the shock profile'),
        ('Number of years', NYR, 'Fixed at %d in v0.2' % NYR),
        ('Study type', 'Impact (modelled direct)', 'Impact (modelled direct) | Contribution (measured direct)'),
        ('Headline multiplier basis', 'Type I (simple)', 'Type I (simple) | Type II (total)'),
        ('Shock price year (FY ending)', 2024, 'The year the $ figures are expressed in'),
        ('Multiplier price year (FY ending)', 2022, 'Vintage of the supplied multiplier set'),
        ('Deflator applied to shock', 1.0, 'Set manually to bring the shock to the multiplier price year')]
hdr(ws, 4, ['Setting', 'Value', 'Notes'], [36, 30, 78])
for i, (a, b, c) in enumerate(sets):
    ws.cell(row=5 + i, column=1, value=a).font = BLK
    v = ws.cell(row=5 + i, column=2, value=b); v.font = BLUE; v.fill = YEL; v.border = BOX
    ws.cell(row=5 + i, column=3, value=c).font = NOTE
dv = DataValidation(type='list', formula1='"' + ','.join(REGIONS) + '"'); ws.add_data_validation(dv); dv.add(ws['B5'])
dv2 = DataValidation(type='list', formula1='"Type I (simple),Type II (total)"'); ws.add_data_validation(dv2); dv2.add(ws['B9'])
ws['A15'] = 'Derived by the engine'; ws['A15'].font = H2
ws['A16'] = 'Flow-table source in use'
ws['B16'] = '=IF(B5="Aus","RAW_T5_Aus / RAW_T8_Aus (real, 2023-24)","RAW_T5_State / RAW_T8_State (PLACEHOLDER - NSW 2021-22)")'
ws['B16'].font = GRN
ws['A17'] = 'Headline effect used for the total'
ws['B17'] = '=IF(B9="Type II (total)","Total","Simple")'; ws['B17'].font = GRN
ws.column_dimensions['B'].width = 60

# ===================================================================== Lists
ws = wb.create_sheet('Lists')
title(ws, 'Lists — the canonical spine and every code the model looks up',
      'IOIG(2022), 115 industries. Column D bridges to the supplied multiplier set, which uses 114 codes.')
hdr(ws, 4, ['#', 'IOIG code', 'Industry / product name', 'Multiplier source code', 'Bridge note'],
    [6, 11, 52, 20, 44])
for i, c in enumerate(CODES):
    src = '6701' if c == '6700' else c
    for j, v in enumerate([i + 1, c, NAMES[c], src,
                           'PLACEHOLDER: borrows 6701 — vendor to supply 6700' if c == '6700' else '']):
        cell = ws.cell(row=5 + i, column=1 + j, value=v); cell.font = BLK; cell.border = BOX
        if c == '6700' and j >= 3:
            cell.fill = YEL
ws.freeze_panes = 'A5'
r0 = 5 + N + 2
ws.cell(row=r0, column=1, value='Purchasing column groups').font = H2
hdr(ws, r0 + 1, ['#', 'Group code', 'ABS column it aggregates', 'When to use it'], [6, 16, 34, 62])
gdesc = [('Intermediate', 'Sum of all 115 industry columns', 'A business buying an input — including a contractor buying materials'),
         ('HFCE', 'Q1 Households FCE', 'Consumer spending, tourism, events'),
         ('GGFCE', 'Q2 General Government FCE', 'Government service delivery'),
         ('GFCF_Priv', 'Q3 Private GFCF', 'Private capital projects — the finished asset'),
         ('GFCF_PubCorp', 'Q4 Public Corporations GFCF', 'GBE capex'),
         ('GFCF_GG', 'Q5 General Government GFCF', 'Government capital projects — the finished asset'),
         ('Inventories', 'Q6 Changes in Inventories', 'Rarely a shock'),
         ('Exports', 'Q7 Exports', 'Export sales — import share unreliable, see QA gate 13')]
for i, (g, src, use) in enumerate(gdesc):
    for j, v in enumerate([i + 1, g, src, use]):
        ws.cell(row=r0 + 2 + i, column=1 + j, value=v).font = BLK
r1 = r0 + 2 + len(COLG) + 2
ws.cell(row=r1, column=1, value='Margin destinations — which industry earns each ABS margin').font = H2
hdr(ws, r1 + 1, ['#', 'IOIG', 'Industry', 'ABS table', 'Total $m 2023-24'], [6, 10, 44, 26, 18])
tots = {n: float(R['M'][n].sum()) for n in R['M']}
for i, (code, nm, srcs) in enumerate(MARGIND):
    tbl = ', '.join('Table 23' if s == 'Wholesale' else 'Table 24' if s == 'Retail' else OTHER_TBL[s] for s in srcs)
    for j, v in enumerate([i + 1, code, nm, tbl, round(sum(tots[s] for s in srcs), 0)]):
        cell = ws.cell(row=r1 + 2 + i, column=1 + j, value=v); cell.font = BLK
        if j == 4: cell.number_format = '#,##0'
r2 = r1 + 2 + len(MARGIND)
ws.cell(row=r2, column=3, value='TOTAL — control total for QA gate 17').font = H2
ws.cell(row=r2, column=5, value=round(sum(tots.values()), 0)).font = Font(name=A_, size=10, bold=True)
ws.cell(row=r2, column=5).number_format = '#,##0'
ws.cell(row=r2 + 2, column=1, value='Regions').font = H2
for i, rg in enumerate(REGIONS):
    ws.cell(row=r2 + 3 + i, column=2, value=rg).font = BLK
    ws.cell(row=r2 + 3 + i, column=3,
            value='real multipliers' if rg in LOADED else 'NOT LOADED — model will refuse').font = (
        BLK if rg in LOADED else RED)
LIST_CODE = 'Lists!$B$5:$B$%d' % (4 + N)
LIST_SRC = 'Lists!$D$5:$D$%d' % (4 + N)
LIST_GRP = 'Lists!$B$%d:$B$%d' % (r0 + 2, r0 + 1 + len(COLG))
MARG_CTRL = 'Lists!$E$%d' % r2

# ===================================================================== RAW ABS tables
ws = wb.create_sheet('RAW_T5_Aus')
raw_matrix(ws, R['T5'], 'RAW — ABS Table 5. Industry by industry flow table (DIRECT allocation of imports), 2023-24 ($m)',
           'Layout: rows are products, columns C to DM are the 115 using industries, DN to DT are final use Q1-Q7. Source: ABS 5209.0.55.001, released 25 March 2026, file 520905500105. Intermediate and final use here are '
           'DOMESTIC PRODUCTION ONLY — competing imports are shown separately as a primary input. This is the table '
           'multipliers must be built from, and the numerator of the domestic share in CALC_StripData.')
ws = wb.create_sheet('RAW_T8_Aus')
raw_matrix(ws, R['T8'], 'RAW — ABS Table 8. Industry by industry flow table (INDIRECT allocation of imports), 2023-24 ($m)',
           'Layout: rows are products, columns C to DM are the 115 using industries, DN to DT are final use Q1-Q7. Source: ABS 5209.0.55.001, file 520905500108. Same cells as Table 5 but with competing imports embedded in '
           'the flows. Table 8 less Table 5, cell by cell, IS the import content — that is how CALC_StripData derives '
           'import shares without needing ABS Table 3.')
ws = wb.create_sheet('RAW_T5_State')
raw_matrix(ws, R['S5'], 'RAW — Table 5, state (PLACEHOLDER)',
           'Placeholder content: the NSW block from Master_IO_model_v3_28April26.xlsm, Table 5 tab, vintage 2021-22, '
           '114 industries expanded onto the 115-code spine with 6700 set to zero. Replace with the real nine-region set.',
           warn='*** PLACEHOLDER DATA — NSW 2021-22, USED FOR EVERY NON-AUSTRALIA REGION. NOT A REAL STATE TABLE. ***')
ws = wb.create_sheet('RAW_T8_State')
raw_matrix(ws, R['S8'], 'RAW — Table 8, state (PLACEHOLDER)',
           'Placeholder content: the NSW block from Master_IO_model_v3_28April26.xlsm, Table 8 tab, vintage 2021-22.',
           warn='*** PLACEHOLDER DATA — NSW 2021-22, USED FOR EVERY NON-AUSTRALIA REGION. NOT A REAL STATE TABLE. ***')
ws = wb.create_sheet('RAW_T23_Wholesale')
raw_matrix(ws, R['M']['Wholesale'], 'RAW — ABS Table 23. Wholesale margin by product by using industry and final use, 2023-24 ($m)',
           'Source: ABS file 520905500123. Total on the 115 spine: %s. The wholesale margin is earned by IOIG 3301, '
           'not by the producer of the good.' % f"{tots['Wholesale']:,.0f}")
ws = wb.create_sheet('RAW_T24_Retail')
raw_matrix(ws, R['M']['Retail'], 'RAW — ABS Table 24. Retail margin by product by using industry and final use, 2023-24 ($m)',
           'Source: ABS file 520905500124. Total on the 115 spine: %s. Note how concentrated this is in the Q1 '
           'household column — retail margin against all 115 industry columns combined is only about $6.7bn.'
           % f"{tots['Retail']:,.0f}")
ws = wb.create_sheet('RAW_T35_NetTaxes')
raw_matrix(ws, R['TAX'], 'RAW — ABS Table 35. Net taxes on products by product by using industry and final use, 2023-24 ($m)',
           'Source: ABS file 520905500135. Total on the 115 spine: %s. Taxes less subsidies, so cells can be negative. '
           'GST, duty, excise and subsidies can be separated later using ABS Tables 36-39.' % f"{R['TAX'].sum():,.0f}")

# stacked other margins
ws = wb.create_sheet('RAW_T25_34_Margins')
title(ws, 'RAW — ABS Tables 25 to 34. The ten remaining margin types, 2023-24 ($m)',
      'Stacked, one 115-row block per margin type, in the order listed in column A. Same column layout as the other '
      'RAW tabs. Sources: ABS files 520905500125 to 520905500134.')
ws.cell(row=4, column=1, value='Margin type').font = HD; ws.cell(row=4, column=1).fill = HF
ws.cell(row=4, column=2, value='IOIG').font = HD; ws.cell(row=4, column=2).fill = HF
ws.cell(row=4, column=3, value='Product').font = HD; ws.cell(row=4, column=3).fill = HF
ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 9; ws.column_dimensions['C'].width = 34
for j, c in enumerate(CODES):
    cell = ws.cell(row=4, column=4 + j, value=c); cell.font = Font(name=A_, size=8, bold=True); cell.fill = SUBF
    ws.cell(row=5, column=4 + j, value=NAMES[c]).font = Font(name=A_, size=7)
    ws.column_dimensions[CL(4 + j)].width = 10
for j, q in enumerate(QLAB):
    cell = ws.cell(row=4, column=119 + j, value='Q%d' % (j + 1))
    cell.font = Font(name=A_, size=8, bold=True); cell.fill = ORNG
    ws.cell(row=5, column=119 + j, value=q).font = Font(name=A_, size=7)
    ws.column_dimensions[CL(119 + j)].width = 12
OTHER_ROW0 = {}
rr = 6
for t in OTHER:
    OTHER_ROW0[t] = rr
    arr = R['M'][t]
    for i, c in enumerate(CODES):
        ws.cell(row=rr, column=1, value='%s (%s)' % (t, OTHER_TBL[t])).font = BLK
        ws.cell(row=rr, column=2, value=c).font = BLK
        ws.cell(row=rr, column=3, value=NAMES[c]).font = BLK
        for j in range(122):
            v = float(arr[i, j])
            if v != 0.0:
                cell = ws.cell(row=rr, column=4 + j, value=round(v, 4)); cell.font = BLK; cell.number_format = NUM
        rr += 1
ws.freeze_panes = 'D6'

# ===================================================================== CALC_StripData
ws = wb.create_sheet('CALC_StripData')
title(ws, 'Strip data — the ABS tables aggregated to purchasing-column groups ($m)',
      'Every cell is a formula reading the RAW tabs. Intermediate = the sum of all 115 industry columns; each Q '
      'column is a direct reference. Domestic and Imports switch between the Australian and the placeholder state '
      'tables depending on Settings!B5. Row 4 repeats the column group so CALC_PP can SUMIF across the blocks.')
BLOCKS = ['Domestic_BP', 'Imports', 'NetTaxes'] + ['Margin_' + c for c, _, _ in MARGIND]
NB = len(BLOCKS)
ws.cell(row=3, column=1, value='IOIG').font = HD; ws.cell(row=3, column=1).fill = HF
ws.cell(row=3, column=2, value='Product').font = HD; ws.cell(row=3, column=2).fill = HF
ws.column_dimensions['A'].width = 9; ws.column_dimensions['B'].width = 34
for j, bn in enumerate(BLOCKS):
    c0 = 3 + j * len(COLG)
    ws.cell(row=3, column=c0, value=bn).font = HD; ws.cell(row=3, column=c0).fill = HF
    for k, g in enumerate(COLG):
        c = ws.cell(row=4, column=c0 + k, value=g); c.font = Font(name=A_, size=8, bold=True); c.fill = SUBF
        ws.column_dimensions[CL(c0 + k)].width = 12
IND_A, IND_Z = CL(3), CL(117)          # industry columns on the T5/T8/T23/T24/T35 tabs
IND_A2, IND_Z2 = CL(4), CL(118)        # on the stacked margins tab


def agg(tab, row, k, off=0):
    """Aggregate one ABS tab row to column group k. off=1 for the stacked margins tab."""
    a, z = (IND_A2, IND_Z2) if off else (IND_A, IND_Z)
    if k == 0:
        return '{t}!${a}{r}:${z}{r}'.format(t=tab, a=a, z=z, r=row), True
    return '{t}!${c}{r}'.format(t=tab, c=CL((118 + off) + (k - 1)), r=row), False


def ref(tab, row, k, off=0):
    s, isrange = agg(tab, row, k, off)
    return ('SUM(%s)' % s) if isrange else s


for i, c in enumerate(CODES):
    r = 5 + i; src = 6 + i
    ws.cell(row=r, column=1, value=c).font = BLK
    ws.cell(row=r, column=2, value=NAMES[c]).font = BLK
    for k in range(len(COLG)):
        # Domestic
        f = '=IF(Settings!$B$5="Aus",%s,%s)' % (ref('RAW_T5_Aus', src, k), ref('RAW_T5_State', src, k))
        cell = ws.cell(row=r, column=3 + k, value=f); cell.font = BLK; cell.number_format = NUM
        # Imports = T8 - T5
        f = ('=IF(Settings!$B$5="Aus",%s-%s,%s-%s)'
             % (ref('RAW_T8_Aus', src, k), ref('RAW_T5_Aus', src, k),
                ref('RAW_T8_State', src, k), ref('RAW_T5_State', src, k)))
        cell = ws.cell(row=r, column=3 + len(COLG) + k, value=f); cell.font = BLK; cell.number_format = NUM
        # Net taxes
        cell = ws.cell(row=r, column=3 + 2 * len(COLG) + k, value='=' + ref('RAW_T35_NetTaxes', src, k))
        cell.font = BLK; cell.number_format = NUM
        # margins
        for m, (code, nm, srcs) in enumerate(MARGIND):
            parts = []
            for s in srcs:
                if s == 'Wholesale':
                    parts.append(ref('RAW_T23_Wholesale', src, k))
                elif s == 'Retail':
                    parts.append(ref('RAW_T24_Retail', src, k))
                else:
                    parts.append(ref('RAW_T25_34_Margins', OTHER_ROW0[s] + i, k, off=1))
            cell = ws.cell(row=r, column=3 + (3 + m) * len(COLG) + k, value='=' + '+'.join(parts))
            cell.font = BLK; cell.number_format = NUM
ws.freeze_panes = 'C5'
LASTC = 2 + NB * len(COLG)
tr = 5 + N
ws.cell(row=tr, column=2, value='TOTAL — control totals').font = H2
for j in range(NB * len(COLG)):
    cell = ws.cell(row=tr, column=3 + j, value='=SUM(%s$5:%s$%d)' % (CL(3 + j), CL(3 + j), 4 + N))
    cell.font = Font(name=A_, size=10, bold=True); cell.number_format = NUM
ws.cell(row=tr + 2, column=2, value='All margins, all column groups (should equal Lists control total)').font = BLK
ws.cell(row=tr + 2, column=3, value='=SUM($%s$%d:$%s$%d)' % (CL(3 + 3 * len(COLG)), tr, CL(LASTC), tr))
ws.cell(row=tr + 2, column=3).font = Font(name=A_, size=10, bold=True); ws.cell(row=tr + 2, column=3).number_format = NUM
ws.cell(row=tr + 3, column=2, value='Published control total (Lists)').font = BLK
ws.cell(row=tr + 3, column=3, value='=%s' % MARG_CTRL); ws.cell(row=tr + 3, column=3).number_format = NUM
STRIPDATA_CTRL = 'CALC_StripData!$C$%d' % (tr + 2)
RAW_ALL = 'CALC_StripData!$C$5:$%s$%d' % (CL(LASTC), 4 + N)
RAW_HDR = 'CALC_StripData!$C$4:$%s$4' % CL(LASTC)

# ===================================================================== CALC_PP
ws = wb.create_sheet('CALC_PP')
title(ws, 'Purchasers price by product and column group ($m)',
      'PP = domestic + imports + net taxes + all eleven margin blocks, summed across CALC_StripData by column group. '
      'This is the denominator of every strip rate.')
hdr(ws, 4, ['IOIG', 'Product'] + COLG, [10, 38] + [14] * len(COLG))
for i, c in enumerate(CODES):
    ws.cell(row=5 + i, column=1, value=c).font = BLK
    ws.cell(row=5 + i, column=2, value=NAMES[c]).font = BLK
    for k in range(len(COLG)):
        f = '=SUMIF(%s,%s$4,CALC_StripData!$C%d:$%s%d)' % (RAW_HDR, CL(3 + k), 5 + i, CL(LASTC), 5 + i)
        cell = ws.cell(row=5 + i, column=3 + k, value=f); cell.font = BLK; cell.number_format = NUM
ws.freeze_panes = 'C5'
PP_RANGE = 'CALC_PP!$C$5:$%s$%d' % (CL(2 + len(COLG)), 4 + N)

# ===================================================================== CALC_Rates (visible rates)
ws = wb.create_sheet('CALC_Rates')
title(ws, 'Strip rates — share of each purchasers-price dollar, by product and column group',
      'Shown for inspection and for the sense-checks in reports. CALC_Strip computes the same ratios inline. '
      'The two Check columns must read 100% wherever the ABS cell is non-zero.')
sub = ['Domestic', 'Imports', 'NetTaxes', 'All margins']
hdr(ws, 4, ['IOIG', 'Product'] + ['%s\n%s' % (s, g) for s in sub for g in ['Intermediate', 'HFCE']]
    + ['Check\nIntermediate', 'Check\nHFCE'], [10, 34] + [13] * 10)
for i, c in enumerate(CODES):
    r = 5 + i
    ws.cell(row=r, column=1, value=c).font = BLK
    ws.cell(row=r, column=2, value=NAMES[c]).font = BLK
    for si, s in enumerate(sub):
        for gi, k in enumerate([0, 1]):
            if s == 'All margins':
                num = 'SUM(CALC_StripData!$%s%d:$%s%d)' % (CL(3 + 3 * len(COLG) + k), r, CL(LASTC), r)
                num = 'SUMIF(%s,CALC_PP!%s$4,CALC_StripData!$%s%d:$%s%d)-CALC_StripData!$%s%d-CALC_StripData!$%s%d-CALC_StripData!$%s%d' % (
                    RAW_HDR, CL(3 + k), CL(3), r, CL(LASTC), r,
                    CL(3 + k), r, CL(3 + len(COLG) + k), r, CL(3 + 2 * len(COLG) + k), r)
            else:
                num = 'CALC_StripData!$%s%d' % (CL(3 + si * len(COLG) + k), r)
            f = '=IFERROR((%s)/CALC_PP!$%s%d,"")' % (num, CL(3 + k), r)
            cell = ws.cell(row=r, column=3 + si * 2 + gi, value=f); cell.font = BLK; cell.number_format = PCT
    for gi in range(2):
        f = '=IFERROR($%s%d+$%s%d+$%s%d+$%s%d,"")' % (CL(3 + gi), r, CL(5 + gi), r, CL(7 + gi), r, CL(9 + gi), r)
        cell = ws.cell(row=r, column=11 + gi, value=f); cell.font = BLK; cell.number_format = PCT
ws.freeze_panes = 'C5'

# ===================================================================== RAW_Multipliers
ws = wb.create_sheet('RAW_Multipliers')
title(ws, 'RAW — supplied multiplier set. PASTE TARGET.',
      'One row per region and industry. Aus, NSW, Vic and SA carry real data from the master model. QLD, WA, Tas, NT '
      'and ACT are intentionally EMPTY so that QA gate 2 blocks a run rather than returning national numbers under a '
      'state label. Do not compute anything on this tab.')
ws.cell(row=3, column=1, value='Key').font = HD; ws.cell(row=3, column=1).fill = HF
for j, lab in enumerate(['Region', 'Code', 'Name']):
    ws.cell(row=3, column=2 + j, value=lab).font = HD; ws.cell(row=3, column=2 + j).fill = HF
for m, bn in enumerate(MU['blocks']):
    c0 = 5 + m * 11
    ws.cell(row=3, column=c0, value=bn).font = HD; ws.cell(row=3, column=c0).fill = HF
    for k, e in enumerate(MU['effects']):
        c = ws.cell(row=4, column=c0 + k, value=e); c.font = Font(name=A_, size=8, bold=True); c.fill = SUBF
        ws.column_dimensions[CL(c0 + k)].width = 11
for w, col in zip([16, 8, 8, 34], ['A', 'B', 'C', 'D']):
    ws.column_dimensions[col].width = w
rr = 5
for rg in REGIONS:
    for c in CODES:
        src = '6701' if c == '6700' else c
        ws.cell(row=rr, column=1, value='%s|%s' % (rg, c)).font = BLK
        ws.cell(row=rr, column=2, value=rg).font = BLK
        ws.cell(row=rr, column=3, value=c).font = BLK
        ws.cell(row=rr, column=4, value=NAMES[c]).font = BLK
        vals = MU['data'].get('%s|%s' % (rg, src)) if rg in LOADED else None
        if vals:
            for k, v in enumerate(vals):
                if v is not None:
                    cell = ws.cell(row=rr, column=5 + k, value=float(v)); cell.font = BLK; cell.number_format = '0.0000'
        else:
            ws.cell(row=rr, column=5).fill = YEL
        rr += 1
ws.freeze_panes = 'E5'
MULT_LAST = 4 + len(MU['blocks']) * 11
MULT_KEY = 'RAW_Multipliers!$A$5:$A$%d' % (rr - 1)
MULT_DATA = 'RAW_Multipliers!$E$5:$%s$%d' % (CL(MULT_LAST), rr - 1)

# ===================================================================== ENG_SelectedMult
ws = wb.create_sheet('ENG_SelectedMult')
title(ws, 'Multipliers for the selected region',
      'INDEX/MATCH into RAW_Multipliers on Region|Code. Blank means the region is not loaded. Column C shows the '
      'bridge code actually used, which differs from column A only for 6700.')
cols = []
for mi, mname, unit in MEAS:
    for ename in ['Initial', 'ProdInduced', 'ConsInduced', 'Total']:
        cols.append((mname, ename, (mi - 1) * 11 + EFF[ename]))
for ename in ['FirstRound', 'Simple', 'IndSupport']:
    cols.append(('Output', ename, EFF[ename]))
hdr(ws, 4, ['IOIG', 'Industry', 'Source code'] + ['%s\n%s' % (a, b) for a, b, _ in cols],
    [10, 34, 12] + [13] * len(cols))
for i, c in enumerate(CODES):
    ws.cell(row=5 + i, column=1, value=c).font = BLK
    ws.cell(row=5 + i, column=2, value=NAMES[c]).font = BLK
    ws.cell(row=5 + i, column=3, value='=INDEX(%s,%d)' % (LIST_SRC, i + 1)).font = GRN
    for j, (_, _, off) in enumerate(cols):
        f = '=IFERROR(INDEX(%s,MATCH(Settings!$B$5&"|"&$A%d,%s,0),%d),"")' % (MULT_DATA, 5 + i, MULT_KEY, off)
        cell = ws.cell(row=5 + i, column=4 + j, value=f); cell.font = BLK; cell.number_format = '0.0000'
ws.freeze_panes = 'D5'


def mcol(mname, ename):
    for j, (a, b, _) in enumerate(cols):
        if a == mname and b == ename:
            return CL(4 + j)
    raise KeyError((mname, ename))


# ===================================================================== IN_Shock
ws = wb.create_sheet('IN_Shock')
title(ws, 'INPUT — the spending shock',
      'One row per spending line, at PURCHASERS PRICES. Allocate each line to the industry you immediately pay, '
      'never to its suppliers — the multiplier handles the supply chain.')
gnote = [
    'CHOOSING THE PURCHASING COLUMN GROUP:',
    '  HFCE / GGFCE   the buyer is a household, or a government department buying a service.',
    '  GFCF_*         the buyer is acquiring a finished capital asset. These ABS columns hold construction services, '
    'machinery and software — NOT steel, concrete or aggregate. Nationally the whole General Government GFCF column '
    'shows $20m of iron and steel against $35bn of civil construction.',
    '  Intermediate   the buyer is a business purchasing an input. A contractor buying steel sits here.',
    '',
    'TWO WAYS TO MODEL A CAPITAL PROJECT:',
    '  Top-down (recommended)  one line: contract value against 3101 or 3002, column group GFCF_GG or GFCF_Priv. '
    'The multiplier generates the steel, concrete, plant and design automatically.',
    '  Bottom-up               the bill of quantities against the Intermediate group. The contractor\'s own wages and '
    'margin line must be set to Treatment = Direct only, or you double count the supply chain the multiplier builds.',
]
for i, t in enumerate(gnote):
    c = ws.cell(row=4 + i, column=1, value=t)
    c.font = H2 if t.endswith(':') else NOTE
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=4 + i, start_column=1, end_row=4 + i, end_column=9)
    ws.row_dimensions[4 + i].height = 30 if len(t) > 110 else 14
HR = 4 + len(gnote) + 1
hdr(ws, HR, ['Line', 'Description of spending', 'IOIG product bought', 'Purchasing column group', 'Include?', 'Treatment']
    + ['FY%d' % (2026 + i) for i in range(NYR)], [6, 40, 16, 20, 9, 14] + [11] * NYR)
for i in range(NYR):
    ws.cell(row=HR - 1, column=7 + i, value='=Settings!$B$6+%d' % i).font = GRN
example = [
    ('Groceries bought by households', '1101', 'HFCE', 'Y', 'Flow-on', [10, 10, 10, 0, 0, 0, 0, 0]),
    ('Clothing bought by households', '1305', 'HFCE', 'Y', 'Flow-on', [4, 4, 4, 0, 0, 0, 0, 0]),
    ('Electricity bought by households', '2601', 'HFCE', 'Y', 'Flow-on', [6, 6, 6, 0, 0, 0, 0, 0]),
    ('Bridge contract — civil construction (top-down)', '3101', 'GFCF_GG', 'Y', 'Flow-on', [0, 140, 190, 140, 0, 0, 0, 0]),
    ('Design and engineering engaged directly', '6901', 'GFCF_GG', 'Y', 'Flow-on', [8, 10, 6, 3, 0, 0, 0, 0]),
    ('EXAMPLE bottom-up: structural steel', '2101', 'Intermediate', 'N', 'Flow-on', [0, 20, 30, 20, 0, 0, 0, 0]),
    ('EXAMPLE bottom-up: ready-mixed concrete', '2003', 'Intermediate', 'N', 'Flow-on', [0, 15, 25, 15, 0, 0, 0, 0]),
    ('EXAMPLE bottom-up: contractor wages and margin', '3101', 'Intermediate', 'N', 'Direct only', [0, 40, 50, 30, 0, 0, 0, 0]),
]
for i in range(NLINE):
    r = HR + 1 + i
    ws.cell(row=r, column=1, value=i + 1).font = BLK
    if i < len(example):
        d, code, grp, inc, trt, yrs = example[i]
        for j, v in enumerate([d, code, grp, inc, trt]):
            cell = ws.cell(row=r, column=2 + j, value=v); cell.font = BLUE; cell.fill = YEL; cell.border = BOX
        for j, v in enumerate(yrs):
            cell = ws.cell(row=r, column=7 + j, value=float(v))
            cell.font = BLUE; cell.fill = YEL; cell.border = BOX; cell.number_format = NUM
    else:
        for j in range(5):
            cell = ws.cell(row=r, column=2 + j); cell.fill = YEL; cell.border = BOX; cell.font = BLUE
        for j in range(NYR):
            cell = ws.cell(row=r, column=7 + j); cell.fill = YEL; cell.border = BOX
            cell.font = BLUE; cell.number_format = NUM
for dvspec, rng in [(LIST_CODE, 'C'), (LIST_GRP, 'D')]:
    d = DataValidation(type='list', formula1=dvspec); ws.add_data_validation(d)
    d.add('%s%d:%s%d' % (rng, HR + 1, rng, HR + NLINE))
for f1, rng in [('"Y,N"', 'E'), ('"Flow-on,Direct only"', 'F')]:
    d = DataValidation(type='list', formula1=f1); ws.add_data_validation(d)
    d.add('%s%d:%s%d' % (rng, HR + 1, rng, HR + NLINE))
tr = HR + NLINE + 2
ws.cell(row=tr, column=2, value='TOTAL SHOCK INCLUDED (purchasers prices, $m)').font = H2
for j in range(NYR):
    cell = ws.cell(row=tr, column=7 + j, value='=SUMIF($E$%d:$E$%d,"Y",%s$%d:%s$%d)'
                   % (HR + 1, HR + NLINE, CL(7 + j), HR + 1, CL(7 + j), HR + NLINE))
    cell.font = BLK; cell.number_format = NUM
ws.cell(row=tr + 1, column=2, value='Example rows are illustrative — overwrite them. Lines 6-8 show the bottom-up '
                                    'alternative and are switched off.').font = NOTE
ws.freeze_panes = 'G%d' % (HR + 1)
SHOCK_R0 = HR + 1

# ===================================================================== CALC_Strip
ws = wb.create_sheet('CALC_Strip')
title(ws, 'Step 3 — strip each line to domestic basic prices',
      'One row per spending line and year. RowIdx and ColIdx locate the ABS cell; PP is read from CALC_PP; each '
      'component is the line value times that component divided by PP. The Check column must be zero.')
comp = ['Domestic', 'Imports', 'NetTaxes'] + ['Marg_' + c for c, _, _ in MARGIND]
hdr(ws, 4, ['Line', 'Year#', 'Year', 'Product', 'Group', 'Include', 'Treatment', 'Spend $m', 'RowIdx', 'ColIdx', 'PP $m']
    + comp + ['Check'], [6, 7, 8, 10, 14, 9, 13, 11, 8, 8, 11] + [11] * len(comp) + [10])
r = 5
for li in range(NLINE):
    for y in range(NYR):
        ws.cell(row=r, column=1, value=li + 1).font = BLK
        ws.cell(row=r, column=2, value=y + 1).font = BLK
        ws.cell(row=r, column=3, value='=Settings!$B$6+%d' % y).font = GRN
        for cc, sc in [(4, 'C'), (5, 'D'), (6, 'E'), (7, 'F')]:
            ws.cell(row=r, column=cc, value='=IN_Shock!$%s$%d' % (sc, SHOCK_R0 + li)).font = GRN
        ws.cell(row=r, column=8, value='=IF($F%d="Y",IFERROR(IN_Shock!%s%d,0)*Settings!$B$12,0)'
                % (r, CL(7 + y), SHOCK_R0 + li)).font = GRN
        ws.cell(row=r, column=9, value='=IFERROR(MATCH($D%d,%s,0),0)' % (r, LIST_CODE)).font = BLK
        ws.cell(row=r, column=10, value='=IFERROR(MATCH($E%d,%s,0),0)' % (r, LIST_GRP)).font = BLK
        ws.cell(row=r, column=11, value='=IF(OR($I%d=0,$J%d=0),0,INDEX(%s,$I%d,$J%d))' % (r, r, PP_RANGE, r, r)).font = BLK
        for k in range(len(comp)):
            f = ('=IF(OR($I{r}=0,$J{r}=0,$K{r}=0),0,$H{r}*INDEX({rng},$I{r},{off}+$J{r})/$K{r})'
                 .format(r=r, rng=RAW_ALL, off=k * len(COLG)))
            cell = ws.cell(row=r, column=12 + k, value=f); cell.font = BLK; cell.number_format = NUM
        ws.cell(row=r, column=12 + len(comp),
                value='=ROUND(SUM($L%d:$%s%d)-$H%d,6)' % (r, CL(11 + len(comp)), r, r)).font = BLK
        r += 1
ws.freeze_panes = 'I5'
STRIP_LAST = r - 1

# ===================================================================== CALC_Margins
ws = wb.create_sheet('CALC_Margins')
title(ws, 'Margin dollars by earning industry and year ($m)',
      'Margin stripped out of a line is not lost — it is income to the industry that earned it, and it flows into '
      'CALC_Vector as direct domestic production for that industry.')
hdr(ws, 4, ['IOIG', 'Industry'] + ['Y%d' % (i + 1) for i in range(NYR)], [10, 44] + [12] * NYR)
for i, (code, nm, _) in enumerate(MARGIND):
    ws.cell(row=5 + i, column=1, value=code).font = BLK
    ws.cell(row=5 + i, column=2, value=nm).font = BLK
    for y in range(NYR):
        col = CL(12 + 3 + i)
        f = '=SUMIF(CALC_Strip!$B$5:$B$%d,%d,CALC_Strip!$%s$5:$%s$%d)' % (STRIP_LAST, y + 1, col, col, STRIP_LAST)
        cell = ws.cell(row=5 + i, column=3 + y, value=f); cell.font = BLK; cell.number_format = NUM
MARG_RANGE = 'CALC_Margins!$C$5:$%s$%d' % (CL(2 + NYR), 4 + len(MARGIND))
MARG_CODES = 'CALC_Margins!$A$5:$A$%d' % (4 + len(MARGIND))

# ===================================================================== CALC_Vector
ws = wb.create_sheet('CALC_Vector')
title(ws, 'Step 3 output — direct domestic production by industry and year ($m, basic prices)',
      'Left block: lines that receive the full multiplier. Right block: lines flagged Direct only. Margins always sit '
      'in the flow-on block.')
ws.cell(row=3, column=3, value='FLOW-ON LINES').font = H2
ws.cell(row=3, column=3 + NYR, value='DIRECT-ONLY LINES').font = H2
hdr(ws, 4, ['IOIG', 'Industry'] + ['FY%d' % (2026 + i) for i in range(NYR)]
    + ['FY%d' % (2026 + i) for i in range(NYR)] + ['Total all'], [10, 42] + [12] * (2 * NYR) + [13])
for i, c in enumerate(CODES):
    r = 5 + i
    ws.cell(row=r, column=1, value=c).font = BLK
    ws.cell(row=r, column=2, value=NAMES[c]).font = BLK
    for y in range(NYR):
        f = ('=SUMIFS(CALC_Strip!$L$5:$L$%d,CALC_Strip!$D$5:$D$%d,$A%d,CALC_Strip!$B$5:$B$%d,%d,'
             'CALC_Strip!$G$5:$G$%d,"Flow-on")+IFERROR(INDEX(%s,MATCH($A%d,%s,0),%d),0)'
             % (STRIP_LAST, STRIP_LAST, r, STRIP_LAST, y + 1, STRIP_LAST, MARG_RANGE, r, MARG_CODES, y + 1))
        ws.cell(row=r, column=3 + y, value=f).number_format = NUM
        ws.cell(row=r, column=3 + y).font = BLK
        f2 = ('=SUMIFS(CALC_Strip!$L$5:$L$%d,CALC_Strip!$D$5:$D$%d,$A%d,CALC_Strip!$B$5:$B$%d,%d,'
              'CALC_Strip!$G$5:$G$%d,"Direct only")' % (STRIP_LAST, STRIP_LAST, r, STRIP_LAST, y + 1, STRIP_LAST))
        ws.cell(row=r, column=3 + NYR + y, value=f2).number_format = NUM
        ws.cell(row=r, column=3 + NYR + y).font = BLK
    ws.cell(row=r, column=3 + 2 * NYR, value='=SUM($C%d:$%s%d)' % (r, CL(2 + 2 * NYR), r)).font = BLK
    ws.cell(row=r, column=3 + 2 * NYR).number_format = NUM
tr = 5 + N
ws.cell(row=tr, column=2, value='TOTAL DIRECT DOMESTIC').font = H2
for y in range(2 * NYR + 1):
    cell = ws.cell(row=tr, column=3 + y, value='=SUM(%s$5:%s$%d)' % (CL(3 + y), CL(3 + y), 4 + N))
    cell.font = Font(name=A_, size=10, bold=True); cell.number_format = NUM
ws.freeze_panes = 'C5'
VEC = lambda y: 'CALC_Vector!$%s$5:$%s$%d' % (CL(3 + y), CL(3 + y), 4 + N)
VECD = lambda y: 'CALC_Vector!$%s$5:$%s$%d' % (CL(3 + NYR + y), CL(3 + NYR + y), 4 + N)
VEC_TOTCOL = CL(3 + 2 * NYR)

# ===================================================================== CALC_Impacts
ws = wb.create_sheet('CALC_Impacts')
title(ws, 'Step 4 — apply the supplied multipliers',
      'Direct = initial effect x (flow-on + direct-only vectors). Indirect and induced apply to the flow-on vector only.')
hdr(ws, 4, ['Measure', 'Effect'] + ['FY%d' % (2026 + i) for i in range(NYR)] + ['Total'], [30, 24] + [12] * NYR + [13])
r = 5; imp_rows = {}
for mi, mname, unit in MEAS:
    for label, ename, both in [('Direct', 'Initial', True), ('Indirect (production induced)', 'ProdInduced', False),
                               ('Induced (consumption)', 'ConsInduced', False)]:
        ws.cell(row=r, column=1, value='%s (%s)' % (mname, unit)).font = BLK
        ws.cell(row=r, column=2, value=label).font = BLK
        col = mcol(mname, ename)
        for y in range(NYR):
            mr = 'IFERROR(ENG_SelectedMult!$%s$5:$%s$%d,0)' % (col, col, 4 + N)
            f = '=SUMPRODUCT(%s,%s)' % (VEC(y), mr)
            if both:
                f += '+SUMPRODUCT(%s,%s)' % (VECD(y), mr)
            ws.cell(row=r, column=3 + y, value=f).number_format = NUM
            ws.cell(row=r, column=3 + y).font = BLK
        ws.cell(row=r, column=3 + NYR, value='=SUM($C%d:$%s%d)' % (r, CL(2 + NYR), r)).font = BLK
        ws.cell(row=r, column=3 + NYR).number_format = NUM
        imp_rows[(mname, label)] = r; r += 1
    for label, parts in [('Type I total (direct + indirect)', ['Direct', 'Indirect (production induced)']),
                         ('Total (Type II)', ['Direct', 'Indirect (production induced)', 'Induced (consumption)'])]:
        ws.cell(row=r, column=1, value='%s (%s)' % (mname, unit)).font = BLK
        ws.cell(row=r, column=2, value=label).font = BLK
        for y in range(NYR + 1):
            f = '=' + '+'.join('%s%d' % (CL(3 + y), imp_rows[(mname, p)]) for p in parts)
            ws.cell(row=r, column=3 + y, value=f).number_format = NUM
            ws.cell(row=r, column=3 + y).font = BLK
        imp_rows[(mname, label)] = r; r += 1
    imp_rows[(mname, 'Type I total')] = imp_rows[(mname, 'Type I total (direct + indirect)')]
    r += 1
ws.freeze_panes = 'C5'

# ===================================================================== OUT_Summary
ws = wb.create_sheet('OUT_Summary')
title(ws, 'Results', 'Read QA_Checks before using any number on this page.')
for i, (a, b) in enumerate([('Region', '=Settings!$B$5'), ('Flow tables in use', '=Settings!$B$16'),
                            ('Headline basis', '=Settings!$B$9')]):
    ws.cell(row=4 + i, column=1, value=a).font = BLK
    ws.cell(row=4 + i, column=2, value=b).font = GRN
ws['A7'] = 'QA status'
ws['B7'] = '=IF(QA_Checks!$B$4="ALL OK","ALL OK",QA_Checks!$B$4)'; ws['B7'].font = RED
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 30
ws['A9'] = 'Cumulative impact over the whole profile'; ws['A9'].font = H2
hdr(ws, 10, ['Measure', 'Direct', 'Indirect', 'Type I total', 'Induced', 'Type II total'], [34, 15, 15, 15, 15, 15])
for i, (mi, mname, unit) in enumerate(MEAS):
    r = 11 + i
    ws.cell(row=r, column=1, value='%s (%s)' % (mname, unit)).font = BLK
    tc = CL(3 + NYR)
    for j, key in enumerate(['Direct', 'Indirect (production induced)', 'Type I total',
                             'Induced (consumption)', 'Total (Type II)']):
        ws.cell(row=r, column=2 + j, value='=CALC_Impacts!$%s$%d' % (tc, imp_rows[(mname, key)])).number_format = NUM
        ws.cell(row=r, column=2 + j).font = BLK
r0 = 11 + len(MEAS) + 2
ws.cell(row=r0, column=1, value='Shock reconciliation ($m, cumulative)').font = H2
hdr(ws, r0 + 1, ['Item', 'Value', 'Share of shock'], [40, 16, 16])
recon = [('Shock at purchasers prices', '=SUM(CALC_Strip!$H$5:$H$%d)' % STRIP_LAST),
         ('Less: net taxes on products', '=-SUM(CALC_Strip!$N$5:$N$%d)' % STRIP_LAST),
         ('Less: imports', '=-SUM(CALC_Strip!$M$5:$M$%d)' % STRIP_LAST),
         ('Direct domestic production (basic prices)', '=CALC_Vector!$%s$%d' % (VEC_TOTCOL, 5 + N)),
         ('   of which producer of the product', '=SUM(CALC_Strip!$L$5:$L$%d)' % STRIP_LAST),
         ('   of which margin industries', '=SUM(%s)' % MARG_RANGE)]
for i, (lab, f) in enumerate(recon):
    ws.cell(row=r0 + 2 + i, column=1, value=lab).font = BLK
    ws.cell(row=r0 + 2 + i, column=2, value=f).number_format = NUM
    ws.cell(row=r0 + 2 + i, column=2).font = BLK
    p = ws.cell(row=r0 + 2 + i, column=3, value='=IFERROR($B%d/$B$%d,"")' % (r0 + 2 + i, r0 + 2))
    p.font = BLK; p.number_format = PCT
r1 = r0 + 2 + len(recon) + 2
ws.cell(row=r1, column=1, value='Value added (basic prices) by year').font = H2
hdr(ws, r1 + 1, ['Effect'] + ['FY%d' % (2026 + i) for i in range(NYR)], [30] + [12] * NYR)
for i, key in enumerate(['Direct', 'Indirect (production induced)', 'Type I total',
                         'Induced (consumption)', 'Total (Type II)']):
    ws.cell(row=r1 + 2 + i, column=1, value=key).font = BLK
    for y in range(NYR):
        ws.cell(row=r1 + 2 + i, column=2 + y,
                value='=CALC_Impacts!$%s$%d' % (CL(3 + y), imp_rows[('Value added (basic prices)', key)])
                ).number_format = NUM
        ws.cell(row=r1 + 2 + i, column=2 + y).font = BLK

# ===================================================================== OUT_Detail
ws = wb.create_sheet('OUT_Detail')
title(ws, 'Value added by industry ($m, cumulative)')
hdr(ws, 4, ['IOIG', 'Industry', 'Direct domestic $m', 'Direct VA $m', 'Type I total VA $m', 'Type II total VA $m'],
    [10, 44, 18, 16, 20, 20])
ci, cp, ct = (mcol('Value added (basic prices)', e) for e in ['Initial', 'ProdInduced', 'Total'])
for i, c in enumerate(CODES):
    r = 5 + i
    ws.cell(row=r, column=1, value=c).font = BLK
    ws.cell(row=r, column=2, value=NAMES[c]).font = BLK
    ws.cell(row=r, column=3, value='=CALC_Vector!$%s%d' % (VEC_TOTCOL, r)).number_format = NUM
    ws.cell(row=r, column=3).font = BLK
    ws.cell(row=r, column=4, value='=$C%d*IFERROR(ENG_SelectedMult!$%s%d,0)' % (r, ci, r)).number_format = NUM
    ws.cell(row=r, column=5, value='=$C%d*(IFERROR(ENG_SelectedMult!$%s%d,0)+IFERROR(ENG_SelectedMult!$%s%d,0))'
            % (r, ci, r, cp, r)).number_format = NUM
    ws.cell(row=r, column=6, value='=$C%d*IFERROR(ENG_SelectedMult!$%s%d,0)' % (r, ct, r)).number_format = NUM
    for cc in range(4, 7):
        ws.cell(row=r, column=cc).font = BLK
for j in range(4):
    cell = ws.cell(row=5 + N, column=3 + j, value='=SUM(%s$5:%s$%d)' % (CL(3 + j), CL(3 + j), 4 + N))
    cell.font = Font(name=A_, size=10, bold=True); cell.number_format = NUM
ws.freeze_panes = 'C5'

# ===================================================================== QA_Checks
ws = wb.create_sheet('QA_Checks')
title(ws, 'QA gates — every one must read OK before any result is used')
ws['A4'] = 'OVERALL'; ws['A4'].font = H2
hdr(ws, 6, ['#', 'Check', 'Result', 'Status', 'What a fail means'], [5, 48, 16, 12, 66])
OS = mcol('Output', 'Simple'); OI = mcol('Output', 'Initial'); OP = mcol('Output', 'ProdInduced')
OF = mcol('Output', 'FirstRound'); OIS = mcol('Output', 'IndSupport'); OT = mcol('Output', 'Total')
E = lambda c: 'IFERROR(ENG_SelectedMult!$%s$5:$%s$%d,0)' % (c, c, 4 + N)
AUSCOL = ('IFERROR(INDEX(%s,MATCH("Aus|"&Lists!$B$5:$B$%d,%s,0),%d),0)' % (MULT_DATA, 4 + N, MULT_KEY, EFF['Simple']))
checks = [
    ('Spine has 115 unique codes', '=COUNTA(%s)-SUMPRODUCT((COUNTIF(%s,%s)>1)*1)' % (LIST_CODE, LIST_CODE, LIST_CODE),
     '=IF(C{R}=115,"OK","FAIL")', 'The industry list is corrupted. Nothing downstream is reliable.'),
    ('Multipliers loaded for the selected region', '=COUNT(ENG_SelectedMult!$D$5:$D$%d)' % (4 + N),
     '=IF(C{R}>=114,"OK","FAIL")', 'The region on Settings has no multiplier data. Paste the vendor set in.'),
    ('Selected region is not a copy of Australia',
     '=IF(Settings!$B$5="Aus",-1,SUMPRODUCT(ABS(%s-%s)))' % (E(OS), AUSCOL),
     '=IF(OR(C{R}=-1,C{R}>0.001),"OK","FAIL")',
     'The state multipliers are identical to the national set — a placeholder, not a regionalised table.'),
    ('State multipliers are smaller than national',
     '=IF(Settings!$B$5="Aus",0,SUMPRODUCT((%s>%s+0.0001)*1))' % (E(OS), AUSCOL),
     '=IF(C{R}=0,"OK","REVIEW")',
     'A state multiplier exceeds the national one. Smaller economies leak more, so this points to a data error.'),
    ('Multiplier identity: simple = initial + production induced',
     '=MAX(ABS(%s-%s-%s))' % (E(OS), E(OI), E(OP)), '=IF(C{R}<0.0001,"OK","FAIL")',
     'The supplied multiplier block is not internally consistent — a paste or column-alignment error.'),
    ('Multiplier identity: production induced = first round + industrial support',
     '=MAX(ABS(%s-%s-%s))' % (E(OP), E(OF), E(OIS)), '=IF(C{R}<0.0001,"OK","FAIL")',
     'Same as above — the effect columns do not decompose correctly.'),
    ('Type II is not smaller than Type I', '=SUMPRODUCT((%s<%s-0.0001)*1)' % (E(OT), E(OS)),
     '=IF(C{R}=0,"OK","FAIL")', 'Adding household consumption cannot reduce the impact.'),
    ('Every shock line has a valid product code',
     '=SUMPRODUCT((IN_Shock!$E$%d:$E$%d="Y")*(IFERROR(MATCH(IN_Shock!$C$%d:$C$%d,%s,0),0)=0)*1)'
     % (SHOCK_R0, SHOCK_R0 + NLINE - 1, SHOCK_R0, SHOCK_R0 + NLINE - 1, LIST_CODE),
     '=IF(C{R}=0,"OK","FAIL")', 'An included line has a code that is not in IOIG(2022). It will be silently dropped.'),
    ('Every shock line has a valid column group',
     '=SUMPRODUCT((IN_Shock!$E$%d:$E$%d="Y")*(IFERROR(MATCH(IN_Shock!$D$%d:$D$%d,%s,0),0)=0)*1)'
     % (SHOCK_R0, SHOCK_R0 + NLINE - 1, SHOCK_R0, SHOCK_R0 + NLINE - 1, LIST_GRP),
     '=IF(C{R}=0,"OK","FAIL")', 'An included line has no purchasing column group, so no strip rates apply.'),
    ('Strip identity: components sum back to the purchasers price',
     '=MAX(ABS(CALC_Strip!$%s$5:$%s$%d))' % (CL(12 + len(comp)), CL(12 + len(comp)), STRIP_LAST),
     '=IF(C{R}<0.000001,"OK","FAIL")', 'Domestic + imports + taxes + margins does not equal the spend.'),
    ('Direct vector reconciles to the strip',
     '=ROUND(CALC_Vector!$%s$%d-SUM(CALC_Strip!$L$5:$L$%d)-SUM(%s),6)' % (VEC_TOTCOL, 5 + N, STRIP_LAST, MARG_RANGE),
     '=IF(ABS(C{R})<0.000001,"OK","FAIL")', 'Margin dollars are being lost or double counted.'),
    ('Conversion ratio is plausible',
     '=IFERROR(CALC_Vector!$%s$%d/SUM(CALC_Strip!$H$5:$H$%d),0)' % (VEC_TOTCOL, 5 + N, STRIP_LAST),
     '=IF(AND(C{R}>0.5,C{R}<0.995),"OK","REVIEW")',
     'Outside 50-99.5% usually means a wrong purchasing column group or a mis-coded product. Service and '
     'construction shocks legitimately sit high; goods shocks sit lower.'),
    ('No export-column lines (import share unreliable in v0.2)',
     '=COUNTIF(IN_Shock!$D$%d:$D$%d,"Exports")' % (SHOCK_R0, SHOCK_R0 + NLINE - 1),
     '=IF(C{R}=0,"OK","REVIEW")', 'Table 5 and Table 8 treat re-exports differently, so T8-T5 is zero for exports.'),
    ('6700 Imputed rent is borrowing 6701 multipliers',
     '=IF(INDEX(%s,MATCH("6700",%s,0))="6701",1,0)' % (LIST_SRC, LIST_CODE),
     '=IF(C{R}=0,"OK","REVIEW")', 'PLACEHOLDER bridge in place. Ask the vendor for a 6700 row.'),
    ('No included line falls on an empty ABS cell',
     '=SUMPRODUCT((CALC_Strip!$F$5:$F$%d="Y")*(CALC_Strip!$H$5:$H$%d<>0)*(CALC_Strip!$K$5:$K$%d=0)*1)'
     % (STRIP_LAST, STRIP_LAST, STRIP_LAST), '=IF(C{R}=0,"OK","FAIL")',
     'A line has spend but no ABS purchasers-price value for that product and column group, so it is silently '
     'dropped. Usually a materials line put in a GFCF column — see the guidance on IN_Shock.'),
    ('No negative domestic or import share on an included line',
     '=SUMPRODUCT((CALC_Strip!$F$5:$F$%d="Y")*(CALC_Strip!$H$5:$H$%d<>0)*'
     '((CALC_Strip!$L$5:$L$%d<0)+(CALC_Strip!$M$5:$M$%d<0)>0)*1)'
     % (STRIP_LAST, STRIP_LAST, STRIP_LAST, STRIP_LAST), '=IF(C{R}=0,"OK","REVIEW")',
     'The ABS cell has a negative value (net disposals or an import adjustment). The rate is not meaningful.'),
    ('Margin control total: CALC_StripData equals ABS Tables 23-34',
     '=ROUND(%s-%s,0)' % (STRIPDATA_CTRL, MARG_CTRL), '=IF(C{R}=0,"OK","FAIL")',
     'The aggregation of the ABS margin tabs has lost or duplicated cells. Check the RAW tabs were pasted whole.'),
    ('Flow tables in use are real, not the state placeholder',
     '=IF(Settings!$B$5="Aus",0,1)', '=IF(C{R}=0,"OK","PLACEHOLDER")',
     'RAW_T5_State and RAW_T8_State hold NSW 2021-22 data standing in for every state. Import shares, and therefore '
     'the conversion ratio, are not that state\'s. Load the real nine-region set.'),
    ('Shock price year matches multiplier price year', '=Settings!$B$10-Settings!$B$11',
     '=IF(C{R}=0,"OK","REVIEW")', 'Mixed vintages. Deflate the shock on Settings B12 or get a matching set.'),
    ('Margins and taxes are national even for a state run', '=IF(Settings!$B$5="Aus",0,1)',
     '=IF(C{R}=0,"OK","NOTE")', 'The ABS publishes no state margin or tax matrices. State runs use national margin '
     'and tax rates with state import shares. Disclose this in the report.'),
]
for i, (lab, res, stat, meaning) in enumerate(checks):
    r = 7 + i
    ws.cell(row=r, column=1, value=i + 1).font = BLK
    ws.cell(row=r, column=2, value=lab).font = BLK
    c = ws.cell(row=r, column=3, value=res); c.font = BLK; c.number_format = '0.000000'
    s = ws.cell(row=r, column=4, value=stat.replace('{R}', str(r)))
    s.font = Font(name=A_, size=10, bold=True); s.alignment = Alignment(horizontal='center')
    m = ws.cell(row=r, column=5, value=meaning); m.font = NOTE; m.alignment = Alignment(wrap_text=True)
LQ = 6 + len(checks)
ws['B4'] = ('=IF(COUNTIF($D$7:$D$%d,"FAIL")>0,"FAIL - DO NOT USE RESULTS",'
            'IF(COUNTIF($D$7:$D$%d,"PLACEHOLDER")>0,"RUNNING ON PLACEHOLDER DATA",'
            'IF(COUNTIF($D$7:$D$%d,"REVIEW")>0,"REVIEW ITEMS","ALL OK")))' % (LQ, LQ, LQ))
ws['B4'].font = Font(name=A_, size=12, bold=True, color='C00000')
ws['C4'] = ('=COUNTIF($D$7:$D$%d,"FAIL")&" fail, "&COUNTIF($D$7:$D$%d,"REVIEW")&" review, "'
            '&COUNTIF($D$7:$D$%d,"PLACEHOLDER")&" placeholder"' % (LQ, LQ, LQ))
ws['C4'].font = BLK

# ===================================================================== Assumptions
ws = wb.create_sheet('Assumptions')
title(ws, 'Assumptions register — attach this to every report')
hdr(ws, 4, ['Item', 'Value / setting', 'Source or basis'], [40, 40, 66])
asum = [('Region modelled', '=Settings!$B$5', 'Selected by the analyst'),
        ('Flow tables used', '=Settings!$B$16', 'Aus is real 2023-24; any state is the NSW 2021-22 placeholder'),
        ('Years modelled', '=Settings!$B$6&" to "&(Settings!$B$6+Settings!$B$7-1)', 'Settings'),
        ('Study type', '=Settings!$B$8', 'Settings'),
        ('Headline multiplier basis', '=Settings!$B$9', 'Type I is the recommended headline'),
        ('Shock price year', '=Settings!$B$10', 'Settings'),
        ('Multiplier price year', '=Settings!$B$11', 'Vintage of the supplied multiplier set'),
        ('Deflator applied', '=Settings!$B$12', 'Manual'),
        ('Total shock, purchasers prices ($m)', '=SUM(CALC_Strip!$H$5:$H$%d)' % STRIP_LAST, 'IN_Shock'),
        ('Direct domestic production ($m)', '=CALC_Vector!$%s$%d' % (VEC_TOTCOL, 5 + N), 'CALC_Vector'),
        ('Conversion ratio', '=IFERROR(B14/B13,"")', 'Direct domestic / shock'),
        ('Multiplier source', 'Supplied set — not derived by this model', 'RAW_Multipliers'),
        ('Flow table source', 'ABS 5209.0.55.001, 2023-24, Tables 5 and 8', 'RAW_T5_Aus / RAW_T8_Aus'),
        ('Margin source', 'ABS Tables 23-34, 2023-24, national', 'RAW_T23, RAW_T24, RAW_T25_34_Margins'),
        ('Tax source', 'ABS Table 35, 2023-24, national', 'RAW_T35_NetTaxes'),
        ('Import derivation', 'Table 8 less Table 5, cell by cell', 'CALC_StripData'),
        ('Margin treatment', 'SNA1968 basis — changes in the 2024-25 ABS release, due March 2027',
         'ABS Input-Output Tables methodology 2023-24'),
        ('Type II closure', 'As supplied in the multiplier set — not recomputed here', 'Vendor'),
        ('PLACEHOLDERS IN USE', '=IF(Settings!$B$5="Aus","6700 bridged to 6701 only",'
                                '"State flow tables (NSW 2021-22) AND 6700 bridged to 6701")', 'See README')]
for i, (a, b, c) in enumerate(asum):
    ws.cell(row=5 + i, column=1, value=a).font = BLK
    cell = ws.cell(row=5 + i, column=2, value=b)
    cell.font = GRN if str(b).startswith('=') else BLK
    cell.alignment = Alignment(wrap_text=True)
    ws.cell(row=5 + i, column=3, value=c).font = NOTE

for s in wb.worksheets:
    s.sheet_view.showGridLines = False
OUTFILE.parent.mkdir(exist_ok=True)
wb.save(OUTFILE)
print('wrote %s  (strip rows=%d, qa gates=%d)' % (OUTFILE, STRIP_LAST, LQ))
