"""
Multi-region variant: every shock line carries its own region.

    python scripts/build_model_multiregion.py

v0.3 (scripts/build_model.py) is the single-region model and remains the main
version. This one exists because a study often spans jurisdictions - a project
with a NSW construction component, a Victorian manufacturing component, and a
national overlay - and running it three times loses the combined view.

The difference is where region lives. In v0.3 it is one cell on Settings and
every lookup in the workbook is bound to it. Here it is a column on IN_Shock,
so each line is a (region, product, column group) triple resolved on its own.

That inverts the sizing, which is the real reason this design works. v0.3
resolves all 115 products x 8 groups for one region whether or not you use
them: 49,220 formulas across MAP_StripData, CALC_Rates and MAP_Multipliers.
Doing that for nine regions would be roughly 440,000. Demand-driven, only the
cells actually shocked are resolved - 40 lines is 40 strips, not 1,035 - so
this model is SMALLER than the single-region one while covering nine regions.

What gets pulled, per line:
  domestic   that line's REGION Table 5
  imports    that line's REGION Table 8, less Table 5
  net taxes  ABS Table 35        - national, the ABS publishes no state matrix
  margins    ABS Tables 23-34    - national, same reason
and the margin is then earned by the mapped industry IN THE SAME REGION, which
is an assumption worth reading in Assumptions.

Summing regional results is NOT the national impact of the combined shock.
State multipliers are smaller because a state leaks to the rest of Australia,
so a sum across states understates what the same spend does nationally. Both
are shown, and never added together.
"""
import os
import pickle
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation

import build_raw_stacked as RS

ROOT = Path(__file__).resolve().parent.parent
SOURCES = ROOT / 'build' / 'sources.pkl'
OUTFILE = ROOT / 'output' / 'IO_Impact_Model_MultiRegion_v0-4.xlsx'
# A study spanning nine jurisdictions needs a line per (region, IOIG code), so
# IOMODEL_LINES builds a taller IN_Shock. 32 household categories concorded to
# ~50 IOIG codes across 9 jurisdictions is ~450 lines.
SUBSET_OUT = ROOT / 'build' / 'IO_MultiRegion_subset.xlsx'

H1 = Font(bold=True, size=14, color='FFFFFF')
H2 = Font(bold=True, size=11)
H3 = Font(bold=True, size=10)
BLUE = Font(color='0000CC')
GREEN = Font(color='006100')
NOTE = Font(italic=True, size=9, color='666666')
TITLEFILL = PatternFill('solid', fgColor='1F3864')
HDRFILL = PatternFill('solid', fgColor='DDEBF7')
GREYFILL = PatternFill('solid', fgColor='F2F2F2')
YELLOW = PatternFill('solid', fgColor='FFF2CC')
MONEY = '#,##0.0'
PCT = '0.0%'
NUM4 = '0.0000'

REGIONS = RS.REGIONS
GROUPS = [
    ('Intermediate', 'Businesses buying inputs. A bill of quantities belongs here',
     RS.FLOW_IND_FIRST, RS.FLOW_IND_LAST, RS.MARG_IND_FIRST, RS.MARG_IND_LAST),
    ('Q1_HFCE', 'Households; final consumption expenditure', 128, 128, 119, 119),
    ('Q2_GGFCE', 'General government; final consumption expenditure', 129, 129, 120, 120),
    ('Q3_GFCF_Priv', 'Private; gross fixed capital formation', 130, 130, 121, 121),
    ('Q4_GFCF_PubCorp', 'Public corporations; GFCF', 131, 131, 122, 122),
    ('Q5_GFCF_GG', 'General government; GFCF', 132, 132, 123, 123),
    ('Q6_Inventories', 'Changes in inventories', 133, 133, 124, 124),
    ('Q7_Exports', 'Exports of goods and services', 134, 134, 125, 125),
]
NG = len(GROUPS)
MARGIN_ORDER = ['Wholesale', 'Retail', 'RestHotelClub', 'Road', 'Rail', 'Pipeline',
                'Water', 'Air', 'PortHandling', 'MarineIns', 'Gas', 'Electricity']
MEASURES = ['Domestic', 'Imports', 'NetTaxes'] + MARGIN_ORDER
NM = len(MEASURES)
NYR = int(os.environ.get('IOMODEL_YEARS', '8'))
NLINE = int(os.environ.get('IOMODEL_LINES', '40'))
REPORT = [('Output multipliers', 'Output, $m', MONEY),
          ('Income multipliers', 'Wages and salaries, $m', MONEY),
          ('Value added multipliers', 'Value added at basic prices, $m', MONEY),
          ('Employed multipliers', 'Employment, FTE', '#,##0'),
          ('Gross operating surplus & mixed income multipliers', 'GOS and mixed income, $m',
           MONEY)]
EXAMPLE = [
    ('NSW', 'New road, NSW', '3101', 'Q5_GFCF_GG', 'N', 100.0),
    ('Vic', 'Household food, Vic', '1101', 'Q1_HFCE', 'N', 20.0),
    ('QLD', 'Engineering design, Qld', '6901', 'Intermediate', 'N', 15.0),
    ('Aus', 'National program overlay', '3101', 'Q5_GFCF_GG', 'N', 50.0),
]


def band(ws, row, text, width):
    ws.cell(row=row, column=1, value=text).font = H1
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = TITLEFILL


def hdr(ws, row, labels, col0=1):
    for i, h in enumerate(labels, col0):
        c = ws.cell(row=row, column=i, value=h)
        c.font = H3
        c.fill = HDRFILL
        c.alignment = Alignment(wrap_text=True, vertical='bottom')


def main():
    src = pickle.load(open(SOURCES, 'rb'))
    abs_t, mult = src['abs'], src['multipliers']
    spine = [(m['code'], m['label']) for m in abs_t['T23']['meta']
             if m['row_type'] == 'Product']
    subset = bool(os.environ.get('IOMODEL_SUBSET'))
    if subset:
        keep = {e[2] for e in EXAMPLE} | {v[1] for v in src['margin_tables'].values()}
        keep |= {'6700', '6701'}
        spine = [x for x in spine if x[0] in keep]
        print(f'  SUBSET BUILD: {len(spine)} spine rows')
    NS = len(spine)
    mult_codes = {m['code'] for m in mult['regions']['Aus']['meta']
                  if m['row_type'] == 'Industry'}
    blocks, effects = mult['blocks'], mult['effects']
    NE = len(effects)
    i_init = effects.index('Initial effect')
    i_prod = effects.index('Production-induced effect')
    i_cons = effects.index('Consumption-induced effect')
    i_totl = effects.index('Total multiplier')
    i_simp = effects.index('Simple multiplier')
    EFF_OUT = [('Direct', i_init), ('Indirect (production-induced)', i_prod),
               ('Induced (consumption-induced)', i_cons), ('TOTAL', i_totl)]
    mt = src['margin_tables']
    mcols = MARGIN_ORDER + ['NetTaxes']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ======================================================== README
    ws = wb.create_sheet('README')
    band(ws, 1, 'Australian IO impact model - MULTI-REGION variant, v0.4', 8)
    for i, (a, b) in enumerate([
        ('What is different', 'Region is a column on IN_Shock, not a single setting. Every '
         'line is resolved as its own (region, product, column group).'),
        ('Why it is smaller', 'It resolves only the cells you actually shock. The '
         'single-region model resolves all 115 products x 8 groups whether used or not.'),
        ('', ''),
        ('THE CHAIN', ''),
        ('  RAW_*', 'All nine regions, verbatim, stacked and keyed. Unchanged from v0.3'),
        ('  MAP_ShockKeys', 'Per line: resolve region+product+group to row and column positions'),
        ('  MAP_LineStrip', 'Per line: the RAW values for that one cell, and the shares'),
        ('  CALC_Strip', 'Per line x measure x year, $m'),
        ('  CALC_Margins', 'Margin earned, by region and margin type'),
        ('  CALC_Vector', 'Direct domestic by (region, industry), with its multipliers'),
        ('  CALC_Impacts', 'By region, measure and effect'),
        ('  OUT_Summary', 'By region, and the cross-region total'),
        ('', ''),
        ('READ THIS', 'Summing regions is NOT the national impact of the combined shock. '
         'A state leaks to the rest of Australia, so state multipliers are smaller and a '
         'sum across states understates the national effect. Run the shock against Aus if '
         'you want the national number. Never add the two.'),
        ('Margins', 'Margin rates are national - the ABS publishes no state margin or tax '
         'matrices - and the margin is assumed earned in the same region as the purchase.'),
    ], 3):
        ws.cell(row=i, column=1, value=a).font = H3 if a and not a.startswith(' ') else Font()
        ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 96

    # ======================================================== Settings
    ws = wb.create_sheet('Settings')
    band(ws, 1, 'Settings', 6)
    hdr(ws, 4, ['Setting', 'Value', 'Note'])
    for i, (k, v, n) in enumerate([
        ('Study type', 'Impact', 'Impact or Contribution'),
        ('Headline GVA basis', 'Value added', 'Basic prices, the ABS headline. Not market prices'),
        ('Shock price year', '2022-23', 'The price year your spending lines are in'),
        ('Multiplier vintage', '2022-23', 'From the supplied set'),
        ('ABS strip vintage', '2023-24', 'From the ABS margin and tax tables'),
    ], 5):
        ws.cell(row=i, column=1, value=k).font = H3
        c = ws.cell(row=i, column=2, value=v)
        c.font = BLUE
        c.fill = YELLOW
        ws.cell(row=i, column=3, value=n).font = NOTE
    ws.cell(row=11, column=1, value='Region is NOT set here. It is a column on IN_Shock, '
                                    'one per line.').font = Font(bold=True, color='9C0006')
    for col, w in zip('ABC', [22, 16, 76]):
        ws.column_dimensions[col].width = w

    # ======================================================== Lists
    ws = wb.create_sheet('Lists')
    band(ws, 1, 'Lists - the 115-code spine, the 114-code bridge, column groups, regions', 12)
    hdr(ws, 4, ['#', 'IOIG code', 'Industry name', 'Code for the supplied set', 'Bridged?'])
    L0 = 5
    for i, (code, name) in enumerate(spine):
        r = L0 + i
        ws.cell(row=r, column=1, value=i + 1)
        c = ws.cell(row=r, column=2, value=code)
        c.number_format = '@'
        c.font = BLUE
        ws.cell(row=r, column=3, value=name)
        use = code if code in mult_codes else '6701'
        c = ws.cell(row=r, column=4, value=use)
        c.number_format = '@'
        c.font = BLUE
        f = ws.cell(row=r, column=5, value=(1 if use != code else 0))
        if use != code:
            f.fill = YELLOW
    L1 = L0 + NS - 1
    G0 = 5
    hdr(ws, 4, ['Group', 'What it is', 'Flow first', 'Flow last', 'Margin first',
                'Margin last'], col0=7)
    for i, g in enumerate(GROUPS):
        for j, v in enumerate(g, 7):
            ws.cell(row=G0 + i, column=j, value=v)
    G1 = G0 + NG - 1
    hdr(ws, 4, ['Region'], col0=14)
    for i, rg in enumerate(REGIONS):
        ws.cell(row=5 + i, column=14, value=rg)
    RG0, RG1 = 5, 5 + len(REGIONS) - 1
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'G', 'H', 'N'],
                      [5, 10, 44, 13, 9, 17, 40, 9]):
        ws.column_dimensions[col].width = w
    GRPNAMES = f'Lists!$G${G0}:$G${G1}'
    CODES = f'Lists!$B${L0}:$B${L1}'
    USECODES = f'Lists!$D${L0}:$D${L1}'
    REGLIST = f'Lists!$N${RG0}:$N${RG1}'

    # ======================================================== RAW
    _sub = {c for c, _ in spine} if subset else None
    geo = RS.write_raw_tabs(wb, src, verbose=False, subset=_sub)
    mm = RS.write_margin_map(wb, src)

    def rng(tab, first_col):
        g = geo[tab]
        return (f"{tab}!${CL(first_col)}${g['first']}:"
                f"${CL(first_col + g['width'] - 1)}${g['last']}")

    def keyrng(tab, keycol):
        g = geo[tab]
        return f"{tab}!${CL(keycol)}${g['first']}:${CL(keycol)}${g['last']}"

    T5R, T8R = rng('RAW_T5', RS.FLOW_VERB_COL), rng('RAW_T8', RS.FLOW_VERB_COL)
    MUR, MGR = rng('RAW_Multipliers', RS.FLOW_VERB_COL), rng('RAW_Margins', RS.MARG_VERB_COL)
    T5K, T8K = keyrng('RAW_T5', 5), keyrng('RAW_T8', 5)
    MUK, MGK = keyrng('RAW_Multipliers', 5), keyrng('RAW_Margins', 6)

    # ======================================================== IN_Shock
    ws = wb.create_sheet('IN_Shock')
    band(ws, 1, 'IN_Shock - YOUR INPUT. One line per item of spending, $m at purchasers '
                'prices. EACH LINE CARRIES ITS OWN REGION', 15)
    ws.cell(row=2, column=1, value=(
        'Region is per line, so a study can span jurisdictions in one run. Code is the IOIG '
        'of the PRODUCT bought. Group is who is buying - a bill of quantities goes to '
        'Intermediate, because the contractor is a business buying inputs.')).font = NOTE
    ws.cell(row=3, column=1, value=(
        'Direct only = Y for a line whose supply chain you have already itemised. It is '
        'counted as direct but never multiplied.')).font = NOTE
    hdr(ws, 5, ['#', 'Region', 'Description', 'IOIG code', 'Column group', 'Direct only']
        + [f'Year {y + 1}' for y in range(NYR)])
    I0 = 6
    for i in range(NLINE):
        r = I0 + i
        ws.cell(row=r, column=1, value=i + 1)
        for j in range(2, 7 + NYR):
            c = ws.cell(row=r, column=j)
            c.fill = YELLOW
            c.font = BLUE
            if j >= 7:
                c.number_format = MONEY
        ws.cell(row=r, column=4).number_format = '@'
        if i < len(EXAMPLE):
            rg, d, code, grp, only, amt = EXAMPLE[i]
            for j, v in zip(range(2, 8), [rg, d, code, grp, only, amt]):
                ws.cell(row=r, column=j, value=v)
        else:
            ws.cell(row=r, column=6, value='N')
    I1 = I0 + NLINE - 1
    for f1, rf in ((REGLIST, f'B{I0}:B{I1}'), (GRPNAMES, f'E{I0}:E{I1}'),
                   ('"Y,N"', f'F{I0}:F{I1}')):
        dv = DataValidation(type='list', formula1=f1)
        ws.add_data_validation(dv)
        dv.add(rf)
    TR = I1 + 2
    ws.cell(row=TR, column=3, value='TOTAL').font = H2
    for y in range(NYR):
        ws.cell(row=TR, column=7 + y, value=(
            f'=SUM({CL(7 + y)}{I0}:{CL(7 + y)}{I1})')).number_format = MONEY
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F'], [4, 9, 32, 11, 18, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'G6'

    # ======================================================== MAP_ShockKeys
    ws = wb.create_sheet('MAP_ShockKeys')
    band(ws, 1, 'MAP_ShockKeys - resolve each line to positions in the RAW blocks. '
                'This is the whole demand-driven idea: only shocked cells are looked up', 26)
    ws.cell(row=2, column=1, value=(
        'One MATCH per line per block, instead of one per product per block. A line with no '
        'code resolves to 0 and contributes nothing.')).font = NOTE
    hdr(ws, 4, ['Line', 'Region', 'Code', 'Use code', 'Group', 'Direct only',
                'Flow first', 'Flow last', 'Marg first', 'Marg last', 'T5 row', 'T8 row']
        + [m[:9] + ' r' for m in mcols])
    K0 = 5
    for i in range(NLINE):
        r = K0 + i
        ir = I0 + i
        ws.cell(row=r, column=1, value=f'=IN_Shock!$A{ir}').font = GREEN
        ws.cell(row=r, column=2, value=f'=IN_Shock!$B{ir}').font = GREEN
        c = ws.cell(row=r, column=3, value=f'=IN_Shock!$D{ir}')
        c.number_format = '@'
        c.font = GREEN
        c = ws.cell(row=r, column=4, value=(
            f'=IFERROR(INDEX({USECODES},MATCH($C{r},{CODES},0)),"")'))
        c.number_format = '@'
        ws.cell(row=r, column=5, value=f'=IN_Shock!$E{ir}').font = GREEN
        ws.cell(row=r, column=6, value=f'=IN_Shock!$F{ir}').font = GREEN
        for j, srccol in enumerate(['I', 'J', 'K', 'L']):
            ws.cell(row=r, column=7 + j, value=(
                f'=IFERROR(INDEX(Lists!${srccol}${G0}:${srccol}${G1},'
                f'MATCH($E{r},{GRPNAMES},0)),0)'))
        ws.cell(row=r, column=11, value=(
            f'=IFERROR(MATCH($B{r}&"|"&$D{r},{T5K},0),0)'))
        ws.cell(row=r, column=12, value=(
            f'=IFERROR(MATCH($B{r}&"|"&$D{r},{T8K},0),0)'))
        for j, mn in enumerate(mcols, 13):
            ws.cell(row=r, column=j, value=(
                f'=IFERROR(MATCH("{mn}|"&$C{r},{MGK},0),0)'))
    K1 = K0 + NLINE - 1
    ws.freeze_panes = 'G5'

    # ======================================================== MAP_LineStrip
    ws = wb.create_sheet('MAP_LineStrip')
    band(ws, 1, 'MAP_LineStrip - the RAW values for each shocked cell, and the shares', 40)
    ws.cell(row=2, column=1, value=(
        'Domestic and imports come from the LINE REGION Table 5 and Table 8. Net taxes and '
        'the 12 margins are national ABS data - the ABS publishes no state matrices - and '
        'the margin is assumed earned in the same region as the purchase.')).font = NOTE
    hdr(ws, 5, ['Line', 'Region', 'Code', 'Group'])
    c = ws.cell(row=4, column=5, value='Component values, $m')
    c.font = H3
    c.fill = GREYFILL
    for mi, meas in enumerate(MEASURES):
        h = ws.cell(row=5, column=5 + mi, value=meas)
        h.font = H3
        h.fill = HDRFILL
        h.alignment = Alignment(wrap_text=True)
    PPC = 5 + NM
    ws.cell(row=5, column=PPC, value='PP total').font = H3
    SH0 = PPC + 1
    c = ws.cell(row=4, column=SH0, value='Shares')
    c.font = H3
    c.fill = GREYFILL
    for mi, meas in enumerate(MEASURES):
        h = ws.cell(row=5, column=SH0 + mi, value=meas)
        h.font = H3
        h.fill = HDRFILL
        h.alignment = Alignment(wrap_text=True)
    CHKC = SH0 + NM
    ws.cell(row=5, column=CHKC, value='Check = 100%').font = H3
    S0 = 6
    for i in range(NLINE):
        r = S0 + i
        kr = K0 + i
        for j, srccol in enumerate(['A', 'B', 'C', 'E'], 1):
            c = ws.cell(row=r, column=j, value=f'=MAP_ShockKeys!${srccol}{kr}')
            c.font = GREEN
            if j == 3:
                c.number_format = '@'
        t5 = (f'SUM(INDEX({T5R},MAP_ShockKeys!$K{kr},MAP_ShockKeys!$G{kr}):'
              f'INDEX({T5R},MAP_ShockKeys!$K{kr},MAP_ShockKeys!$H{kr}))')
        t8 = (f'SUM(INDEX({T8R},MAP_ShockKeys!$L{kr},MAP_ShockKeys!$G{kr}):'
              f'INDEX({T8R},MAP_ShockKeys!$L{kr},MAP_ShockKeys!$H{kr}))')
        ws.cell(row=r, column=5, value=(
            f'=IF(MAP_ShockKeys!$K{kr}=0,0,{t5})')).number_format = MONEY
        ws.cell(row=r, column=6, value=(
            f'=IF(OR(MAP_ShockKeys!$K{kr}=0,MAP_ShockKeys!$L{kr}=0),0,({t8})-({t5}))')
        ).number_format = MONEY
        for mi, meas in enumerate(MEASURES[2:], 2):
            mrow = f'MAP_ShockKeys!${CL(13 + mcols.index(meas))}{kr}'
            mg = (f'SUM(INDEX({MGR},{mrow},MAP_ShockKeys!$I{kr}):'
                  f'INDEX({MGR},{mrow},MAP_ShockKeys!$J{kr}))')
            ws.cell(row=r, column=5 + mi, value=(
                f'=IF({mrow}=0,0,{mg})')).number_format = MONEY
        ws.cell(row=r, column=PPC, value=(
            f'=SUM({CL(5)}{r}:{CL(5 + NM - 1)}{r})')).number_format = MONEY
        for mi in range(NM):
            ws.cell(row=r, column=SH0 + mi, value=(
                f'=IFERROR({CL(5 + mi)}{r}/${CL(PPC)}{r},0)')).number_format = PCT
        ws.cell(row=r, column=CHKC, value=(
            f'=IF(${CL(PPC)}{r}=0,1,SUM({CL(SH0)}{r}:{CL(SH0 + NM - 1)}{r}))')
        ).number_format = PCT
    S1 = S0 + NLINE - 1
    ws.freeze_panes = 'E6'

    # ======================================================== CALC_Strip
    ws = wb.create_sheet('CALC_Strip')
    band(ws, 1, 'CALC_Strip - line x measure x year, $m', 14)
    hdr(ws, 5, ['Line', 'Region', 'Code', 'Measure', 'Direct only']
        + [f'Year {y + 1}' for y in range(NYR)])
    P0 = 6
    for li in range(NLINE):
        for mi, meas in enumerate(MEASURES):
            r = P0 + li * NM + mi
            sr = S0 + li
            ir = I0 + li
            ws.cell(row=r, column=1, value=f'=MAP_LineStrip!$A{sr}').font = GREEN
            ws.cell(row=r, column=2, value=f'=MAP_LineStrip!$B{sr}').font = GREEN
            c = ws.cell(row=r, column=3, value=f'=MAP_LineStrip!$C{sr}')
            c.number_format = '@'
            c.font = GREEN
            ws.cell(row=r, column=4, value=meas)
            ws.cell(row=r, column=5, value=f'=MAP_ShockKeys!$F{K0 + li}').font = GREEN
            for y in range(NYR):
                ws.cell(row=r, column=6 + y, value=(
                    f'=IN_Shock!{CL(7 + y)}{ir}*MAP_LineStrip!{CL(SH0 + mi)}{sr}')
                ).number_format = MONEY
    P1 = P0 + NLINE * NM - 1
    ws.freeze_panes = 'F6'

    # ======================================================== CALC_Margins
    ws = wb.create_sheet('CALC_Margins')
    band(ws, 1, 'CALC_Margins - margin earned, by region and margin type, $m', 13)
    ws.cell(row=2, column=1, value=(
        'A margin is revenue of the wholesaler, retailer or carrier, not of the producer, so '
        'it is reallocated to that industry before the multipliers are applied. It is '
        'assumed earned in the region where the purchase happened.')).font = NOTE
    hdr(ws, 5, ['Region', 'Margin', 'Earning IOIG'] + [f'Year {y + 1}' for y in range(NYR)])
    MG0 = 6
    r = MG0
    for rg in REGIONS:
        for mn in MARGIN_ORDER:
            ws.cell(row=r, column=1, value=rg)
            ws.cell(row=r, column=2, value=mn)
            c = ws.cell(row=r, column=3, value=(
                f'=INDEX(Lists_MarginMap!$C${mm["first"]}:$C${mm["last"]},'
                f'MATCH($B{r},Lists_MarginMap!$B${mm["first"]}:$B${mm["last"]},0))'))
            c.number_format = '@'
            for y in range(NYR):
                ws.cell(row=r, column=4 + y, value=(
                    f'=SUMIFS(CALC_Strip!{CL(6 + y)}${P0}:{CL(6 + y)}${P1},'
                    f'CALC_Strip!$B${P0}:$B${P1},$A{r},'
                    f'CALC_Strip!$D${P0}:$D${P1},$B{r},'
                    f'CALC_Strip!$E${P0}:$E${P1},"N")')).number_format = MONEY
            r += 1
    MG1 = r - 1

    # ======================================================== CALC_Vector
    ws = wb.create_sheet('CALC_Vector')
    band(ws, 1, 'CALC_Vector - direct domestic by (region, industry), with that regions '
                'multipliers alongside', 30)
    ws.cell(row=2, column=1, value=(
        'One row per shock line for the domestic content, plus one row per region and margin '
        'type for the reallocated margin. The multiplier columns are looked up on '
        'region + code, so each row carries its own regions multipliers.')).font = NOTE
    hdr(ws, 5, ['Source', 'Region', 'Code'] + [f'Year {y + 1}' for y in range(NYR)])
    mcolstart = 4 + NYR
    ci = mcolstart
    mult_col = {}
    for bn, label, fmt in REPORT:
        for en, ei in EFF_OUT:
            h = ws.cell(row=5, column=ci, value=f'{label.split(",")[0]} {en}')
            h.font = H3
            h.fill = GREYFILL
            h.alignment = Alignment(wrap_text=True)
            mult_col[(bn, en)] = ci
            ci += 1
    V0 = 6
    r = V0
    vec_rows = []
    for li in range(NLINE):
        sr = S0 + li
        ws.cell(row=r, column=1, value='line')
        ws.cell(row=r, column=2, value=f'=MAP_LineStrip!$B{sr}').font = GREEN
        c = ws.cell(row=r, column=3, value=f'=MAP_ShockKeys!$D{K0 + li}')
        c.number_format = '@'
        c.font = GREEN
        for y in range(NYR):
            ws.cell(row=r, column=4 + y, value=(
                f'=IF(MAP_ShockKeys!$F{K0 + li}="Y",0,'
                f'IN_Shock!{CL(7 + y)}{I0 + li}*MAP_LineStrip!{CL(SH0)}{sr})')
            ).number_format = MONEY
        vec_rows.append(r)
        r += 1
    for i in range(MG1 - MG0 + 1):
        mr = MG0 + i
        ws.cell(row=r, column=1, value='margin')
        ws.cell(row=r, column=2, value=f'=CALC_Margins!$A{mr}').font = GREEN
        c = ws.cell(row=r, column=3, value=f'=CALC_Margins!$C{mr}')
        c.number_format = '@'
        c.font = GREEN
        for y in range(NYR):
            ws.cell(row=r, column=4 + y, value=(
                f'=CALC_Margins!{CL(4 + y)}{mr}')).number_format = MONEY
        vec_rows.append(r)
        r += 1
    V1 = r - 1
    for rr in range(V0, V1 + 1):
        for (bn, en), col in mult_col.items():
            ei = dict(EFF_OUT)[en]
            sc = next(b['first_col'] for b in blocks if b['name'] == bn) + ei
            ws.cell(row=rr, column=col, value=(
                f'=IFERROR(INDEX({MUR},MATCH($B{rr}&"|"&$C{rr},{MUK},0),{sc}),0)')
            ).number_format = NUM4
    ws.freeze_panes = 'D6'
    for col, w in zip(['A', 'B', 'C'], [9, 9, 10]):
        ws.column_dimensions[col].width = w

    # ======================================================== CALC_Impacts
    ws = wb.create_sheet('CALC_Impacts')
    band(ws, 1, 'CALC_Impacts - by region, measure and effect', 13)
    ws.cell(row=2, column=1, value=(
        'SUMPRODUCT over the vector rows belonging to each region. A region with no shock '
        'lines simply totals zero.')).font = NOTE
    hdr(ws, 5, ['Region', 'Measure', 'Effect'] + [f'Year {y + 1}' for y in range(NYR)]
        + ['Total'])
    A0 = 6
    r = A0
    imp = {}
    for rg in REGIONS + ['ALL REGIONS']:
        for bn, label, fmt in REPORT:
            for en, ei in EFF_OUT:
                ws.cell(row=r, column=1, value=rg)
                ws.cell(row=r, column=2, value=label)
                ws.cell(row=r, column=3, value=en)
                mc = mult_col[(bn, en)]
                for y in range(NYR):
                    if rg == 'ALL REGIONS':
                        f = (f'=SUMPRODUCT(CALC_Vector!{CL(4 + y)}${V0}:{CL(4 + y)}${V1},'
                             f'CALC_Vector!${CL(mc)}${V0}:${CL(mc)}${V1})')
                    else:
                        f = (f'=SUMPRODUCT((CALC_Vector!$B${V0}:$B${V1}="{rg}")*'
                             f'CALC_Vector!{CL(4 + y)}${V0}:{CL(4 + y)}${V1}*'
                             f'CALC_Vector!${CL(mc)}${V0}:${CL(mc)}${V1})')
                    ws.cell(row=r, column=4 + y, value=f).number_format = fmt
                ws.cell(row=r, column=4 + NYR, value=(
                    f'=SUM({CL(4)}{r}:{CL(3 + NYR)}{r})')).number_format = fmt
                imp[(rg, bn, en)] = r
                r += 1
    A1 = r - 1
    for col, w in zip(['A', 'B', 'C'], [12, 28, 26]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'D6'

    # ======================================================== OUT_Summary
    ws = wb.create_sheet('OUT_Summary')
    band(ws, 1, 'OUT_Summary - results by region', 13)
    ws.cell(row=2, column=1, value='=IF(QA_Checks!$B$4="OK","",'
                                   '"CHECK QA_Checks BEFORE USING THESE NUMBERS")'
            ).font = Font(bold=True, color='9C0006')
    ws.cell(row=3, column=1, value=(
        'ALL REGIONS is the sum of the regional results, each computed with its own regions '
        'multipliers. It is NOT the national impact of the combined shock - a state leaks to '
        'the rest of Australia, so this understates what the same spend does nationally. For '
        'the national number, put the lines against Aus. Never add the two together.')
            ).font = Font(italic=True, size=9, color='9C0006')
    hdr(ws, 5, ['Region', 'Measure', 'Effect'] + [f'Year {y + 1}' for y in range(NYR)]
        + ['Total'])
    O0 = 6
    r = O0
    for rg in REGIONS + ['ALL REGIONS']:
        for bn, label, fmt in REPORT:
            for en, ei in EFF_OUT:
                sr = imp[(rg, bn, en)]
                ws.cell(row=r, column=1, value=rg)
                ws.cell(row=r, column=2, value=label)
                b = ws.cell(row=r, column=3, value=en)
                for y in range(NYR + 1):
                    c = ws.cell(row=r, column=4 + y, value=f'=CALC_Impacts!{CL(4 + y)}{sr}')
                    c.number_format = fmt
                    c.font = Font(bold=True, color='006100') if en == 'TOTAL' else GREEN
                if en == 'TOTAL':
                    b.font = H3
                r += 1
    O1 = r - 1
    ws.cell(row=r + 1, column=1, value='Spend by region, year 1').font = H2
    for i, rg in enumerate(REGIONS):
        ws.cell(row=r + 2 + i, column=1, value=rg)
        ws.cell(row=r + 2 + i, column=4, value=(
            f'=SUMIFS(IN_Shock!$G${I0}:$G${I1},IN_Shock!$B${I0}:$B${I1},$A{r + 2 + i})')
        ).number_format = MONEY
        ws.cell(row=r + 2 + i, column=5, value=(
            f'=IFERROR(SUMIFS(CALC_Vector!$D${V0}:$D${V1},'
            f'CALC_Vector!$B${V0}:$B${V1},$A{r + 2 + i})/$D{r + 2 + i},"")')
        ).number_format = PCT
    ws.cell(row=r + 1, column=4, value='Spend $m').font = H3
    ws.cell(row=r + 1, column=5, value='Conversion').font = H3
    for col, w in zip(['A', 'B', 'C'], [12, 28, 26]):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f'A5:C{O1}'
    ws.freeze_panes = 'D6'

    # ======================================================== QA_Checks
    ws = wb.create_sheet('QA_Checks')
    band(ws, 1, 'QA_Checks', 6)
    hdr(ws, 6, ['#', 'Gate', 'Result', 'What it catches'])
    Q0 = 7
    gates = [
        ('Every line with spend has a region',
         f'=IF(SUMPRODUCT(--(IN_Shock!$G${I0}:$G${I1}<>0),'
         f'--(IN_Shock!$B${I0}:$B${I1}=""))=0,"PASS","FAIL")',
         'A line that will silently resolve to nothing'),
        ('Every line region is a known region',
         f'=IF(SUMPRODUCT(--(IN_Shock!$B${I0}:$B${I1}<>""),'
         f'--ISNA(MATCH(IN_Shock!$B${I0}:$B${I1},{REGLIST},0)))=0,"PASS","FAIL")',
         'A typo in the region column'),
        ('Every line resolves in its regions Table 5',
         f'=IF(SUMPRODUCT(--(IN_Shock!$G${I0}:$G${I1}<>0),'
         f'--(MAP_ShockKeys!$K${K0}:$K${K1}=0))=0,"PASS","FAIL")',
         'A region and product pair with no row in the flow table'),
        ('Every line resolves in the margin tables',
         f'=IF(SUMPRODUCT(--(IN_Shock!$G${I0}:$G${I1}<>0),'
         f'--(MAP_ShockKeys!$M${K0}:$M${K1}=0))=0,"PASS","FAIL")',
         'A code outside the ABS 115 spine'),
        ('Shares sum to 100% on every line',
         f'=IF(SUMPRODUCT(--(ABS(MAP_LineStrip!${CL(CHKC)}${S0}:${CL(CHKC)}${S1}-1)'
         f'>0.0001))=0,"PASS","FAIL")',
         'A broken rate calculation'),
        ('No line lands on an empty ABS cell',
         f'=IF(SUMPRODUCT(--(IN_Shock!$G${I0}:$G${I1}<>0),'
         f'--(MAP_LineStrip!${CL(PPC)}${S0}:${CL(PPC)}${S1}=0))=0,"PASS","REVIEW")',
         'The bill-of-quantities-into-GFCF trap: no ABS cell to strip against'),
        ('No negative domestic or import share',
         f'=IF(SUMPRODUCT(--(MAP_LineStrip!${CL(SH0)}${S0}:${CL(SH0 + 1)}${S1}<0))=0,'
         f'"PASS","REVIEW")',
         'ABS cells with net disposals or import adjustments'),
        ('Strip components sum back to the line total',
         f'=IF(ABS(SUM(CALC_Strip!$F${P0}:$F${P1})-IN_Shock!$G${TR})<0.01,"PASS","FAIL")',
         'Margins lost or double counted'),
        ('Direct vector reconciles to the strip',
         f'=IF(ABS(SUM(CALC_Vector!$D${V0}:$D${V1})-'
         f'(SUMIFS(CALC_Strip!$F${P0}:$F${P1},CALC_Strip!$D${P0}:$D${P1},"Domestic",'
         f'CALC_Strip!$E${P0}:$E${P1},"N")+SUM(CALC_Margins!$D${MG0}:$D${MG1})))'
         f'<0.01,"PASS","FAIL")',
         'Margins lost between the strip and the vector'),
        ('Every vector row found its multipliers',
         f'=IF(SUMPRODUCT(--(CALC_Vector!$D${V0}:$D${V1}<>0),'
         f'--(CALC_Vector!${CL(mult_col[("Output multipliers", "TOTAL")])}${V0}:'
         f'${CL(mult_col[("Output multipliers", "TOTAL")])}${V1}=0))=0,"PASS","FAIL")',
         'A region and industry pair missing from the multiplier set'),
        ('State results are not being read as national',
         f'=IF(COUNTIF(IN_Shock!$B${I0}:$B${I1},"Aus")=COUNTIF(IN_Shock!$B${I0}:$B${I1},"<>")'
         f',"PASS","REVIEW")',
         'ALL REGIONS sums regional effects and understates the national impact'),
        ('Vintages match',
         '=IF(Settings!$B$8=Settings!$B$9,"PASS","REVIEW")',
         'Multipliers 2022-23 against ABS strip data 2023-24'),
        ('6700 bridged to 6701',
         f'=IF(SUM(Lists!$E${L0}:$E${L1})=0,"PASS","REVIEW")',
         'Imputed rent borrows actual rent'),
        ('Margin and tax rates are national',
         f'=IF(COUNTIF(IN_Shock!$B${I0}:$B${I1},"Aus")=COUNTIF(IN_Shock!$B${I0}:$B${I1},"<>"),'
         f'"PASS","REVIEW")',
         'The ABS publishes no state margin or tax matrices. Disclose it'),
    ]
    for i, (d, f, c) in enumerate(gates):
        r = Q0 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=d)
        ws.cell(row=r, column=3, value=f).font = H3
        ws.cell(row=r, column=4, value=c).font = NOTE
    Q1 = Q0 + len(gates) - 1
    ws.cell(row=3, column=1, value='OVERALL').font = H2
    ws.cell(row=4, column=1, value='STATUS').font = H2
    ws.cell(row=4, column=2, value=(
        f'=IF(COUNTIF($C${Q0}:$C${Q1},"FAIL")>0,"FAIL - do not use",'
        f'IF(COUNTIF($C${Q0}:$C${Q1},"REVIEW")>0,"REVIEW - disclose the items below","OK"))')
    ).font = Font(bold=True, size=12)
    ws.cell(row=4, column=3, value=(
        f'=COUNTIF($C${Q0}:$C${Q1},"FAIL")&" fail, "&COUNTIF($C${Q0}:$C${Q1},"REVIEW")'
        f'&" review, "&COUNTIF($C${Q0}:$C${Q1},"PASS")&" pass"'))
    for col, w in zip(['A', 'B', 'C', 'D'], [5, 50, 22, 60]):
        ws.column_dimensions[col].width = w

    # ======================================================== Assumptions
    ws = wb.create_sheet('Assumptions')
    band(ws, 1, 'Assumptions - attach to every report', 4)
    hdr(ws, 4, ['Item', 'What to disclose'])
    for i, (a, b) in enumerate([
        ('Multi-region totals', 'ALL REGIONS sums results computed with each regions own '
         'multipliers. It is NOT the national impact of the combined shock. A state leaks to '
         'the rest of Australia, so state multipliers are smaller and the sum understates '
         'the national effect. Never add ALL REGIONS to an Aus run.'),
        ('Margin region', 'Margin rates are national. The margin is assumed earned in the '
         'same region as the purchase, which will overstate local retail and wholesale '
         'activity where goods are distributed from another state.'),
        ('Margin and tax vintage', 'ABS 2023-24, against 2022-23 multipliers and flow tables.'),
        ('Multiplier source', 'Supplied by the provider, never derived here. Verified: all '
         'nine regions distinct, reconciling to the supplied Table 5 at 114/114 within 0.001.'),
        ('6700 Imputed rent', 'No row in the supplied 114-code set. Bridged to 6701. Carries '
         '$21m of net taxes and no margins, so the effect is small, but it is wrong for a '
         'housing study.'),
        ('Employment units', 'FTE. Australia has no employment column in the supplied set.'),
        ('GVA basis', 'Value added at basic prices (P1+P2+P4), the ABS headline.'),
        ('ABS caveat', 'Input-output multipliers assume no supply constraints, fixed prices '
         'and fixed input ratios, and are likely to significantly overstate impacts. Quote '
         'this, do not paraphrase it.'),
    ], 5):
        ws.cell(row=i, column=1, value=a).font = H3
        ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[i].height = 42
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 106

    n = sum(1 for s_ in wb.worksheets for row in s_.iter_rows()
            for c in row if isinstance(c.value, str) and c.value.startswith('='))
    print(f'  {len(wb.sheetnames)} tabs, {n:,} formulas, {len(vec_rows)} vector rows')
    return wb


if __name__ == '__main__':
    wb = main()
    out = SUBSET_OUT if os.environ.get('IOMODEL_SUBSET') else OUTFILE
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f'wrote {out}  ({out.stat().st_size / 1e6:.1f} MB)')
