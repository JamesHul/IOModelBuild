"""
Generate the IO impact model workbook from build/sources.pkl.

    python scripts/build_model.py

v0.3. Supersedes the v0.2 generator, which predates the verbatim RAW rule.
This one builds the whole chain on the stacked RAW layer:

    RAW_*  (verbatim, stacked, region-keyed)
      -> MAP_Spine        where each code sits in each RAW block, for the
                          selected region. One MATCH per code, reused everywhere
      -> MAP_Multipliers  selected region on the 115 spine, 6700 bridged to
                          6701, 'n.a.' preserved as text
      -> MAP_StripData    RAW aggregated to 8 purchasing column groups
      -> CALC_Rates       the shares, which must sum to 1
      -> IN_Shock         the spending lines
      -> CALC_Strip       each line split into domestic / imports / tax / margins
      -> CALC_Vector      margins reallocated to earning industries
      -> CALC_Impacts     vector x multipliers
      -> OUT_Summary / OUT_Detail / QA_Checks

Design notes that matter:

  * The region switch is an INDEX/MATCH on a Key column, never INDIRECT and
    never a nine-deep IF. That is the whole reason the RAW tabs are stacked.
  * Primary inputs are summed with SUMIF on RowType, never by row number,
    because Table 5 and Table 8 disagree about where they sit.
  * Every n.a. in the supplied multiplier set is in a Type 1A/1B/2A/2B ratio
    column - those ratios are undefined where the initial effect is zero. The
    seven effect columns are n.a.-free, and the impact arithmetic uses only
    those. MAP_Multipliers keeps the text so nothing pretends a ratio is 0.
  * 6700 Imputed rent has no row in the supplied 114-code set, so it is bridged
    to 6701 and flagged. It carries $21m of net taxes and no margins at all, so
    the bridge is close to immaterial - but it is still disclosed.
"""
import os
import pickle
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation

import build_raw_stacked as RS

ROOT = Path(__file__).resolve().parent.parent
SOURCES = ROOT / 'build' / 'sources.pkl'
OUTFILE = ROOT / 'output' / 'IO_Impact_Model_v0-3.xlsx'
SUBSET_OUT = ROOT / 'build' / 'IO_Model_subset.xlsx'

H1 = Font(bold=True, size=14, color='FFFFFF')
H2 = Font(bold=True, size=11)
H3 = Font(bold=True, size=10)
BLUE = Font(color='0000CC')            # hardcoded input
BLACK = Font(color='000000')           # formula
GREEN = Font(color='006100')           # cross-sheet link
NOTE = Font(italic=True, size=9, color='666666')
TITLEFILL = PatternFill('solid', fgColor='1F3864')
HDRFILL = PatternFill('solid', fgColor='DDEBF7')
YELLOW = PatternFill('solid', fgColor='FFF2CC')
GREYFILL = PatternFill('solid', fgColor='F2F2F2')
OKFILL = PatternFill('solid', fgColor='C6EFCE')
BADFILL = PatternFill('solid', fgColor='FFC7CE')
MONEY = '#,##0.0'
PCT = '0.0%'
NUM4 = '0.0000'

REGIONS = RS.REGIONS
# 8 purchasing column groups. first/last are 1-based columns WITHIN the source
# block, so one formula shape covers both a range (Intermediate) and a single
# column (each Q). Flow blocks carry 10 dummy industry columns after the 114;
# they are all zero and are included so the sum ties to the source's own total.
GROUPS = [
    ('Intermediate', 'Businesses buying inputs. A bill of quantities belongs here',
     RS.FLOW_IND_FIRST, RS.FLOW_IND_LAST, RS.MARG_IND_FIRST, RS.MARG_IND_LAST),
    ('Q1_HFCE', 'Households; final consumption expenditure', 128, 128, 119, 119),
    ('Q2_GGFCE', 'General government; final consumption expenditure', 129, 129, 120, 120),
    ('Q3_GFCF_Priv', 'Private; gross fixed capital formation', 130, 130, 121, 121),
    ('Q4_GFCF_PubCorp', 'Public corporations; GFCF', 131, 131, 122, 122),
    ('Q5_GFCF_GG', 'General government; GFCF', 132, 132, 123, 123),
    ('Q6_Inventories', 'Changes in inventories', 133, 133, 124, 124),
    ('Q7_Exports', 'Exports of goods and services', 134, 134, 125, 125),
]
NG = len(GROUPS)
MARGIN_ORDER = ['Wholesale', 'Retail', 'RestHotelClub', 'Road', 'Rail', 'Pipeline',
                'Water', 'Air', 'PortHandling', 'MarineIns', 'Gas', 'Electricity']
MEASURES = ['Domestic', 'Imports', 'NetTaxes'] + MARGIN_ORDER          # 15
NM = len(MEASURES)
# The strip block also carries the Australian basic values, because the
# national margin and tax rates have to be measured against the national cell.
STRIP_MEASURES = MEASURES + ['Aus_Domestic', 'Aus_Imports']
NYR = 8
NLINE = 40
EXAMPLE = [
    ('Contract value, new road', '3101', 'Q5_GFCF_GG', 'N', 100.0),
    ('Household food spending', '1101', 'Q1_HFCE', 'N', 20.0),
    ('Engineering design fees', '6901', 'Intermediate', 'N', 15.0),
]


def band(ws, row, text, width, font=H1):
    ws.cell(row=row, column=1, value=text).font = font
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = TITLEFILL


def hdr(ws, row, labels, col0=1, fill=HDRFILL):
    for i, h in enumerate(labels, col0):
        c = ws.cell(row=row, column=i, value=h)
        c.font = H3
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical='bottom')


def main():
    if not SOURCES.exists():
        raise SystemExit(f"{SOURCES} not found. Run scripts/load_sources.py first.")
    src = pickle.load(open(SOURCES, 'rb'))
    abs_t, flows, mult = src['abs'], src['flows'], src['multipliers']
    if not flows or not mult or not abs_t:
        raise SystemExit('sources.pkl incomplete. Run load_sources.py.')

    # ---- the 115-code spine, from the ABS margin tables --------------------
    spine = [(m['code'], m['label']) for m in abs_t['T23']['meta']
             if m['row_type'] == 'Product']
    # IOMODEL_SUBSET builds a structurally identical but much smaller workbook -
    # same formula shapes, fewer spine rows - so the whole chain can actually be
    # recalculated and checked. It is a test harness, never a delivery build.
    if os.environ.get('IOMODEL_SUBSET'):
        keep = {e[1] for e in EXAMPLE} | {v[1] for v in src['margin_tables'].values()}
        keep |= {'6700', '6701'}
        spine = [x for x in spine if x[0] in keep]
        print(f'  SUBSET BUILD: {len(spine)} spine rows')
    NS = len(spine)
    mult_codes = {m['code'] for m in mult['regions']['Aus']['meta']
                  if m['row_type'] == 'Industry'}
    blocks = mult['blocks']
    effects = mult['effects']
    NE = len(effects)
    i_init = effects.index('Initial effect')
    i_prod = effects.index('Production-induced effect')
    i_cons = effects.index('Consumption-induced effect')
    i_totl = effects.index('Total multiplier')
    i_simp = effects.index('Simple multiplier')
    mt = src['margin_tables']
    earner = {v[0]: v[1] for v in mt.values()}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ======================================================== README
    ws = wb.create_sheet('README')
    band(ws, 1, 'Australian Input-Output impact model  -  v0.3', 8)
    rows = [
        ('Built', f'{date.today():%Y-%m-%d} by scripts/build_model.py from build/sources.pkl'),
        ('', ''),
        ('THE CHAIN', 'each tab reads only the ones above it'),
        ('  RAW_T5 / RAW_T8', 'Provider flow tables, 9 regions stacked, verbatim, $m 2022-23'),
        ('  RAW_Multipliers', 'Provider multiplier set, 9 regions, 15 measures x 11 effects'),
        ('  RAW_Margins', 'ABS Tables 23-34 and 35, verbatim, $m 2023-24, NATIONAL only'),
        ('  RAW_T21_Control', 'ABS Table 21, the independent margin control total'),
        ('  Lists / Lists_MarginMap', 'The 115 spine, the 114 bridge, column groups, margin map'),
        ('  MAP_Spine', 'Where each code sits in each RAW block for the selected region'),
        ('  MAP_Multipliers', 'Selected region on the 115 spine, 6700 bridged, n.a. kept'),
        ('  MAP_StripData', 'RAW aggregated to the 8 purchasing column groups'),
        ('  CALC_Rates', 'The shares. The Check column must read 100%'),
        ('  IN_Shock', 'YOUR INPUT. Spending lines at purchasers prices'),
        ('  CALC_Strip', 'Each line split into domestic / imports / tax / 12 margins'),
        ('  CALC_Vector', 'Margins reallocated to earning industries; direct domestic vector'),
        ('  CALC_Impacts', 'Vector x multipliers'),
        ('  OUT_Summary / OUT_Detail', 'Results'),
        ('  QA_Checks', 'The gates. Read this before believing any number'),
        ('', ''),
        ('HOW TO RUN', ''),
        ('  1', 'Set the region on Settings'),
        ('  2', 'Enter spending lines on IN_Shock (yellow cells)'),
        ('  3', 'Read QA_Checks. If the banner is not OK, do not use the result'),
        ('  4', 'Read OUT_Summary'),
        ('', ''),
        ('COLOUR KEY', 'blue = hardcoded input, black = formula, green = cross-sheet link, '
                       'yellow fill = you must complete it'),
        ('', ''),
        ('KNOWN LIMITATIONS', ''),
        ('  Vintage', 'Multipliers and flow tables are 2022-23; ABS margin and tax data is '
                      '2023-24. One year apart. QA gate flags it'),
        ('  State margins', 'The ABS publishes no state margin or tax matrices. State runs use '
                            'national margin and tax rates'),
        ('  6700', 'Imputed rent for owner-occupiers has no row in the supplied 114-code set. '
                   'Bridged to 6701 and flagged'),
        ('  Exports', 'Table 5 and Table 8 treat re-exports differently, so the Q7 import share '
                      'is unreliable'),
        ('  ABS caveat', 'Multipliers assume no supply constraints, fixed prices and fixed input '
                         'ratios, and are likely to overstate impacts. Quote this in every report'),
    ]
    for i, (a, b) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=a).font = H3 if a and not a.startswith(' ') else BLACK
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 100

    # ======================================================== Settings
    ws = wb.create_sheet('Settings')
    band(ws, 1, 'Settings', 6)
    setrows = [
        ('Region', 'Aus', 'Which region the flow tables and multipliers come from'),
        ('Study type', 'Impact', 'Impact or Contribution'),
        ('Headline GVA basis', 'Value added', 'Basic prices. NOT market prices, which include P3'),
        ('Shock price year', '2022-23', 'The price year your spending lines are in'),
        ('Multiplier vintage', '2022-23', 'From the supplied set'),
        ('ABS strip vintage', '2023-24', 'From the ABS margin and tax tables'),
        ('Years modelled', NYR, f'Up to {NYR}'),
    ]
    hdr(ws, 4, ['Setting', 'Value', 'Note'])
    for i, (k, v, n) in enumerate(setrows, 5):
        ws.cell(row=i, column=1, value=k).font = H3
        c = ws.cell(row=i, column=2, value=v)
        c.font = BLUE
        c.fill = YELLOW
        ws.cell(row=i, column=3, value=n).font = NOTE
    dv = DataValidation(type='list', formula1='"%s"' % ','.join(REGIONS), allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=5, column=2))
    dv2 = DataValidation(type='list', formula1='"Impact,Contribution"')
    ws.add_data_validation(dv2)
    dv2.add(ws.cell(row=6, column=2))
    for col, w in zip('ABC', [22, 16, 74]):
        ws.column_dimensions[col].width = w
    REGION = 'Settings!$B$5'

    # ======================================================== Lists
    ws = wb.create_sheet('Lists')
    band(ws, 1, 'Lists  -  the 115-code IOIG(2022) spine, the 114-code bridge, and the '
                'purchasing column groups', 12)
    ws.cell(row=2, column=1, value=(
        'IOIG(2022) has 115 industries. The supplied set uses 114: it has no 6700 and '
        'treats ownership of dwellings as a single 6701. That is the only spine '
        'difference, so the bridge is one code.')).font = NOTE
    hdr(ws, 4, ['#', 'IOIG code', 'Industry name', 'Code to use for the supplied set',
                'Bridged?'])
    L0 = 5
    for i, (code, name) in enumerate(spine):
        r = L0 + i
        ws.cell(row=r, column=1, value=i + 1)
        c = ws.cell(row=r, column=2, value=code)
        c.number_format = '@'
        c.font = BLUE
        ws.cell(row=r, column=3, value=name)
        use = code if code in mult_codes else '6701'
        c = ws.cell(row=r, column=4, value=use)
        c.number_format = '@'
        c.font = BLUE
        f = ws.cell(row=r, column=5, value=(1 if use != code else 0))
        if use != code:
            f.fill = YELLOW
    L1 = L0 + NS - 1
    # column groups
    G0 = 5
    hdr(ws, 4, ['Group', 'What it is', 'Flow first', 'Flow last', 'Margin first',
                'Margin last'], col0=7)
    for i, g in enumerate(GROUPS):
        r = G0 + i
        for j, v in enumerate(g, 7):
            ws.cell(row=r, column=j, value=v).font = BLUE if j > 8 else BLACK
    G1 = G0 + NG - 1
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'G', 'H', 'I', 'J', 'K', 'L'],
                      [5, 10, 46, 12, 9, 17, 44, 10, 10, 11, 11]):
        ws.column_dimensions[col].width = w
    CODES = f'Lists!$B${L0}:$B${L1}'
    USECODE = f'Lists!$D${L0}:$D${L1}'
    GRPNAMES = f'Lists!$G${G0}:$G${G1}'

    # ======================================================== RAW tabs
    _sub = {c for c, _ in spine} if os.environ.get('IOMODEL_SUBSET') else None
    geo = RS.write_raw_tabs(wb, src, verbose=False, subset=_sub,
                            regions=['Aus'] if _sub else None)
    mm = RS.write_margin_map(wb, src)

    def rng(tab, first_col):
        """Full verbatim rectangle of a stacked tab, as an absolute range."""
        g = geo[tab]
        c0 = first_col
        c1 = first_col + g['width'] - 1
        return f"{tab}!${CL(c0)}${g['first']}:${CL(c1)}${g['last']}"

    def keyrng(tab, keycol):
        g = geo[tab]
        return f"{tab}!${CL(keycol)}${g['first']}:${CL(keycol)}${g['last']}"

    T5R, T8R = rng('RAW_T5', RS.FLOW_VERB_COL), rng('RAW_T8', RS.FLOW_VERB_COL)
    MUR = rng('RAW_Multipliers', RS.FLOW_VERB_COL)
    MGR = rng('RAW_Margins', RS.MARG_VERB_COL)
    T5K, T8K = keyrng('RAW_T5', 5), keyrng('RAW_T8', 5)
    MUK, MGK = keyrng('RAW_Multipliers', 5), keyrng('RAW_Margins', 6)

    # ======================================================== MAP_Spine
    ws = wb.create_sheet('MAP_Spine')
    band(ws, 1, 'MAP_Spine  -  where each code sits in each RAW block, for the selected '
                'region. One MATCH per code, reused by every tab downstream', 20)
    ws.cell(row=2, column=1, value=(
        'Position, resolved once. Everything downstream indexes off these row numbers '
        'instead of repeating the MATCH. 0 means the code is not in that block.')).font = NOTE
    mcols = MARGIN_ORDER + ['NetTaxes']
    hdr(ws, 4, ['Code', 'Name', 'Use code', 'Bridged', 'T5 row', 'T8 row', 'Mult row']
        + [m[:11] + ' row' for m in mcols] + ['Aus T5 row', 'Aus T8 row'])
    S0 = 5
    for i in range(NS):
        r = S0 + i
        lr = L0 + i
        c = ws.cell(row=r, column=1, value=f'=Lists!$B${lr}')
        c.number_format = '@'
        c.font = GREEN
        ws.cell(row=r, column=2, value=f'=Lists!$C${lr}').font = GREEN
        c = ws.cell(row=r, column=3, value=f'=Lists!$D${lr}')
        c.number_format = '@'
        c.font = GREEN
        ws.cell(row=r, column=4, value=f'=Lists!$E${lr}').font = GREEN
        ws.cell(row=r, column=5, value=f'=IFERROR(MATCH({REGION}&"|"&$C{r},{T5K},0),0)')
        ws.cell(row=r, column=6, value=f'=IFERROR(MATCH({REGION}&"|"&$C{r},{T8K},0),0)')
        ws.cell(row=r, column=7, value=f'=IFERROR(MATCH({REGION}&"|"&$C{r},{MUK},0),0)')
        for j, mn in enumerate(mcols, 8):
            ws.cell(row=r, column=j, value=f'=IFERROR(MATCH("{mn}|"&$A{r},{MGK},0),0)')
        # The Australian rows too. Margins and net taxes are published only
        # nationally, so their RATES must be taken against the national cell.
        # A national margin over a state-sized denominator makes the rate climb
        # as the state shrinks - 31% of a household meat dollar nationally but
        # 77% in Tasmania. A margin rate is a rate.
        ws.cell(row=r, column=21, value=f'=IFERROR(MATCH("Aus|"&$C{r},{T5K},0),0)')
        ws.cell(row=r, column=22, value=f'=IFERROR(MATCH("Aus|"&$C{r},{T8K},0),0)')
    S1 = S0 + NS - 1
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 40
    ws.freeze_panes = 'C5'

    # ======================================================== MAP_Multipliers
    ws = wb.create_sheet('MAP_Multipliers')
    band(ws, 1, 'MAP_Multipliers  -  the selected region on the 115 spine. 15 measure '
                'blocks x 11 effects', 8)
    ws.cell(row=2, column=1, value=(
        "Multipliers are an INPUT and are never derived here. 6700 reads 6701's row and is "
        "flagged in the Bridged column. 'n.a.' is kept as text, not turned into 0 - every "
        "n.a. in the supplied set is a Type 1A/1B/2A/2B ratio that is undefined because the "
        "initial effect is zero. The seven effect columns carry no n.a., and the impact "
        "arithmetic uses only those.")).font = NOTE
    ws.cell(row=3, column=1, value='Headline GVA is the "Value added" block (basic prices). '
                                   'NOT "at market prices", which includes taxes on products.').font = NOTE
    hdr(ws, 5, ['Code', 'Name', 'Bridged'])
    for bi, b in enumerate(blocks):
        c0 = 4 + bi * NE
        c = ws.cell(row=4, column=c0, value=b['name'])
        c.font = H3
        c.fill = GREYFILL
        for j, e in enumerate(effects):
            h = ws.cell(row=5, column=c0 + j, value=e)
            h.font = H3
            h.fill = HDRFILL
            h.alignment = Alignment(wrap_text=True)
    M0 = 6
    for i in range(NS):
        r = M0 + i
        sr = S0 + i
        c = ws.cell(row=r, column=1, value=f'=MAP_Spine!$A{sr}')
        c.number_format = '@'
        c.font = GREEN
        ws.cell(row=r, column=2, value=f'=MAP_Spine!$B{sr}').font = GREEN
        f = ws.cell(row=r, column=3, value=f'=MAP_Spine!$D{sr}')
        f.font = GREEN
        for bi, b in enumerate(blocks):
            for j in range(NE):
                col = 4 + bi * NE + j
                sc = b['first_col'] + j
                idx = f'INDEX({MUR},MAP_Spine!$G{sr},{sc})'
                ws.cell(row=r, column=col, value=(
                    f'=IF(MAP_Spine!$G{sr}=0,"",IF(ISNUMBER({idx}),{idx},"n.a."))')
                ).number_format = NUM4
    M1 = M0 + NS - 1
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 36
    ws.freeze_panes = 'D6'

    def mcol(block_name, eff_idx):
        bi = next(i for i, b in enumerate(blocks) if b['name'] == block_name)
        return 4 + bi * NE + eff_idx

    # ======================================================== MAP_StripData
    ws = wb.create_sheet('MAP_StripData')
    band(ws, 1, 'MAP_StripData  -  RAW aggregated to the 8 purchasing column groups, $m', 10)
    ws.cell(row=2, column=1, value=(
        'Domestic is Table 5. Imports is Table 8 less Table 5, which is why no ABS import '
        'table is needed. Net taxes and the 12 margins are national ABS data - the ABS '
        'publishes no state margin or tax matrices, so a state run uses national rates.')).font = NOTE
    hdr(ws, 5, ['Code', 'Name'])
    for mi, meas in enumerate(STRIP_MEASURES):
        c0 = 3 + mi * NG
        c = ws.cell(row=4, column=c0, value=meas)
        c.font = H3
        c.fill = GREYFILL
        for gi, g in enumerate(GROUPS):
            h = ws.cell(row=5, column=c0 + gi, value=g[0])
            h.font = H3
            h.fill = HDRFILL
            h.alignment = Alignment(wrap_text=True)
    D0 = 6
    for i in range(NS):
        r = D0 + i
        sr = S0 + i
        c = ws.cell(row=r, column=1, value=f'=MAP_Spine!$A{sr}')
        c.number_format = '@'
        c.font = GREEN
        ws.cell(row=r, column=2, value=f'=MAP_Spine!$B{sr}').font = GREEN
        for gi, g in enumerate(GROUPS):
            gr = G0 + gi
            ff, fl = f'Lists!$I${gr}', f'Lists!$J${gr}'
            mf, ml = f'Lists!$K${gr}', f'Lists!$L${gr}'
            t5 = (f'SUM(INDEX({T5R},MAP_Spine!$E{sr},{ff}):'
                  f'INDEX({T5R},MAP_Spine!$E{sr},{fl}))')
            t8 = (f'SUM(INDEX({T8R},MAP_Spine!$F{sr},{ff}):'
                  f'INDEX({T8R},MAP_Spine!$F{sr},{fl}))')
            # Domestic
            ws.cell(row=r, column=3 + 0 * NG + gi, value=(
                f'=IF(MAP_Spine!$E{sr}=0,0,{t5})')).number_format = MONEY
            # Imports = T8 - T5. Parenthesised: precedence has bitten before.
            ws.cell(row=r, column=3 + 1 * NG + gi, value=(
                f'=IF(OR(MAP_Spine!$E{sr}=0,MAP_Spine!$F{sr}=0),0,({t8})-({t5}))')
            ).number_format = MONEY
            # the same two cells for AUSTRALIA, which is what the national
            # margin and tax rates must be measured against
            t5a = (f'SUM(INDEX({T5R},MAP_Spine!$U{sr},{ff}):'
                   f'INDEX({T5R},MAP_Spine!$U{sr},{fl}))')
            t8a = (f'SUM(INDEX({T8R},MAP_Spine!$V{sr},{ff}):'
                   f'INDEX({T8R},MAP_Spine!$V{sr},{fl}))')
            ws.cell(row=r, column=3 + 15 * NG + gi, value=(
                f'=IF(MAP_Spine!$U{sr}=0,0,{t5a})')).number_format = MONEY
            ws.cell(row=r, column=3 + 16 * NG + gi, value=(
                f'=IF(OR(MAP_Spine!$U{sr}=0,MAP_Spine!$V{sr}=0),0,({t8a})-({t5a}))')
            ).number_format = MONEY
            # Net taxes and the 12 margins, all from RAW_Margins
            for mi, meas in enumerate(MEASURES[2:], 2):
                which = 'NetTaxes' if meas == 'NetTaxes' else meas
                mrow = f'MAP_Spine!${CL(8 + mcols.index(which))}{sr}'
                mg = (f'SUM(INDEX({MGR},{mrow},{mf}):INDEX({MGR},{mrow},{ml}))')
                ws.cell(row=r, column=3 + mi * NG + gi, value=(
                    f'=IF({mrow}=0,0,{mg})')).number_format = MONEY
    D1 = D0 + NS - 1
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 36
    ws.freeze_panes = 'C6'

    # ======================================================== CALC_Rates
    ws = wb.create_sheet('CALC_Rates')
    band(ws, 1, 'CALC_Rates  -  the share of each purchasers-price dollar. The Check '
                'columns must all read 100%', 10)
    ws.cell(row=2, column=1, value=(
        'Margins and net taxes are published nationally only, so their rates are taken '
        'against the NATIONAL purchasers-price cell. What the selected region supplies is '
        'the split of the basic-price portion between domestic and imported. The four kinds '
        'of share sum to exactly 1, which is the gate on the right.')).font = NOTE
    hdr(ws, 5, ['Code', 'Name'])
    c = ws.cell(row=4, column=3, value='National purchasers price, $m')
    c.font = H3
    c.fill = GREYFILL
    for gi, g in enumerate(GROUPS):
        h = ws.cell(row=5, column=3 + gi, value=g[0])
        h.font = H3
        h.fill = HDRFILL
        h.alignment = Alignment(wrap_text=True)
    c = ws.cell(row=4, column=3 + NG, value='Basic share of the national cell')
    c.font = H3
    c.fill = GREYFILL
    for gi, g in enumerate(GROUPS):
        h = ws.cell(row=5, column=3 + NG + gi, value=g[0])
        h.font = H3
        h.fill = HDRFILL
        h.alignment = Alignment(wrap_text=True)
    RATE0 = 3 + 2 * NG
    for mi, meas in enumerate(MEASURES):
        c0 = RATE0 + mi * NG
        c = ws.cell(row=4, column=c0, value=meas + ' share')
        c.font = H3
        c.fill = GREYFILL
        for gi, g in enumerate(GROUPS):
            h = ws.cell(row=5, column=c0 + gi, value=g[0])
            h.font = H3
            h.fill = HDRFILL
            h.alignment = Alignment(wrap_text=True)
    CHK0 = RATE0 + NM * NG
    c = ws.cell(row=4, column=CHK0, value='Check - must be 100%')
    c.font = H3
    c.fill = GREYFILL
    for gi, g in enumerate(GROUPS):
        h = ws.cell(row=5, column=CHK0 + gi, value=g[0])
        h.font = H3
        h.fill = HDRFILL
        h.alignment = Alignment(wrap_text=True)
    R0 = 6
    for i in range(NS):
        r = R0 + i
        dr = D0 + i
        c = ws.cell(row=r, column=1, value=f'=MAP_StripData!$A{dr}')
        c.number_format = '@'
        c.font = GREEN
        ws.cell(row=r, column=2, value=f'=MAP_StripData!$B{dr}').font = GREEN
        for gi in range(NG):
            # The national purchasers-price cell: Australian basic value plus the
            # national net taxes and margins. Margin and tax RATES are measured
            # against this, never against a state-sized denominator.
            nat = ['MAP_StripData!%s%d' % (CL(3 + 15 * NG + gi), dr),
                   'MAP_StripData!%s%d' % (CL(3 + 16 * NG + gi), dr)]
            nat += ['MAP_StripData!%s%d' % (CL(3 + mi * NG + gi), dr)
                    for mi in range(2, NM)]
            ws.cell(row=r, column=3 + gi, value=f'=SUM({",".join(nat)})').number_format = MONEY
            # basic (domestic + imports) share of the national cell
            ws.cell(row=r, column=3 + NG + gi, value=(
                f'=IFERROR((MAP_StripData!{CL(3 + 15 * NG + gi)}{dr}'
                f'+MAP_StripData!{CL(3 + 16 * NG + gi)}{dr})/${CL(3 + gi)}{r},0)')
            ).number_format = PCT
            # domestic and imports: the national basic share, split by the
            # SELECTED REGION's own domestic/import ratio. That split is the real
            # gain from holding state Table 5 and Table 8.
            reg_d = f'MAP_StripData!{CL(3 + 0 * NG + gi)}{dr}'
            reg_m = f'MAP_StripData!{CL(3 + 1 * NG + gi)}{dr}'
            ws.cell(row=r, column=RATE0 + 0 * NG + gi, value=(
                f'=IFERROR(${CL(3 + NG + gi)}{r}*{reg_d}/({reg_d}+{reg_m}),0)')
            ).number_format = PCT
            ws.cell(row=r, column=RATE0 + 1 * NG + gi, value=(
                f'=IFERROR(${CL(3 + NG + gi)}{r}*{reg_m}/({reg_d}+{reg_m}),0)')
            ).number_format = PCT
            # net taxes and the 12 margins: national rates
            for mi in range(2, NM):
                ws.cell(row=r, column=RATE0 + mi * NG + gi, value=(
                    f'=IFERROR(MAP_StripData!{CL(3 + mi * NG + gi)}{dr}/${CL(3 + gi)}{r},0)')
                ).number_format = PCT
            sh = ','.join(f'{CL(RATE0 + mi * NG + gi)}{r}' for mi in range(NM))
            ws.cell(row=r, column=CHK0 + gi, value=(
                f'=IF(${CL(3 + gi)}{r}=0,1,SUM({sh}))')).number_format = PCT
    R1 = R0 + NS - 1
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 36
    ws.freeze_panes = 'C6'
    RATES = f'CALC_Rates!${CL(RATE0)}${R0}:${CL(RATE0 + NM * NG - 1)}${R1}'
    RCODES = f'CALC_Rates!$A${R0}:$A${R1}'

    # ======================================================== IN_Shock
    ws = wb.create_sheet('IN_Shock')
    band(ws, 1, 'IN_Shock  -  YOUR INPUT. Spending lines at purchasers prices, $m', 14)
    ws.cell(row=2, column=1, value=(
        'One line per item of spending. Code is the IOIG of the PRODUCT bought. Group is '
        'who is buying: a bill of quantities goes to Intermediate, because the contractor '
        'is a business buying inputs. Putting materials in a GFCF column returns zero and '
        'silently drops the line - the GFCF columns hold finished capital assets.')).font = NOTE
    ws.cell(row=3, column=1, value=(
        "Direct only = Y for a line whose supply chain you have already itemised, such as "
        "the contractor's own wages. It is counted as direct but not put through the "
        "multiplier.")).font = NOTE
    hdr(ws, 5, ['#', 'Description', 'IOIG code', 'Column group', 'Direct only']
        + [f'Year {y + 1}' for y in range(NYR)])
    I0 = 6
    for i in range(NLINE):
        r = I0 + i
        ws.cell(row=r, column=1, value=i + 1)
        for j in range(2, 6 + NYR):
            cell = ws.cell(row=r, column=j)
            cell.fill = YELLOW
            cell.font = BLUE
            if j >= 6:
                cell.number_format = MONEY
        ws.cell(row=r, column=3).number_format = '@'
        if i < len(EXAMPLE):
            d, code, grp, only, amt = EXAMPLE[i]
            ws.cell(row=r, column=2, value=d)
            ws.cell(row=r, column=3, value=code)
            ws.cell(row=r, column=4, value=grp)
            ws.cell(row=r, column=5, value=only)
            ws.cell(row=r, column=6, value=amt)
        else:
            ws.cell(row=r, column=5, value='N')
    # Column N shows the purchasers-price value of the ABS cell this line strips
    # against. Zero means there is nothing there - the bill-of-quantities-into-GFCF
    # trap - and QA gate 7 counts them. It also replaces an array-MATCH gate that
    # returned #VALUE!: INDEX(range, MATCH(array,...)) does not evaluate as an array.
    hdr(ws, 5, ['ABS cell $m'], col0=6 + NYR)
    for i in range(NLINE):
        r = I0 + i
        ws.cell(row=r, column=6 + NYR, value=(
            f'=IFERROR(INDEX(CALC_Rates!$C${R0}:$J${R1},'
            f'MATCH($C{r},{RCODES},0),MATCH($D{r},{GRPNAMES},0)),0)')).number_format = MONEY
    ws.column_dimensions[CL(6 + NYR)].width = 13
    I1 = I0 + NLINE - 1
    # No leading '=' : xlsx data validation takes a bare range, and a stray
    # equals sign makes the whole workbook unloadable.
    dv3 = DataValidation(type='list', formula1=GRPNAMES)
    ws.add_data_validation(dv3)
    dv3.add(f'D{I0}:D{I1}')
    dv4 = DataValidation(type='list', formula1='"Y,N"')
    ws.add_data_validation(dv4)
    dv4.add(f'E{I0}:E{I1}')
    tr = I1 + 2
    ws.cell(row=tr, column=2, value='TOTAL').font = H2
    for y in range(NYR):
        ws.cell(row=tr, column=6 + y, value=(
            f'=SUM({CL(6 + y)}{I0}:{CL(6 + y)}{I1})')).number_format = MONEY
    for col, w in zip(['A', 'B', 'C', 'D', 'E'], [4, 38, 11, 18, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'F6'

    # ======================================================== CALC_Strip
    ws = wb.create_sheet('CALC_Strip')
    band(ws, 1, 'CALC_Strip  -  each line split into domestic, imports, net taxes and the '
                '12 margins, $m', 12)
    ws.cell(row=2, column=1, value=(
        'Line amount x the share for that product and column group. Components sum back to '
        'the line, which is QA gate 6.')).font = NOTE
    hdr(ws, 5, ['Line', 'Code', 'Group', 'Measure', 'Direct only']
        + [f'Year {y + 1}' for y in range(NYR)])
    P0 = 6
    for li in range(NLINE):
        for mi, meas in enumerate(MEASURES):
            r = P0 + li * NM + mi
            ir = I0 + li
            ws.cell(row=r, column=1, value=f'=IN_Shock!$A{ir}').font = GREEN
            c = ws.cell(row=r, column=2, value=f'=IN_Shock!$C{ir}')
            c.number_format = '@'
            c.font = GREEN
            ws.cell(row=r, column=3, value=f'=IN_Shock!$D{ir}').font = GREEN
            ws.cell(row=r, column=4, value=meas)
            ws.cell(row=r, column=5, value=f'=IN_Shock!$E{ir}').font = GREEN
            for y in range(NYR):
                ws.cell(row=r, column=6 + y, value=(
                    f'=IFERROR(IN_Shock!{CL(6 + y)}{ir}*INDEX({RATES},'
                    f'MATCH($B{r},{RCODES},0),{mi * NG}+MATCH($C{r},{GRPNAMES},0)),0)')
                ).number_format = MONEY
    P1 = P0 + NLINE * NM - 1
    for col, w in zip(['A', 'B', 'C', 'D', 'E'], [6, 9, 18, 15, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'F6'

    # ======================================================== CALC_Margins
    ws = wb.create_sheet('CALC_Margins')
    band(ws, 1, 'CALC_Margins  -  margin totals, and the industry that earns them', 12)
    ws.cell(row=2, column=1, value=(
        'A margin is not paid to the producer of the good. It is revenue of the wholesaler, '
        'retailer or carrier, so it is reallocated to that industry before the multipliers '
        'are applied. The mapping is Lists_MarginMap, verified against ABS Table 21.')).font = NOTE
    hdr(ws, 5, ['Margin', 'Earning IOIG'] + [f'Year {y + 1}' for y in range(NYR)])
    MG0 = 6
    for i, mn in enumerate(MARGIN_ORDER):
        r = MG0 + i
        ws.cell(row=r, column=1, value=mn)
        c = ws.cell(row=r, column=2, value=(
            f'=INDEX(Lists_MarginMap!$C${mm["first"]}:$C${mm["last"]},'
            f'MATCH($A{r},Lists_MarginMap!$B${mm["first"]}:$B${mm["last"]},0))'))
        c.number_format = '@'
        c.font = GREEN
        for y in range(NYR):
            ws.cell(row=r, column=3 + y, value=(
                f'=SUMIFS(CALC_Strip!{CL(6 + y)}${P0}:{CL(6 + y)}${P1},'
                f'CALC_Strip!$D${P0}:$D${P1},$A{r},'
                f'CALC_Strip!$E${P0}:$E${P1},"N")')).number_format = MONEY
    MG1 = MG0 + len(MARGIN_ORDER) - 1
    for col, w in zip(['A', 'B'], [16, 13]):
        ws.column_dimensions[col].width = w

    # ======================================================== CALC_Vector
    ws = wb.create_sheet('CALC_Vector')
    band(ws, 1, 'CALC_Vector  -  the direct domestic vector by industry, $m. This is what '
                'goes through the multipliers', 12)
    ws.cell(row=2, column=1, value=(
        'Domestic content of every line landing on this product, plus any margin this '
        'industry earns from other lines. Direct-only lines are excluded here - they are '
        'counted as direct on OUT_Summary but never multiplied.')).font = NOTE
    hdr(ws, 5, ['Code', 'Name'] + [f'Year {y + 1}' for y in range(NYR)])
    V0 = 6
    for i in range(NS):
        r = V0 + i
        sr = S0 + i
        c = ws.cell(row=r, column=1, value=f'=MAP_Spine!$A{sr}')
        c.number_format = '@'
        c.font = GREEN
        ws.cell(row=r, column=2, value=f'=MAP_Spine!$B{sr}').font = GREEN
        for y in range(NYR):
            ws.cell(row=r, column=3 + y, value=(
                f'=SUMIFS(CALC_Strip!{CL(6 + y)}${P0}:{CL(6 + y)}${P1},'
                f'CALC_Strip!$B${P0}:$B${P1},$A{r},'
                f'CALC_Strip!$D${P0}:$D${P1},"Domestic",'
                f'CALC_Strip!$E${P0}:$E${P1},"N")'
                f'+SUMIFS(CALC_Margins!{CL(3 + y)}${MG0}:{CL(3 + y)}${MG1},'
                f'CALC_Margins!$B${MG0}:$B${MG1},$A{r})')).number_format = MONEY
    V1 = V0 + NS - 1
    tr = V1 + 2
    ws.cell(row=tr, column=2, value='TOTAL DIRECT DOMESTIC').font = H2
    for y in range(NYR):
        ws.cell(row=tr, column=3 + y, value=(
            f'=SUM({CL(3 + y)}{V0}:{CL(3 + y)}{V1})')).font = H2
        ws.cell(row=tr, column=3 + y).number_format = MONEY
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 36
    ws.freeze_panes = 'C6'
    VTOT = tr

    # ======================================================== CALC_Impacts
    REPORT = [('Output multipliers', 'Output, $m', MONEY),
              ('Income multipliers', 'Wages and salaries, $m', MONEY),
              ('Value added multipliers', 'Value added at basic prices, $m', MONEY),
              ('Employed multipliers', 'Employment, FTE', '#,##0'),
              ('Gross operating surplus & mixed income multipliers', 'GOS and mixed income, $m',
               MONEY)]
    EFFECTS_OUT = [('Direct', i_init), ('Indirect (production-induced)', i_prod),
                   ('Induced (consumption-induced)', i_cons), ('TOTAL', i_totl)]
    ws = wb.create_sheet('CALC_Impacts')
    band(ws, 1, 'CALC_Impacts  -  the direct vector times the selected regions '
                'multipliers', 12)
    ws.cell(row=2, column=1, value=(
        'SUMPRODUCT of the direct vector and the multiplier column, industry by industry. '
        'Only the seven effect columns are used - the Type 1A/1B/2A/2B ratios carry the '
        'n.a. cells and are presentational.')).font = NOTE
    hdr(ws, 5, ['Measure', 'Effect'] + [f'Year {y + 1}' for y in range(NYR)] + ['Total'])
    C0 = 6
    r = C0
    impact_rows = {}
    for bn, label, fmt in REPORT:
        for en, ei in EFFECTS_OUT:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=en).font = H3 if en == 'TOTAL' else BLACK
            col = mcol(bn, ei)
            for y in range(NYR):
                ws.cell(row=r, column=3 + y, value=(
                    f'=SUMPRODUCT(CALC_Vector!{CL(3 + y)}${V0}:{CL(3 + y)}${V1},'
                    f'MAP_Multipliers!${CL(col)}${M0}:${CL(col)}${M1})')).number_format = fmt
            ws.cell(row=r, column=3 + NYR, value=(
                f'=SUM({CL(3)}{r}:{CL(2 + NYR)}{r})')).number_format = fmt
            impact_rows[(bn, en)] = r
            r += 1
    C1 = r - 1
    for col, w in zip(['A', 'B'], [30, 28]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'C6'

    # ======================================================== OUT_Summary
    ws = wb.create_sheet('OUT_Summary')
    band(ws, 1, 'OUT_Summary  -  headline results', 12)
    ws.cell(row=2, column=1, value=f'=\"Region: \"&{REGION}&\"   |   multipliers \"&'
                                   'Settings!$B$9&"   |   strip data "&Settings!$B$10').font = H2
    ws.cell(row=3, column=1, value='=IF(QA_Checks!$B$4="OK","",'
                                   '"CHECK QA_Checks BEFORE USING THESE NUMBERS")').font = \
        Font(bold=True, color='9C0006')
    hdr(ws, 5, ['Measure', 'Effect'] + [f'Year {y + 1}' for y in range(NYR)] + ['Total'])
    O0 = 6
    r = O0
    for bn, label, fmt in REPORT:
        for en, ei in EFFECTS_OUT:
            ws.cell(row=r, column=1, value=label)
            b = ws.cell(row=r, column=2, value=en)
            src_r = impact_rows[(bn, en)]
            for y in range(NYR + 1):
                c = ws.cell(row=r, column=3 + y, value=(
                    f'=CALC_Impacts!{CL(3 + y)}{src_r}'))
                c.number_format = fmt
                c.font = GREEN
                if en == 'TOTAL':
                    c.font = Font(bold=True, color='006100')
            if en == 'TOTAL':
                b.font = H3
                for cc in range(1, 4 + NYR):
                    ws.cell(row=r, column=cc).border = Border(top=Side(style='thin'))
            r += 1
        r += 1
    ws.cell(row=r + 1, column=1, value='Direct-only lines (not multiplied), $m').font = H3
    for y in range(NYR):
        ws.cell(row=r + 1, column=3 + y, value=(
            f'=SUMIFS(CALC_Strip!{CL(6 + y)}${P0}:{CL(6 + y)}${P1},'
            f'CALC_Strip!$D${P0}:$D${P1},"Domestic",'
            f'CALC_Strip!$E${P0}:$E${P1},"Y")')).number_format = MONEY
    ws.cell(row=r + 2, column=1, value='Conversion ratio: direct domestic / total spend').font = H3
    for y in range(NYR):
        ws.cell(row=r + 2, column=3 + y, value=(
            f'=IFERROR(CALC_Vector!{CL(3 + y)}{VTOT}/IN_Shock!{CL(6 + y)}{I1 + 2},0)')
        ).number_format = PCT
    for col, w in zip(['A', 'B'], [30, 28]):
        ws.column_dimensions[col].width = w

    # ======================================================== OUT_Detail
    ws = wb.create_sheet('OUT_Detail')
    band(ws, 1, 'OUT_Detail  -  total output impact by industry, year 1', 8)
    hdr(ws, 5, ['Code', 'Name', 'Direct $m', 'Total output $m', 'Share of total'])
    DT0 = 6
    ocol = mcol('Output multipliers', i_totl)
    for i in range(NS):
        r = DT0 + i
        vr = V0 + i
        c = ws.cell(row=r, column=1, value=f'=CALC_Vector!$A{vr}')
        c.number_format = '@'
        c.font = GREEN
        ws.cell(row=r, column=2, value=f'=CALC_Vector!$B{vr}').font = GREEN
        ws.cell(row=r, column=3, value=f'=CALC_Vector!$C{vr}').number_format = MONEY
        ws.cell(row=r, column=4, value=(
            f'=CALC_Vector!$C{vr}*IFERROR(MAP_Multipliers!${CL(ocol)}${M0 + i},0)')
        ).number_format = MONEY
        ws.cell(row=r, column=5, value=(
            f'=IFERROR($D{r}/SUM($D${DT0}:$D${DT0 + NS - 1}),0)')).number_format = PCT
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 40
    for c in 'CDE':
        ws.column_dimensions[c].width = 15
    ws.freeze_panes = 'C6'

    # ======================================================== QA_Checks
    ws = wb.create_sheet('QA_Checks')
    band(ws, 1, 'QA_Checks  -  read this before believing any number', 6)
    ws.cell(row=3, column=1, value='OVERALL').font = H2
    hdr(ws, 6, ['#', 'Gate', 'Result', 'What it catches'])
    Q0 = 7
    gates = []

    def gate(desc, formula, catches):
        gates.append((desc, formula, catches))

    gate('Multipliers loaded for the selected region',
         f'=IF(COUNT(MAP_Multipliers!${CL(mcol("Output multipliers", i_simp))}${M0}:'
         f'${CL(mcol("Output multipliers", i_simp))}${M1})={NS},"PASS","FAIL")',
         'Running a study on a region with no data')
    gate('Selected region resolves in Table 5',
         f'=IF(COUNTIF(MAP_Spine!$E${S0}:$E${S1},0)=0,"PASS","FAIL")',
         'A region key that does not match the RAW block')
    gate('Every code found in the margin tables',
         f'=IF(COUNTIF(MAP_Spine!$H${S0}:$H${S1},0)=0,"PASS","FAIL")',
         'A spine mismatch against the ABS 115 codes')
    gate('Shares sum to 100% for every product and group',
         f'=IF(SUMPRODUCT(--(ABS(CALC_Rates!${CL(CHK0)}${R0}:${CL(CHK0 + NG - 1)}${R1}-1)>0.0001))'
         f'=0,"PASS","FAIL")',
         'A broken rate table')
    gate('No negative domestic or import share',
         f'=IF(SUMPRODUCT(--(CALC_Rates!${CL(RATE0)}${R0}:${CL(RATE0 + 2 * NG - 1)}${R1}<0))=0,'
         f'"PASS","REVIEW")',
         'ABS cells with net disposals or import adjustments')
    gate('Strip components sum back to the line total',
         f'=IF(ABS(SUM(CALC_Strip!$F${P0}:$F${P1})-IN_Shock!$F${I1 + 2})<0.01,"PASS","FAIL")',
         'Margins lost or double counted')
    gate('Every included line lands on a non-empty ABS cell',
         f'=IF(SUMPRODUCT(--(IN_Shock!$F${I0}:$F${I1}>0),'
         f'--(IN_Shock!${CL(6 + NYR)}${I0}:${CL(6 + NYR)}${I1}=0))=0,"PASS","REVIEW")',
         'The bill-of-quantities-into-GFCF trap: a line with no ABS cell to strip against')
    gate('Simple = initial + production induced',
         f'=IF(SUMPRODUCT(--(ABS(MAP_Multipliers!${CL(mcol("Output multipliers", i_simp))}${M0}:'
         f'${CL(mcol("Output multipliers", i_simp))}${M1}'
         f'-MAP_Multipliers!${CL(mcol("Output multipliers", i_init))}${M0}:'
         f'${CL(mcol("Output multipliers", i_init))}${M1}'
         f'-MAP_Multipliers!${CL(mcol("Output multipliers", i_prod))}${M0}:'
         f'${CL(mcol("Output multipliers", i_prod))}${M1})>0.0001))=0,"PASS","FAIL")',
         'Column misalignment on paste')
    gate('Type II not smaller than Type I',
         f'=IF(SUMPRODUCT(--(MAP_Multipliers!${CL(mcol("Output multipliers", i_totl))}${M0}:'
         f'${CL(mcol("Output multipliers", i_totl))}${M1}'
         f'<MAP_Multipliers!${CL(mcol("Output multipliers", i_simp))}${M0}:'
         f'${CL(mcol("Output multipliers", i_simp))}${M1}-0.0001))=0,"PASS","FAIL")',
         'A block swap on paste')
    gate('6700 is bridged to 6701, not real data',
         f'=IF(SUM(Lists!$E${L0}:$E${L1})=0,"PASS","REVIEW")',
         'Imputed rent borrows actual rent. Wrong for housing studies')
    gate('Vintages match',
         '=IF(Settings!$B$9=Settings!$B$10,"PASS","REVIEW")',
         'Multipliers 2022-23 against ABS strip data 2023-24')
    gate('State run is using national margin and tax rates',
         f'=IF({REGION}="Aus","PASS","REVIEW")',
         'The ABS publishes no state margin or tax matrices. Disclose it')
    gate('Margin control total holds',
         f'=IF(ABS(SUMIFS(RAW_Margins!${CL(RS.marg_col(127))}${geo["RAW_Margins"]["first"]}:'
         f'${CL(RS.marg_col(127))}${geo["RAW_Margins"]["last"]},'
         f'RAW_Margins!$D${geo["RAW_Margins"]["first"]}:$D${geo["RAW_Margins"]["last"]},'
         f'"Product",RAW_Margins!$A${geo["RAW_Margins"]["first"]}:'
         f'$A${geo["RAW_Margins"]["last"]},"<35")-{src["margin_total_spine"]})<1,"PASS","FAIL")',
         'The ABS margin tables are the wrong vintage or incomplete. A subset '
         'build fails this by construction - it holds only part of the spine')
    gate('Direct vector reconciles to the strip',
         f'=IF(ABS(CALC_Vector!$C${VTOT}-(SUMIFS(CALC_Strip!$F${P0}:$F${P1},'
         f'CALC_Strip!$D${P0}:$D${P1},"Domestic",CALC_Strip!$E${P0}:$E${P1},"N")'
         f'+SUM(CALC_Margins!$C${MG0}:$C${MG1})))<0.01,"PASS","FAIL")',
         'Margins lost between the strip and the vector')

    for i, (desc, formula, catches) in enumerate(gates):
        r = Q0 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=desc)
        c = ws.cell(row=r, column=3, value=formula)
        c.font = H3
        ws.cell(row=r, column=4, value=catches).font = NOTE
    Q1 = Q0 + len(gates) - 1
    ws.cell(row=4, column=1, value='STATUS').font = H2
    st = ws.cell(row=4, column=2, value=(
        f'=IF(COUNTIF($C${Q0}:$C${Q1},"FAIL")>0,"FAIL - do not use",'
        f'IF(COUNTIF($C${Q0}:$C${Q1},"REVIEW")>0,"REVIEW - disclose the items below","OK"))'))
    st.font = Font(bold=True, size=12)
    ws.cell(row=4, column=3, value=f'=COUNTIF($C${Q0}:$C${Q1},"FAIL")&" fail, "'
                                   f'&COUNTIF($C${Q0}:$C${Q1},"REVIEW")&" review, "'
                                   f'&COUNTIF($C${Q0}:$C${Q1},"PASS")&" pass"')
    for col, w in zip(['A', 'B', 'C', 'D'], [5, 52, 22, 62]):
        ws.column_dimensions[col].width = w

    # ======================================================== Assumptions
    ws = wb.create_sheet('Assumptions')
    band(ws, 1, 'Assumptions  -  attach this to every report', 4)
    items = [
        ('Multiplier source', 'Supplied by the provider, not derived in this model. '
         'The model reads the set and never computes a Leontief inverse.'),
        ('Multiplier vintage', '2022-23, all nine regions.'),
        ('Flow table vintage', '2022-23. Table 5 is direct allocation, so the intermediate '
         'quadrant is domestic only.'),
        ('Margin and tax vintage', 'ABS 2023-24. One year later than the multipliers.'),
        ('Regionalisation', 'Done by the provider. Verified here: all nine regions distinct, '
         'and the multipliers reconcile to the supplied Table 5 at 114/114 within 0.001.'),
        ('State margins and taxes', 'National rates. The ABS publishes no state margin or tax '
         'matrices.'),
        ('6700 Imputed rent', 'No row in the supplied 114-code set. Bridged to 6701. It carries '
         '$21m of net taxes and no margins, so the effect is small, but it is wrong for a '
         'housing study.'),
        ('Exports column', 'Table 5 and Table 8 treat re-exports differently, so the Q7 import '
         'share is unreliable.'),
        ('Employment units', 'FTE, from the state Table 5 column heading. Australia has no '
         'employment column in the supplied set.'),
        ('GVA basis', 'Value added at basic prices (P1+P2+P4), the ABS headline. Not market '
         'prices, which include taxes on products.'),
        ('ABS caveat on multipliers', 'Input-output multipliers assume no supply constraints, '
         'fixed prices and fixed input ratios, and are likely to significantly overstate '
         'impacts. Quote this, do not paraphrase it.'),
    ]
    hdr(ws, 4, ['Item', 'What to disclose'])
    for i, (a, b) in enumerate(items, 5):
        ws.cell(row=i, column=1, value=a).font = H3
        ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[i].height = 30
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 104

    nform = sum(1 for s_ in wb.worksheets for row in s_.iter_rows()
                for c in row if isinstance(c.value, str) and c.value.startswith('='))
    print(f'  {len(wb.sheetnames)} tabs, {nform:,} formulas')
    return wb


if __name__ == '__main__':
    wb = main()
    out = SUBSET_OUT if os.environ.get('IOMODEL_SUBSET') else OUTFILE
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f'wrote {out}  ({out.stat().st_size / 1e6:.1f} MB)')
