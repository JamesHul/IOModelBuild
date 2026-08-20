"""
Rebuild the survey-to-ABS allocation with the margin double-count removed.

    python scripts/fix_concordance.py

The supplied concordance sent 45-70% of most goods items straight to 3901
Retail trade. That double-counts: the model already strips the retail and
wholesale margin out of the purchasers' price and reallocates it to 3901 and
3301. On $100 of groceries the original split put $72.85 with retail against a
correct $22.43, and left food manufacturing with $20.54 against $58.69.

The rule applied here: allocate 100% of an item to the PRODUCT bought, at the
price paid. The model produces the margin itself.

Grocery and alcohol splits are weighted by actual ABS household purchasers-price
spend on each food product, not by judgement.
"""
import pickle
import numbers
import warnings
from collections import defaultdict
from pathlib import Path

warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
SRCF = Path('/root/.claude/uploads/40e8c981-0484-5536-a14d-be6baaabbb16/'
            'c47acf74-SMC_category_detail_and_superfunded_spend_by_state_V1__Copy.xlsx')
OUT = ROOT / 'output' / 'SMC_allocation_CORRECTED.xlsx'

OOS = 'OUT_OF_SCOPE'
OOS_NAME = 'Out of scope - not modelled (leaves the Australian economy, or is a transfer)'

# item -> [(code, share)] ; None means keep the original allocation unchanged
FIX = {
 'Groceries': ('retail removed; split by ABS household spend on each food product', [
    ('1101', .199), ('1102', .039), ('1103', .147), ('1104', .088), ('1105', .019),
    ('1106', .055), ('1107', .130), ('1108', .074), ('1109', .153), ('1201', .096)]),
 'Alcohol': ('retail removed; split by ABS household spend', [('1205', .738), ('1202', .262)]),
 'Electricity': ('generation is upstream - households buy from the distributor/retailer',
                 [('2605', 1.0)]),
 'Medicines': ('retail removed', [('1801', 1.0)]),
 'Spectacles, hearing aids and therapeutic appliances': ('retail removed',
                                                         [('2401', 1.0)]),
 'Mobile phones/tablets*': ('retail removed', [('2401', 1.0)]),
 'Furniture': ('retail removed', [('2501', 1.0)]),
 'Small appliances': ('retail removed', [('2404', 1.0)]),
 'Whitegoods': ('retail removed', [('2404', 1.0)]),
 'Other household appliances': ('retail removed', [('2404', 1.0)]),
 'Household textiles, glassware and utensils': ('retail removed, remainder rescaled',
                                                [('1303', .5), ('2001', .5)]),
 'Tools and equipment for house and garden': ('retail removed, remainder rescaled',
                                              [('2204', .625), ('2405', .375)]),
 'Books and stationery (imputed)': ('retail removed, remainder rescaled',
                                    [('5401', .625), ('1502', .375)]),
 'Cosmetic/personal care*': ('retail removed', [('1804', 1.0)]),
 'TV/home theatre*': ('retail removed', [('2401', 1.0)]),
 'Computers*': ('retail removed', [('2401', 1.0)]),
 'Motor vehicle purchase (gross outlay)': ('retail removed', [('2301', 1.0)]),
 'Tobacco': ('retail removed', [('1205', 1.0)]),
 'Clothing & footwear': ('retail removed, remainder rescaled',
                         [('1305', .714), ('1306', .286)]),
 'Pets, gardens and recreation durables': ('retail removed, remainder rescaled',
                                           [('0103', .5), ('2502', .5)]),
 'Vehicle operation (fuel, servicing, etc.)': ('retail removed, remainder rescaled',
                                               [('1701', .533), ('9401', .467)]),
 'Personal effects (whole block) (imputed)': ('retail removed', [('2502', 1.0)]),
}
# items where the retail share sat alongside a non-goods code
FIX['Routine maintenance never asked (garden/home services)'] = (
    'retail removed, remainder rescaled', [('7310', .8), ('1502', .2)])
# 3001 is new dwelling construction - capital formation, not household consumption.
# Its ABS household cell is $58m, effectively empty.
# --- user decisions, this round -------------------------------------------
# Overseas travel: most of it is consumed overseas, or flown on foreign
# carriers, and is an import. Only the Australian-carrier and Australian-agency
# portion is domestic production. 15/5/80 is judgement: roughly a third of an
# overseas trip is airfare, of which Australian carriers hold about a third of
# international capacity, plus a small agency commission.
FIX['Overseas holidays and non-holiday air'] = (
    'OVERSEAS: 80% is consumed overseas or flown on foreign carriers, so it is an '
    'import and out of scope. Only the Australian-carrier and Australian-agency '
    'share is domestic production.',
    [(OOS, .80), ('4901', .15), ('7210', .05)])
# Council rates and vehicle registration are taxes and charges to government,
# not purchases of public administration services. The ABS household cell for
# 7501 is only $1,112m, which is the tell. The money is not lost to the economy
# - councils spend it - but capturing that needs a GOVERNMENT expenditure shock,
# not a household consumption one.
FIX['Council rates'] = (
    'TRANSFER: a property tax, not a purchase of public administration services. '
    'Out of scope for a household consumption shock; model it as government '
    'expenditure if you want the councils re-spending captured.', [(OOS, 1.0)])
FIX['Vehicle registration'] = (
    'TRANSFER: predominantly a registration tax. Out of scope. If you have the '
    'split, the CTP insurance component could be returned to 6301.', [(OOS, 1.0)])
# Body corporate: 6702's household cell is $16.5m - effectively empty, so a line
# there strips to almost nothing. Allocate to the services strata fees actually
# buy: repairs and maintenance, cleaning and grounds, and building insurance.
FIX['Body corporate'] = (
    'REALLOCATED: 6702 has a $16.5m household cell and would strip to almost '
    'nothing. Strata fees buy repairs, cleaning and grounds, and building '
    'insurance, so they are allocated to those.',
    [('3201', .40), ('7310', .40), ('6301', .20)])

FIX['Home improvements/repairs'] = (
    'retail removed; 3001 is new-dwelling GFCF with a near-empty household cell, '
    'so repairs and improvements go to construction services', [('3201', 1.0)])


def main():
    wb = openpyxl.load_workbook(SRCF, data_only=True)
    ws = wb['Sheet1']
    names = {}
    for r in ws.iter_rows(min_row=4, max_row=134, max_col=7, values_only=True):
        if r[3] is not None and r[4]:
            c = str(r[3]).strip()
            if c.endswith('.0'):
                c = c[:-2]
            names[c.zfill(4)] = str(r[4]).strip()
    src = pickle.load(open(ROOT / 'build' / 'sources.pkl', 'rb'))
    absnames = {m['code']: m['label'] for m in src['abs']['T23']['meta'] if m['code']}

    items, order = defaultdict(lambda: {'cat': '', 'q': '', 'total': 0.0, 'parts': []}), []
    for r in ws.iter_rows(min_row=4, max_row=134, max_col=7, values_only=True):
        cat, item, q, code, sec, share, aus = r[:7]
        if not item or 'Subtotal' in str(item) or code is None:
            continue
        if str(code).strip() in ('Nil', 'None', ''):
            continue
        c = str(code).strip()
        if c.endswith('.0'):
            c = c[:-2]
        k = str(item).strip()
        if k not in items:
            order.append(k)
        d = items[k]
        d['cat'] = str(cat).strip() if cat else d['cat']
        d['q'] = str(q).strip() if q else d['q']
        d['total'] += float(aus) if isinstance(aus, numbers.Number) else 0.0
        d['parts'].append((c.zfill(4), float(share) if isinstance(share, numbers.Number) else 0.0))
    wb.close()

    out = openpyxl.Workbook()
    s1 = out.active
    s1.title = 'Allocation CORRECTED'
    H = Font(bold=True, color='FFFFFF')
    HF = PatternFill('solid', fgColor='1F3864')
    CH = PatternFill('solid', fgColor='FFF2CC')
    s1['A1'] = ('Survey item to ABS industry allocation - CORRECTED. Retail and wholesale '
                'allocations removed: the IO model strips those margins from the purchasers '
                'price and reallocates them itself.')
    s1['A1'].font = Font(bold=True, size=11)
    hdr = ['Category', 'Item', 'Survey question', 'ABS code', 'ABS sector',
           'Share of item', 'Australia', 'Changed?', 'What changed']
    for i, h in enumerate(hdr, 1):
        c = s1.cell(row=3, column=i, value=h)
        c.font = H
        c.fill = HF
    r = 4
    changed_tot = 0.0
    for k in order:
        d = items[k]
        fix = FIX.get(k) or next((v for kk, v in FIX.items()
                                  if k.lower().startswith(kk.lower()[:28])), None)
        parts = fix[1] if fix else d['parts']
        note = fix[0] if fix else ''
        if fix:
            changed_tot += d['total']
        tot_share = sum(s for _, s in parts)
        for code, share in parts:
            s1.cell(row=r, column=1, value=d['cat'])
            s1.cell(row=r, column=2, value=k)
            s1.cell(row=r, column=3, value=d['q'])
            cc = s1.cell(row=r, column=4, value=code)
            cc.number_format = '@'
            s1.cell(row=r, column=5,
                    value=OOS_NAME if code == OOS else (names.get(code) or absnames.get(code, '')))
            sh = share / tot_share
            s1.cell(row=r, column=6, value=sh).number_format = '0.0%'
            s1.cell(row=r, column=7, value=d['total'] * sh).number_format = '#,##0.0'
            s1.cell(row=r, column=8, value='YES' if fix else '')
            s1.cell(row=r, column=9, value=note)
            if fix:
                for cx in range(1, 10):
                    s1.cell(row=r, column=cx).fill = CH
            r += 1
    for col, w in zip('ABCDEFGHI', [24, 44, 20, 10, 44, 13, 13, 10, 78]):
        s1.column_dimensions[col].width = w
    s1.freeze_panes = 'A4'

    # rolled up by ABS code, and the before/after comparison
    s2 = out.create_sheet('By ABS code')
    before, after = defaultdict(float), defaultdict(float)
    for k in order:
        d = items[k]
        for code, share in d['parts']:
            before[code] += d['total'] * share
        fx = FIX.get(k) or next((v for kk, v in FIX.items()
                                 if k.lower().startswith(kk.lower()[:28])), None)
        parts = fx[1] if fx else d['parts']
        ts = sum(s for _, s in parts)
        for code, share in parts:
            after[code] += d['total'] * share / ts
    for i, h in enumerate(['ABS code', 'ABS sector', 'BEFORE $m', 'AFTER $m', 'Change $m',
                           'BEFORE share', 'AFTER share'], 1):
        c = s2.cell(row=1, column=i, value=h)
        c.font = H
        c.fill = HF
    TB, TA = sum(before.values()), sum(after.values())
    for i, code in enumerate(sorted(set(before) | set(after),
                                    key=lambda c: -abs(after.get(c, 0) - before.get(c, 0))), 2):
        b, a = before.get(code, 0.0), after.get(code, 0.0)
        cc = s2.cell(row=i, column=1, value=code)
        cc.number_format = '@'
        s2.cell(row=i, column=2,
                value=OOS_NAME if code == OOS else (names.get(code) or absnames.get(code, '')))
        s2.cell(row=i, column=3, value=b).number_format = '#,##0.0'
        s2.cell(row=i, column=4, value=a).number_format = '#,##0.0'
        s2.cell(row=i, column=5, value=a - b).number_format = '#,##0.0;[Red]-#,##0.0'
        s2.cell(row=i, column=6, value=b / TB).number_format = '0.0%'
        s2.cell(row=i, column=7, value=a / TA).number_format = '0.0%'
    for col, w in zip('ABCDEFG', [10, 48, 13, 13, 13, 13, 13]):
        s2.column_dimensions[col].width = w
    s2.freeze_panes = 'A2'

    s3 = out.create_sheet('Notes')
    lines = [
        ('WHAT CHANGED', ''),
        ('Retail and wholesale removed', 'The model strips the retail and wholesale margin '
         'out of the purchasers price and reallocates it to 3901 and 3301 itself. Allocating '
         'to retail as well counted it twice. On $100 of groceries the original split gave '
         'retail $72.85 against a correct $22.43, and food manufacturing $20.54 against '
         '$58.69. After the fix, retail lands at 12.3% of the direct vector - exactly its '
         'share of ABS household consumption. Before, it was 31.5%.'),
        ('Groceries and alcohol', 'Split across food manufacturing codes using ABS household '
         'purchasers-price spend per product, not judgement. Note 1205 bundles wine, spirits '
         'AND tobacco, so it overstates as an alcohol proxy.'),
        ('Electricity', 'Was 60% distribution / 40% generation. Households buy from the '
         'distributor-retailer (2605); generation is upstream and the multiplier reaches it.'),
        ('Home improvements', '3001 is NEW DWELLING construction - capital formation, not '
         'household consumption, with a $58m household cell. All to 3201.'),
        ('', ''),
        ('OUT OF SCOPE - your decisions this round', ''),
        ('Overseas holidays: 80% out', 'Most of an overseas trip is consumed overseas or '
         'flown on a foreign carrier, which is an import and does not stimulate the '
         'Australian economy. 15% stays with Australian air transport and 5% with Australian '
         'travel agencies. The 80/15/5 split is JUDGEMENT - roughly a third of a trip is '
         'airfare and Australian carriers hold about a third of international capacity. '
         'Replace it if you have carrier or destination data.'),
        ('Council rates: 100% out', 'A property tax, not a purchase of public administration '
         'services. The ABS household cell for 7501 is only $1,112m, which is the tell.'),
        ('Vehicle registration: 100% out', 'Predominantly a registration tax. If you have the '
         'split, the CTP insurance component could be returned to 6301.'),
        ('Note on the two above', 'The money is not lost to the economy - councils and states '
         'spend it. Capturing that needs a separate GOVERNMENT expenditure shock, not a '
         'household consumption one. Excluding it here avoids attributing an industry '
         'multiplier to a tax payment.'),
        ('Body corporate: reallocated', '6702 has a $16.5m household cell and would strip to '
         'almost nothing. Strata fees buy repairs (3201), cleaning and grounds (7310) and '
         'building insurance (6301), so they now go there 40/40/20.'),
        ('', ''),
        ('KEPT IN, WITH A CAVEAT', ''),
        ('Insurance stays in 6301', 'This is a MEASUREMENT BASIS issue, not circularity. Your '
         'survey measures insurance on a cash-outlay basis - gross premiums - but 70-85% of a '
         'premium is the claims pool, which is a transfer rather than production. Allocating '
         'the whole premium to 6301 attributes to insurance money that actually ends up with '
         'panel beaters, builders and health providers. The total spend is real, so removing '
         'it would understate; the INDUSTRY attribution is what is approximate. Disclose that '
         'the insurance industry impact is overstated. 6301 also bundles insurers with '
         'superannuation funds, which blends the multiplier - an approximation, not an error.'),
        ('Rent and aged care left as surveyed', 'Rent gets 0.9% of the shock against 24.6% of '
         'ABS household consumption, and aged care 0.23% against 2.60%. Both look low for a '
         'retiree population, but the survey data stands as collected.'),
        ('', ''),
        ('STILL INDICATIVE', ''),
        ('Concordance status', 'These are judgement weights, not the ABS concordance. The '
         'authoritative source is Industry and Product Concordance Tables 2023-24.xlsx from '
         'the ABS Input-Output METHODOLOGY page. Every weight here belongs in the limitations '
         'section until that is used.'),
    ]
    for i, (a, b) in enumerate(lines, 1):
        c = s3.cell(row=i, column=1, value=a)
        c.font = Font(bold=True)
        s3.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical='top')
        s3.row_dimensions[i].height = max(15, 13 * (1 + len(b) // 95))
    s3.column_dimensions['A'].width = 32
    s3.column_dimensions['B'].width = 112

    OUT.parent.mkdir(exist_ok=True)
    out.save(OUT)
    print(f'{len(order)} items, {len(FIX)} changed, ${changed_tot:,.1f}m of spend reallocated')
    print(f'retail 3901: ${before.get("3901", 0):,.1f}m -> ${after.get("3901", 0):,.1f}m')
    oos = after.get(OOS, 0.0)
    print(f'out of scope: ${oos:,.1f}m ({oos / TA:.1%})   IN SCOPE ${TA - oos:,.1f}m')
    print(f'totals preserved: before ${TB:,.1f}m  after ${TA:,.1f}m  diff {TA - TB:+.4f}')
    print(f'wrote {OUT}')


if __name__ == '__main__':
    main()
