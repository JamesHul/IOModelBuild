"""
The ONLY file that knows the layout of the input spreadsheets.

Everything downstream consumes build/sources.pkl, so when a file's shape changes
this is the single place to fix.

Two kinds of thing are stored for every source table:

  'verbatim'  the block exactly as it appears in the file, values untouched,
              including Total columns, re-exports rows, primary-input rows and
              any 'n.a.' text. This is what gets written to the RAW_ tabs.
  'index'     row and column labels, so the MAP layer can locate the 115-code
              spine inside the verbatim block without anyone editing the block.

Rule: never drop, pad, re-order or coerce anything on the way in. See CLAUDE.md.

    python scripts/load_sources.py
"""
import pickle
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ABS_DIR = ROOT / 'data' / 'abs'
SUP_DIR = ROOT / 'data' / 'supplied'
OUT = ROOT / 'build' / 'sources.pkl'
CODE_RE = re.compile(r'^\d{3,4}$')


def norm_code(v):
    """IOIG code as 4-char text. ABS files store 0101 as the number 101 in places,
    which is exactly how leading-zero codes get lost. Always normalise on read."""
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s.zfill(4) if CODE_RE.match(s) else None


REGIONS = ['Aus', 'NSW', 'Vic', 'QLD', 'SA', 'WA', 'Tas', 'NT', 'ACT']

# ABS margin and tax tables -> the IOIG that earns the margin. Verified against Table 21.
MARGIN_TABLES = {
    '23': ('Wholesale', '3301'), '24': ('Retail', '3901'), '25': ('RestHotelClub', '4501'),
    '26': ('Road', '4601'), '27': ('Rail', '4701'), '28': ('Pipeline', '4801'),
    '29': ('Water', '4801'), '30': ('Air', '4901'), '31': ('PortHandling', '5201'),
    '32': ('MarineIns', '6301'), '33': ('Gas', '2701'), '34': ('Electricity', '2605'),
}
# 2023-24 control totals on the 115-code spine, $m. Used by check_sources.py.
MARGIN_CONTROL = {'Wholesale': 182778, 'Retail': 146374, 'RestHotelClub': 4790, 'Road': 44056,
                  'Rail': 7171, 'Pipeline': 2392, 'Water': 120, 'Air': 1151, 'PortHandling': 1057,
                  'MarineIns': 23, 'Gas': 3608, 'Electricity': 27890}
MARGIN_TOTAL_SPINE = 421410       # published 422,034 less the re-exports row
NET_TAX_TOTAL_SPINE = 168673


def grid(ws, max_row=400, max_col=200):
    """Read a sheet into a list of tuples, values exactly as stored."""
    return [r for r in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True)]


def find_code_column(rows, min_hits=100):
    hits = {}
    for row in rows:
        for ci, v in enumerate(row, 1):
            if norm_code(v):
                hits[ci] = hits.get(ci, 0) + 1
    if not hits:
        return None
    col = max(hits, key=hits.get)
    return col if hits[col] >= min_hits else col


def spine_index(rows, code_col):
    """Map IOIG code -> row number (1-based) for the first occurrence of each code."""
    out = {}
    for ri, row in enumerate(rows, 1):
        if len(row) >= code_col:
            c = norm_code(row[code_col - 1])
            if c:
                out.setdefault(c, ri)
    return out


# ---------------------------------------------------------------- ABS loaders
def load_abs_table(path, sheet=None):
    """Load an ABS data cube verbatim. Works for Tables 5, 8 and 23-35."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = sheet or (wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0])
    rows = grid(wb[name])
    wb.close()
    code_col = find_code_column(rows)
    return {'verbatim': rows, 'sheet': name, 'code_col': code_col,
            'row_index': spine_index(rows, code_col) if code_col else {},
            'source': str(path.name)}


def load_all_abs():
    """Everything in data/abs/. Filenames are matched loosely so renames survive."""
    got = {}
    for p in sorted(ABS_DIR.glob('*.xls*')):
        stem = p.stem
        m = re.search(r'5209055001(\d{2})', stem)
        key = None
        if m:
            key = 'T' + str(int(m.group(1)))
        else:
            m2 = re.search(r'[Tt]able[ _-]?(\d{1,2})', stem)
            if m2:
                key = 'T' + m2.group(1)
        if key is None:
            print(f"  ? could not identify {p.name} - name it like 520905500123.xlsx or 'Table 23.xlsx'")
            continue
        got[key] = load_abs_table(p)
        print(f"  {key:5s} <- {p.name}  sheet '{got[key]['sheet']}'  "
              f"{len(got[key]['verbatim'])} rows, {len(got[key]['row_index'])} codes")
    return got


# ------------------------------------------------------- supplied data loaders
#
# Layout of "Piggy IO tables and multipliers_V2.xlsx", established by inspection.
# Every sheet in the file shares one grid, which is what makes the stacking safe:
#
#   r3   'Table 5' / 'Table 8'          r10  column IOIG codes
#   r4   '$m 2022-23'  <- the vintage   r11  region name in col B, column names
#   r12  'FROM INDUSTRY'                r13  first data row
#
#   col A  IOIG code        col B  industry name        col C..  the matrix
#
# Rows 13-136 are 114 real codes plus 10 'Dummy' rows (9901-9910). Then the
# primary inputs. Table 5 and Table 8 do NOT agree on where those sit:
#
#   Table 5:  r138 T1, r140-147 P1 P2 P3a P3c P3d P3b P4a P4b, r148 P6,
#             r150 Production, r152 GDP
#   Table 8:  r138 T1, r140-147 same, r148 P5, r150 Production, r151 P6,
#             r153 Total uses
#
# That disagreement is exactly why the stacked tabs carry a RowType column and
# the model sums primary inputs with SUMIF on it rather than by row number.
SUPPLIED_FILE = 'Piggy IO tables and multipliers_V2.xlsx'

# Sheet-name prefix per region, in the file's own order.
SHEET_REGION = {'Aus': 'Aus', 'NSW': 'NSW', 'Vic': 'Vic', 'QLD': 'Qld', 'SA': 'SA',
                'WA': 'WA', 'Tas': 'Tas', 'NT': 'NT', 'ACT': 'ACT'}

FIRST_DATA_ROW = 13
# Last row holding content, taken as the max across all nine regions so that no
# region's block is truncated. NT Table 5 runs to 165 where the others stop at
# 153; Table 8 runs to 158. Short blocks simply leave trailing rows empty.
BLOCK_BOUNDS = {          # (last_row, last_col)
    'T5':   (165, 141),
    'T8':   (158, 136),
    'MULT': (136, 181),
}


def classify_row(code, label):
    """
    RowType for a source row. This is the key the workbook's SUMIF uses to total
    primary inputs, so it must not depend on row position - see the T5/T8
    disagreement above.
    """
    c = (str(code).strip() if code is not None else '')
    l = (str(label).strip() if label is not None else '')
    if norm_code(c):
        return 'Dummy' if c.zfill(4).startswith('99') else 'Industry'
    if c == 'T1':
        return 'Total'
    if re.match(r'^P\d', c):
        return 'Primary'
    if 'Production' in l:
        return 'Production'
    if c.startswith('GDP') or c.startswith('GSP') or 'Gross Domestic Product' in l \
            or 'Gross State Product' in l:
        return 'Product'
    if 'Total uses' in l:
        return 'Total'
    return 'Other' if (c or l) else ''


def load_supplied_block(ws, kind):
    """
    One region's block, exactly as it appears. Values are never dropped, padded,
    re-ordered or coerced - the only thing added is an index so the MAP layer can
    find things without anyone editing the block.
    """
    last_row, last_col = BLOCK_BOUNDS[kind]
    verbatim, meta = [], []
    for ri in range(FIRST_DATA_ROW, last_row + 1):
        row = [ws.cell(row=ri, column=ci).value for ci in range(1, last_col + 1)]
        verbatim.append(row)
        code, label = row[0], (row[1] if len(row) > 1 else None)
        meta.append({'src_row': ri, 'code': norm_code(code),
                     'raw_code': ('' if code is None else str(code).strip()),
                     'label': ('' if label is None else str(label).strip()),
                     'row_type': classify_row(code, label)})
    header = [[ws.cell(row=ri, column=ci).value for ci in range(1, last_col + 1)]
              for ri in (10, 11)]
    row_index = {}
    for m in meta:
        if m['code']:
            row_index.setdefault(m['code'], m['src_row'])
    return {'verbatim': verbatim, 'meta': meta, 'header': header,
            'row_index': row_index, 'code_col': 1, 'first_data_col': 3,
            'vintage': str(ws.cell(row=4, column=2).value or '').strip(),
            'region_label': str(ws.cell(row=11, column=2).value or '').strip()}


def load_supplied_flows():
    """Table 5 and Table 8 for all nine regions, as supplied."""
    out = {}
    path = SUP_DIR / SUPPLIED_FILE
    if not path.exists():
        cand = sorted(SUP_DIR.glob('*.xls*'))
        if not cand:
            print("  ! nothing in data/supplied/")
            return out
        path = cand[0]
        print(f"  ! {SUPPLIED_FILE} not found, using {path.name}")
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    for region, prefix in SHEET_REGION.items():
        for kind, tag in (('T5', 'Table 5'), ('T8', 'Table 8')):
            sheet = f'{prefix} {tag} FY23'
            if sheet not in wb.sheetnames:
                print(f"  ! missing sheet {sheet!r} - {region} {kind} left empty")
                continue
            d = load_supplied_block(wb[sheet], kind)
            d['source'] = f'{path.name} :: {sheet}'
            out[(kind, region)] = d
        got = [k for k in ('T5', 'T8') if (k, region) in out]
        if got:
            n = len([m for m in out[(got[0], region)]['meta'] if m['row_type'] == 'Industry'])
            print(f"  {region:4s} {'+'.join(got):6s} {n} industry codes, "
                  f"vintage {out[(got[0], region)]['vintage']!r}")
    wb.close()
    return out


def load_supplied_multipliers():
    """
    The multiplier set for all nine regions, as supplied.

    15 measure blocks of 11 effects, 12 columns apart (11 effects + 1 spacer),
    starting at column C. 'n.a.' cells are kept AS TEXT - the MAP layer decides
    what they mean, RAW does not.

    Three of the fifteen blocks are the three GVA definitions. 'Value added
    multipliers' is the basic-prices one and the ABS headline; do not reach for
    the market-prices block, which includes taxes on products (P3).
    """
    out = {}
    path = SUP_DIR / SUPPLIED_FILE
    if not path.exists():
        cand = sorted(SUP_DIR.glob('*.xls*'))
        if not cand:
            return out
        path = cand[0]
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    blocks, effects, per_region = [], [], {}
    for region, prefix in SHEET_REGION.items():
        sheet = f'{prefix} multiplier summary FY23'
        if sheet not in wb.sheetnames:
            print(f"  ! missing sheet {sheet!r} - {region} multipliers left empty")
            continue
        ws = wb[sheet]
        d = load_supplied_block(ws, 'MULT')
        d['source'] = f'{path.name} :: {sheet}'
        per_region[region] = d
        if not blocks:      # block and effect names are identical across regions
            for ci in range(1, BLOCK_BOUNDS['MULT'][1] + 1):
                v = ws.cell(row=10, column=ci).value
                if v not in (None, ''):
                    blocks.append({'name': str(v).strip(), 'first_col': ci})
            first = blocks[0]['first_col']
            for ci in range(first, first + 11):
                effects.append(str(ws.cell(row=11, column=ci).value or '').strip())
        na = sum(1 for r in d['verbatim'] for v in r
                 if isinstance(v, str) and v.strip().lower() == 'n.a.')
        print(f"  {region:4s} {len(d['row_index'])} codes, {na} 'n.a.' cells kept as text")
    wb.close()
    if not per_region:
        return out
    # flat lookup for check_sources.py: REGION|CODE -> the 11 effects of block 1
    data = {}
    for region, d in per_region.items():
        first = blocks[0]['first_col']
        for m, row in zip(d['meta'], d['verbatim']):
            if m['code']:
                data[f"{region}|{m['code']}"] = list(row[first - 1:first + 10])
    return {'blocks': blocks, 'effects': effects, 'regions': per_region,
            'data': data, 'source': str(path.name)}


def main():
    ROOT.joinpath('build').mkdir(exist_ok=True)
    print("ABS source tables:")
    abs_tables = load_all_abs()
    print("\nSupplied flow tables:")
    flows = load_supplied_flows()
    print("\nSupplied multipliers:")
    mult = load_supplied_multipliers()

    sources = {'abs': abs_tables, 'flows': flows, 'multipliers': mult,
               'regions': REGIONS, 'margin_tables': MARGIN_TABLES,
               'margin_control': MARGIN_CONTROL,
               'margin_total_spine': MARGIN_TOTAL_SPINE,
               'net_tax_total_spine': NET_TAX_TOTAL_SPINE,
               'block_bounds': BLOCK_BOUNDS, 'first_data_row': FIRST_DATA_ROW}
    with open(OUT, 'wb') as f:
        pickle.dump(sources, f)
    print(f"\nwrote {OUT}")
    if not flows or not mult:
        print("Loaders still incomplete - check_sources.py will tell you what is missing.")


if __name__ == '__main__':
    main()
