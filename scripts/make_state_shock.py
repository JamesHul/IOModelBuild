"""
Push the corrected national allocation out to the states, and write a
model-ready long-format shock file.

    python scripts/make_state_shock.py

State shares are taken per code from the supplied indicative state file, and
the modal share is used for codes that newly gain money under the corrected
allocation. Out-of-scope rows are dropped, not zeroed - they are not spending
on Australian production.
"""
import numbers
import warnings
from collections import defaultdict, Counter
from pathlib import Path

warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
CORR = ROOT / 'output' / 'SMC_allocation_CORRECTED.xlsx'
STATEF = Path('/root/.claude/uploads/40e8c981-0484-5536-a14d-be6baaabbb16/'
              '6fd7af73-indicative_abs_spending_by_state.xlsx')
OUT = ROOT / 'output' / 'Super_shock_CORRECTED_by_state.xlsx'
GEO = ['National', 'NSW', 'VIC', 'QLD', 'WA', 'SA', 'TAS', 'ACT', 'NT']
OOS = 'OUT_OF_SCOPE'


def num(v):
    return float(v) if isinstance(v, numbers.Number) else 0.0


def main():
    wb = openpyxl.load_workbook(CORR, data_only=True)
    alloc, names = defaultdict(float), {}
    for r in wb['Allocation CORRECTED'].iter_rows(min_row=4, max_col=9, values_only=True):
        if r[3] is None:
            continue
        c = str(r[3]).strip()
        alloc[c] += num(r[6])
        names[c] = str(r[4]) if r[4] else ''
    wb.close()
    oos = alloc.pop(OOS, 0.0)

    wb = openpyxl.load_workbook(STATEF, data_only=True)
    shares, natl = {}, {}
    for r in wb['ABS spending long'].iter_rows(min_row=3, values_only=True):
        if not r or not r[0]:
            continue
        g, _t, code, _c, _bn, m = r[:6]
        c = str(code).strip()
        if c.endswith('.0'):
            c = c[:-2]
        c = c.zfill(4)
        shares.setdefault(c, {})[str(g).strip()] = num(m)
    wb.close()
    # per-code state share, and the modal share for codes with no history
    per, tally = {}, defaultdict(Counter)
    for c, v in shares.items():
        n = v.get('National', 0.0)
        if n <= 0:
            continue
        per[c] = {g: v.get(g, 0.0) / n for g in GEO if g != 'National'}
        for g, x in per[c].items():
            tally[g][round(x, 6)] += 1
    modal = {g: tally[g].most_common(1)[0][0] for g in tally}
    print('modal state shares: ' + ', '.join(f'{g} {modal[g]:.4f}' for g in modal))
    print(f'sum of modal shares {sum(modal.values()):.6f}')

    rows, used_modal = [], []
    for c in sorted(alloc):
        if alloc[c] <= 0:
            continue
        sh = per.get(c)
        if sh is None:
            sh = modal
            used_modal.append(c)
        rows.append(('National', 'National', c, names.get(c, ''), alloc[c]))
        for g in GEO[1:]:
            rows.append((g, 'State/Territory', c, names.get(c, ''), alloc[c] * sh[g]))
    print(f'{len(alloc)} in-scope codes; modal share used for {len(used_modal)}: {used_modal}')

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = 'ABS spending long'
    ws['A1'] = ('Superannuation-funded spending, CORRECTED allocation, in-scope only. '
                f'${oos:,.1f}m of out-of-scope spend (overseas travel, council rates, '
                'vehicle registration) has been removed.')
    hdr = ['Geography', 'Geography type', 'ABS code', 'ABS category',
           'Spending ($bn)', 'Spending ($m)']
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F3864')
    for i, (g, gt, code, nm, m) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=gt)
        cc = ws.cell(row=i, column=3, value=code)
        cc.number_format = '@'
        ws.cell(row=i, column=4, value=nm)
        ws.cell(row=i, column=5, value=m / 1000).number_format = '#,##0.000000'
        ws.cell(row=i, column=6, value=m).number_format = '#,##0.000'
    for col, w in zip('ABCDEF', [12, 16, 10, 46, 15, 15]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A3'
    OUT.parent.mkdir(exist_ok=True)
    out.save(OUT)

    tn = sum(m for g, _, _, _, m in rows if g == 'National')
    ts = sum(m for g, _, _, _, m in rows if g != 'National')
    print(f'\nnational ${tn:,.1f}m   states ${ts:,.1f}m   diff {ts - tn:+,.1f}m')
    print(f'out of scope removed ${oos:,.1f}m')
    print(f'{len(rows)} rows -> {OUT}')


if __name__ == '__main__':
    main()
