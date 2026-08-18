"""
Write the stacked RAW layer from build/sources.pkl.

    python scripts/build_raw_stacked.py

One tab per data type, all nine regions stacked inside it, keyed by a Region
column. This is a deliberate design choice, not just tidiness:

  * CLAUDE.md bans INDIRECT and OFFSET and requires INDEX/MATCH on the code.
    With one tab per region a region switch has only two possible shapes - a
    nine-deep nested IF repeated in every formula, or INDIRECT to build the tab
    name. Stacked, the region stops being part of a sheet name and becomes a
    value you can MATCH on, and the whole switch collapses to one pattern.

  * Table 5 and Table 8 disagree about where the primary inputs sit (T5 has P6
    at source row 148; T8 has P5 there and P6 at 151). The RowType column makes
    the output denominator a SUMIF over row type instead of a hardcoded row
    range, which is what CLAUDE.md means by "sum the P-rows dynamically".

  * Lookups are by key, not by position, so a block pasted a row out still
    resolves correctly. Position only matters to the paste workflow and to the
    QA gate that checks it.

Layout of every stacked tab:

    A       B     C        D       | E ...
    Region  Code  RowType  SrcRow  | the source block, pasted verbatim
                                   | (source col A lands in E, B in F, data G+)

The four key columns sit OUTSIDE the pasted rectangle, so re-pasting a region
cannot disturb them and they cannot disturb the data.

RAW holds source data verbatim - rule 1. Nothing here trims, pads, re-spines,
blanks or coerces. 'n.a.' stays text, Total columns stay, Dummy rows stay,
primary-input rows stay. All reshaping belongs in the MAP_ layer.
"""
import pickle
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as CL

ROOT = Path(__file__).resolve().parent.parent
SOURCES = ROOT / 'build' / 'sources.pkl'
OUTFILE = ROOT / 'output' / 'IO_RAW_Stacked.xlsx'

REGIONS = ['Aus', 'NSW', 'Vic', 'QLD', 'SA', 'WA', 'Tas', 'NT', 'ACT']
KEYS = ['Region', 'Code', 'RowType', 'SrcRow']
NKEY = len(KEYS)
HDR_ROWS = 2          # the source's own two header rows (source rows 10 and 11)
FIRST_DATA_ROW = 4 + HDR_ROWS + 1     # banner(3) + key header(1) + 2 source headers

H1 = Font(bold=True, size=13, color='FFFFFF')
H2 = Font(bold=True, size=10)
KEYFONT = Font(bold=False, size=9, color='333333')
BLUE = Font(color='0000CC')           # hardcoded input, per the workbook colour key
NOTE = Font(italic=True, size=9, color='666666')
TITLEFILL = PatternFill('solid', fgColor='1F3864')
KEYFILL = PatternFill('solid', fgColor='DDEBF7')
HDRFILL = PatternFill('solid', fgColor='F2F2F2')
BANDFILL = PatternFill('solid', fgColor='FFF2CC')
THIN = Side(style='thin', color='BFBFBF')


def write_stacked(wb, tabname, title, blocks, note, keydefs=None, hdr_src=(0, 1)):
    """
    blocks:  list of (block_id, loaded_block). Every block gets the same band
             height so a short one leaves trailing rows empty rather than
             shifting the ones below it.
    keydefs: list of (column header, fn(block_id, meta) -> value). Defaults to
             the region key set used by the flow and multiplier tabs; the margin
             tab passes ABS_Table and Margin instead.
    """
    if keydefs is None:
        keydefs = [('Region', lambda bid, m: bid),
                   ('Code', lambda bid, m: m['code'] or m['raw_code']),
                   ('RowType', lambda bid, m: m['row_type']),
                   ('SrcRow', lambda bid, m: m['src_row'])]
    nkey = len(keydefs)
    ws = wb.create_sheet(tabname)
    band = max(len(b['verbatim']) for _, b in blocks)
    width = max(len(r) for _, b in blocks for r in b['verbatim'])

    ws.cell(row=1, column=1, value=title).font = H1
    for c in range(1, nkey + width + 1):
        ws.cell(row=1, column=c).fill = TITLEFILL
    src = blocks[0][1].get('source', '')
    ws.cell(row=2, column=1, value=(
        f"Source: {src.split(' :: ')[0]}   |   vintage {blocks[0][1].get('vintage', '')}"
        f"   |   loaded {date.today():%Y-%m-%d}   |   {len(blocks)} blocks"
        f"   |   band height {band} rows")).font = NOTE
    ws.cell(row=3, column=1, value=note).font = NOTE

    hr = 4
    for i, (k, _) in enumerate(keydefs, 1):
        c = ws.cell(row=hr, column=i, value=k)
        c.font = H2
        c.fill = KEYFILL
        c.border = Border(bottom=THIN)
    ws.cell(row=hr, column=nkey + 1, value='── source block, verbatim ──▶').font = H2
    # the source's own header rows, written once
    for j, hi in enumerate(hdr_src):
        hrow = blocks[0][1]['header'][hi]
        r = hr + 1 + j
        for ci, v in enumerate(hrow, 1):
            cell = ws.cell(row=r, column=nkey + ci, value=v)
            cell.font = H2
            cell.fill = HDRFILL
            cell.alignment = Alignment(horizontal='center', wrap_text=False)

    bands = []
    row = FIRST_DATA_ROW
    for bid, b in blocks:
        start = row
        for m, vals in zip(b['meta'], b['verbatim']):
            for i, (name, fn) in enumerate(keydefs, 1):
                c = ws.cell(row=row, column=i, value=fn(bid, m))
                c.font = KEYFONT
                c.fill = KEYFILL
                if name == 'Code':
                    c.number_format = '@'      # codes are TEXT, always
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=nkey + ci, value=v)
                cell.font = BLUE
                if ci == 1:
                    cell.number_format = '@'   # source code column stays text
            row += 1
        # pad short blocks so every band is the same height
        while row - start < band:
            for i, (name, fn) in enumerate(keydefs, 1):
                c = ws.cell(row=row, column=i,
                            value=(fn(bid, {'code': '', 'raw_code': '', 'row_type': '',
                                            'src_row': None}) if name not in ('Code', 'RowType', 'SrcRow') else ''))
                c.font = KEYFONT
                c.fill = KEYFILL
            row += 1
        bands.append((bid, start, row - 1, b.get('source', '')))
        ws.cell(row=start, column=1).fill = BANDFILL

    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=nkey + 1)
    ws.auto_filter.ref = (f"A{hr}:{CL(nkey)}{row - 1}")
    for i, (name, _) in enumerate(keydefs, 1):
        ws.column_dimensions[CL(i)].width = {'Region': 8, 'Code': 9, 'RowType': 11,
                                             'SrcRow': 8, 'ABS_Table': 10,
                                             'Margin': 15}.get(name, 10)
    ws.column_dimensions[CL(nkey + 1)].width = 9
    ws.column_dimensions[CL(nkey + 2)].width = 34
    return bands, band, width, row - 1


def main():
    if not SOURCES.exists():
        raise SystemExit(f"{SOURCES} not found. Run scripts/load_sources.py first.")
    src = pickle.load(open(SOURCES, 'rb'))
    flows, mult = src['flows'], src['multipliers']
    if not flows or not mult:
        raise SystemExit("sources.pkl has no supplied data. Run load_sources.py.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    idx = wb.create_sheet('INDEX')
    summary = []

    for kind, tab, title, note in [
        ('T5', 'RAW_T5', 'RAW — Table 5, all regions stacked. Industry by industry, '
         'DIRECT allocation of imports (domestic only), $m 2022-23',
         'Verbatim. Domestic multipliers come from this table. Imports sit in the P6 '
         'primary-input row, not in the intermediate quadrant.'),
        ('T8', 'RAW_T8', 'RAW — Table 8, all regions stacked. Industry by industry, '
         'INDIRECT allocation of imports (imports embedded), $m 2022-23',
         'Verbatim. Used ONLY to derive imports as T8 less T5, cell by cell. '
         'Never build multipliers off this table.'),
    ]:
        blocks = [(r, flows[(kind, r)]) for r in REGIONS if (kind, r) in flows]
        bands, band, width, last = write_stacked(wb, tab, title, blocks, note)
        summary.append((tab, bands, band, width, last))
        print(f"  {tab:16s} {len(bands)} regions, band {band} rows, "
              f"{width} source cols, last row {last}")

    mblocks = [(r, mult['regions'][r]) for r in REGIONS if r in mult.get('regions', {})]
    bands, band, width, last = write_stacked(
        wb, 'RAW_Multipliers',
        'RAW — Multiplier set, all regions stacked. 15 measure blocks x 11 effects, '
        'FY23 (2022-23)',
        mblocks,
        "Verbatim, INCLUDING the 'n.a.' text cells - the MAP layer decides what they "
        "mean. Multipliers are an input and are never derived. 'Value added multipliers' "
        "is the basic-prices block and the ABS headline; 'at market prices' includes "
        "taxes on products and is NOT the headline.")
    summary.append(('RAW_Multipliers', bands, band, width, last))
    print(f"  RAW_Multipliers  {len(bands)} regions, band {band} rows, "
          f"{width} source cols, last row {last}")

    # ------------------------------------------------------------- margins
    # ABS Tables 23-34 (the twelve margins) and Table 35 (net taxes), stacked
    # with ABS_Table and Margin as the identity keys. Those two say only WHICH
    # source block a row came from. The margin -> earning-industry assignment
    # is a transformation, not identity, so it lives in Lists_MarginMap where
    # it can be audited in one place - see rule 2.
    #
    # There is no Region key here: the ABS publishes margin and tax matrices
    # nationally only. A state run uses national rates, which is a disclosed
    # limitation, not an oversight.
    abs_t = src.get('abs') or {}
    mt = src['margin_tables']
    mblocks = []
    for k in sorted(mt, key=int) + ['35']:
        key = 'T' + k
        if key not in abs_t:
            print(f"  ! {key} missing from data/abs/ - not stacked")
            continue
        mblocks.append((k, abs_t[key]))
    if mblocks:
        names = dict({k: v[0] for k, v in mt.items()}, **{'35': 'NetTaxes'})
        keydefs = [('ABS_Table', lambda bid, m: int(bid)),
                   ('Margin', lambda bid, m: names.get(bid, '')),
                   ('Code', lambda bid, m: m['code'] or m['raw_code']),
                   ('RowType', lambda bid, m: m['row_type']),
                   ('SrcRow', lambda bid, m: m['src_row'])]
        bands, band, width, last = write_stacked(
            wb, 'RAW_Margins',
            'RAW — ABS Tables 23-34 (margins) and 35 (net taxes), stacked. '
            'Product by using industry and final use, $m 2023-24. NATIONAL ONLY',
            mblocks,
            'Verbatim, including the Re-exports row and the Total row and columns. '
            'Filter on ABS_Table or Margin to isolate one table. Tables 23-34 sum to '
            '$421,410m on the 115-code spine and $422,034m including re-exports; '
            'Table 35 is $168,673m. Margin rates are national - the ABS publishes no '
            'state margin or tax matrices.',
            keydefs=keydefs, hdr_src=(0, 1))
        summary.append(('RAW_Margins', bands, band, width, last))
        print(f"  RAW_Margins      {len(bands)} tables, band {band} rows, "
              f"{width} source cols, last row {last}")

    # Table 21 - the independent control total, by earning industry.
    if 'T21' in abs_t:
        bands, band, width, last = write_stacked(
            wb, 'RAW_T21_Control',
            'RAW — ABS Table 21. Composition of supply by margin commodity, '
            '$m 2023-24. The independent control total',
            [('21', abs_t['T21'])],
            'Verbatim. Margin commodity total is $422,034m, which is Tables 23-34 '
            'including their Re-exports rows. Use as the gate on any re-paste.',
            keydefs=[('ABS_Table', lambda bid, m: int(bid)),
                     ('Code', lambda bid, m: m['code'] or m['raw_code']),
                     ('RowType', lambda bid, m: m['row_type']),
                     ('SrcRow', lambda bid, m: m['src_row'])],
            hdr_src=(0,))
        summary.append(('RAW_T21_Control', bands, band, width, last))
        print(f"  RAW_T21_Control  control totals, last row {last}")

    # Lists_MarginMap - the analytical assignment, kept OUT of RAW.
    lm = wb.create_sheet('Lists_MarginMap')
    lm.cell(row=1, column=1, value='Lists — margin table to earning IOIG').font = \
        Font(bold=True, size=13)
    lm.cell(row=2, column=1, value=(
        'This is a mapping, not source data, so it lives here rather than in a RAW '
        'tab (rule 2). Verified against ABS Table 21 to the dollar. Join to '
        'RAW_Margins on ABS_Table.')).font = NOTE
    for i, h in enumerate(['ABS_Table', 'Margin', 'Earning IOIG', 'Spine $m',
                           'Re-exports $m'], 1):
        c = lm.cell(row=4, column=i, value=h)
        c.font = H2
        c.fill = KEYFILL
    ctrl = src.get('margin_control', {})
    r = 4
    for k in sorted(mt, key=int):
        r += 1
        name, earner = mt[k]
        lm.cell(row=r, column=1, value=int(k))
        lm.cell(row=r, column=2, value=name)
        e = lm.cell(row=r, column=3, value=earner)
        e.number_format = '@'
        lm.cell(row=r, column=4, value=ctrl.get(name))
        if 'T' + k in abs_t:
            rex = next((row[126] for row, m in zip(abs_t['T' + k]['verbatim'],
                                                   abs_t['T' + k]['meta'])
                        if m['row_type'] == 'ReExports'), None)
            lm.cell(row=r, column=5, value=rex)
    r += 1
    lm.cell(row=r, column=2, value='TOTAL 23-34').font = H2
    lm.cell(row=r, column=4, value=src.get('margin_total_spine')).font = H2
    r += 2
    lm.cell(row=r, column=2, value='Net taxes (Table 35)')
    lm.cell(row=r, column=3, value='n/a — leakage to government')
    lm.cell(row=r, column=4, value=src.get('net_tax_total_spine'))
    for col, w in zip('ABCDE', [11, 16, 26, 12, 14]):
        lm.column_dimensions[col].width = w

    # ------------------------------------------------------------------ INDEX
    idx.cell(row=1, column=1, value='RAW stacked layer — index and paste guide').font = \
        Font(bold=True, size=14)
    idx.cell(row=2, column=1, value=(
        f"Generated {date.today():%Y-%m-%d} by scripts/build_raw_stacked.py from "
        f"build/sources.pkl. Do not hand-edit: change the script and regenerate.")).font = NOTE
    r = 4
    idx.cell(row=r, column=1, value='How to load a new vintage').font = H2
    for line in [
        'Every tab is: four key columns (A-D), then the source block pasted verbatim from column E.',
        'The keys sit outside the pasted rectangle, so re-pasting cannot disturb them.',
        'To replace a region: copy the source rows listed below from the provider file '
        '(starting at its column A), and paste into column E of that region\'s first row.',
        'Lookups are INDEX/MATCH on Region + Code, so a block pasted a row out still '
        'resolves - but fix it anyway, and the QA gate will flag it.',
        'Codes are TEXT. Paste values only; do not let Excel turn 0101 into 101.',
    ]:
        r += 1
        idx.cell(row=r, column=1, value='• ' + line)
    r += 2
    idx.cell(row=r, column=1, value='Row bands').font = H2
    r += 1
    for h, w in zip(['Tab', 'Region', 'First row', 'Last row', 'Source sheet'],
                    [18, 9, 10, 10, 52]):
        pass
    for i, h in enumerate(['Tab', 'Region', 'First row', 'Last row', 'Source sheet'], 1):
        c = idx.cell(row=r, column=i, value=h)
        c.font = H2
        c.fill = KEYFILL
    for tab, bands, band, width, last in summary:
        for region, s, e, source in bands:
            r += 1
            idx.cell(row=r, column=1, value=tab)
            idx.cell(row=r, column=2, value=region)
            idx.cell(row=r, column=3, value=s)
            idx.cell(row=r, column=4, value=e)
            idx.cell(row=r, column=5, value=source.split(' :: ')[-1])
    for col, w in zip('ABCDE', [18, 9, 10, 10, 52]):
        idx.column_dimensions[col].width = w
    idx.freeze_panes = 'A5'

    OUTFILE.parent.mkdir(exist_ok=True)
    wb.save(OUTFILE)
    print(f"\nwrote {OUTFILE}  ({OUTFILE.stat().st_size / 1e6:.1f} MB)")


if __name__ == '__main__':
    main()
