"""
Prove the stacked RAW tabs are verbatim, cell for cell, against the source file.

    python scripts/verify_stacked.py

Claiming "verbatim" is worthless without this. The check reads the generated
workbook back and compares every cell of every region band to the provider file
it came from.

Note on int vs float: this has produced false discrepancies twice before, when a
check used isinstance(x, float) and skipped values Excel had stored as int. Use
numbers.Number and compare numerically.
"""
import numbers
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SUPPLIED = ROOT / 'data' / 'supplied' / 'Piggy IO tables and multipliers_V2.xlsx'
STACKED = ROOT / 'output' / 'IO_RAW_Stacked.xlsx'

REGIONS = ['Aus', 'NSW', 'Vic', 'QLD', 'SA', 'WA', 'Tas', 'NT', 'ACT']
SHEET = {'Aus': 'Aus', 'NSW': 'NSW', 'Vic': 'Vic', 'QLD': 'Qld', 'SA': 'SA',
         'WA': 'WA', 'Tas': 'Tas', 'NT': 'NT', 'ACT': 'ACT'}
NKEY = 5   # Region, Code, RowType, SrcRow, Key
TABS = {'RAW_T5': ('%s Table 5 FY23', 165, 141),
        'RAW_T8': ('%s Table 8 FY23', 158, 136),
        'RAW_Multipliers': ('%s multiplier summary FY23', 136, 181)}
FIRST_SRC_ROW = 13


def empty(v):
    """An empty cell and a cell holding an empty string are the same thing.

    openpyxl writes '' back as an empty cell, which is a representation
    difference, not a data change. Collapsing the two here is what stops this
    check reporting 1,244 phantom mismatches - the same trap as int vs float.
    Note this is deliberately narrow: only '' and whitespace count as empty.
    A '0', a 0, or an 'n.a.' is real content and must still compare exactly.
    """
    return v is None or (isinstance(v, str) and v.strip() == '')


def same(a, b):
    """Equal as stored. Numeric compared numerically so int 5 == float 5.0."""
    if empty(a) and empty(b):
        return True
    if isinstance(a, numbers.Number) and isinstance(b, numbers.Number):
        if a == b:
            return True
        scale = max(abs(a), abs(b), 1.0)
        return abs(a - b) <= 1e-9 * scale
    if empty(a) or empty(b):
        return False
    return str(a).strip() == str(b).strip()


def main():
    for p in (SUPPLIED, STACKED):
        if not p.exists():
            sys.exit(f"missing {p}")
    srcwb = openpyxl.load_workbook(SUPPLIED, read_only=True, data_only=True)
    outwb = openpyxl.load_workbook(STACKED, read_only=True, data_only=True)

    total_cells = total_bad = 0
    na_seen = 0
    failures = []

    for tab, (pat, last_row, last_col) in TABS.items():
        ws = outwb[tab]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        # region -> first sheet row (1-based) in the stacked tab
        bands = {}
        for i, row in enumerate(grid, 1):
            if row and row[0] in REGIONS and isinstance(row[3], int):
                bands.setdefault(row[0], i)
        tab_bad = 0
        for region in REGIONS:
            sname = pat % SHEET[region]
            if sname not in srcwb.sheetnames or region not in bands:
                failures.append(f'{tab}/{region}: band or source sheet missing')
                continue
            sws = srcwb[sname]
            src = [list(r) for r in sws.iter_rows(
                min_row=FIRST_SRC_ROW, max_row=last_row, max_col=last_col,
                values_only=True)]
            start = bands[region]
            for ri, srow in enumerate(src):
                orow = grid[start - 1 + ri] if start - 1 + ri < len(grid) else []
                for ci, sv in enumerate(srow):
                    ov = orow[NKEY + ci] if NKEY + ci < len(orow) else None
                    total_cells += 1
                    if isinstance(sv, str) and sv.strip().lower() == 'n.a.':
                        na_seen += 1
                    if not same(sv, ov):
                        tab_bad += 1
                        if len(failures) < 15:
                            failures.append(
                                f'{tab}/{region} srcrow{FIRST_SRC_ROW + ri} '
                                f'col{ci + 1}: source={sv!r} stacked={ov!r}')
        total_bad += tab_bad
        print(f'  {tab:16s} {"OK" if not tab_bad else f"{tab_bad} MISMATCH"}')

    # key-column integrity: codes must be text, regions must all be present
    print('\nKey columns')
    for tab in TABS:
        ws = outwb[tab]
        rows = [r for r in ws.iter_rows(values_only=True) if r and r[0] in REGIONS]
        regs = sorted({r[0] for r in rows})
        nonstr = sum(1 for r in rows if r[1] not in (None, '') and not isinstance(r[1], str))
        rtypes = sorted({r[2] for r in rows if r[2]})
        print(f'  {tab:16s} regions={len(regs)}/9  non-text codes={nonstr}  rowtypes={rtypes}')
        if len(regs) != 9:
            failures.append(f'{tab}: only {len(regs)} regions')
        if nonstr:
            failures.append(f'{tab}: {nonstr} codes not stored as text')

    # ------------------------------------------------------------- margins
    # Same round-trip for the ABS margin and tax tables, which carry five key
    # columns (ABS_Table, Margin, Code, RowType, SrcRow) rather than four.
    import glob
    import os
    MKEY = 6   # ABS_Table, Margin, Code, RowType, SrcRow, Key
    ABS_FIRST, ABS_LAST, ABS_COLS = 4, 121, 127
    absfiles = {}
    for f in glob.glob(str(ROOT / 'data' / 'abs' / '*.xls*')):
        m = re.search(r'5209055001(\d{2})', os.path.basename(f))
        if m:
            absfiles[str(int(m.group(1)))] = f
    if 'RAW_Margins' in outwb.sheetnames and absfiles:
        ws = outwb['RAW_Margins']
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        bands = {}
        for i, row in enumerate(grid, 1):
            if row and isinstance(row[0], int) and isinstance(row[4], int):
                bands.setdefault(str(row[0]), i)
        mbad = 0
        for tbl, start in sorted(bands.items(), key=lambda kv: int(kv[0])):
            if tbl not in absfiles:
                failures.append(f'RAW_Margins/{tbl}: no source file')
                continue
            swb = openpyxl.load_workbook(absfiles[tbl], read_only=True, data_only=True)
            sn = next(n for n in swb.sheetnames if n.lower().startswith('table'))
            srows = [list(r) for r in swb[sn].iter_rows(
                min_row=ABS_FIRST, max_row=ABS_LAST, max_col=ABS_COLS, values_only=True)]
            swb.close()
            for ri, srow in enumerate(srows):
                orow = grid[start - 1 + ri] if start - 1 + ri < len(grid) else []
                for ci, sv in enumerate(srow):
                    ov = orow[MKEY + ci] if MKEY + ci < len(orow) else None
                    total_cells += 1
                    if not same(sv, ov):
                        mbad += 1
                        if len(failures) < 15:
                            failures.append(
                                f'RAW_Margins/T{tbl} srcrow{ABS_FIRST + ri} '
                                f'col{ci + 1}: source={sv!r} stacked={ov!r}')
        total_bad += mbad
        print(f'  {"RAW_Margins":16s} {"OK" if not mbad else f"{mbad} MISMATCH"}')

        # control totals read back out of the generated workbook
        print('\nMargin control totals (read from the workbook, not the source)')
        CTRL = {'Wholesale': 182778, 'Retail': 146374, 'RestHotelClub': 4790,
                'Road': 44056, 'Rail': 7171, 'Pipeline': 2392, 'Water': 120,
                'Air': 1151, 'PortHandling': 1057, 'MarineIns': 23, 'Gas': 3608,
                'Electricity': 27890}
        tot = rex_tot = 0.0
        for row in grid:
            if not row or not isinstance(row[0], int):
                continue
            if row[3] == 'Product' and isinstance(row[MKEY + 126], numbers.Number):
                if row[0] != 35:
                    tot += float(row[MKEY + 126])
            if row[3] == 'ReExports' and isinstance(row[MKEY + 126], numbers.Number):
                if row[0] != 35:
                    rex_tot += float(row[MKEY + 126])
        per = {}
        for row in grid:
            if row and isinstance(row[0], int) and row[3] == 'Product' \
                    and isinstance(row[MKEY + 126], numbers.Number):
                per[row[1]] = per.get(row[1], 0.0) + float(row[MKEY + 126])
        for name, gate in CTRL.items():
            got = per.get(name, 0.0)
            ok = abs(got - gate) < 0.5
            print(f'  {name:16s} {got:12,.0f}  gate {gate:9,.0f}  {"ok" if ok else "MISMATCH"}')
            if not ok:
                failures.append(f'margin control {name}: {got:,.0f} vs {gate:,.0f}')
        for label, got, gate in [('TOTAL 23-34 spine', tot, 421410),
                                 ('incl re-exports', tot + rex_tot, 422034),
                                 ('Table 35 net taxes', per.get('NetTaxes', 0.0), 168673)]:
            ok = abs(got - gate) < 0.5
            print(f'  {label:16s} {got:12,.0f}  gate {gate:9,.0f}  {"ok" if ok else "MISMATCH"}')
            if not ok:
                failures.append(f'{label}: {got:,.0f} vs {gate:,.0f}')

    srcwb.close()
    outwb.close()
    print('\n' + '=' * 64)
    print(f'{total_cells:,} cells compared, {total_bad} mismatches, '
          f"{na_seen:,} 'n.a.' text cells carried through")
    if failures:
        print('\nFAILURES:')
        for f in failures:
            print('  ' + f)
        sys.exit(1)
    print('VERBATIM CONFIRMED — stacked tabs match the source cell for cell.')


if __name__ == '__main__':
    main()
