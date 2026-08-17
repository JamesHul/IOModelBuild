"""
Report the shape of every spreadsheet in data/, so the loaders can be written
against what is actually there rather than what we assume is there.

Run this first, and again whenever a new data drop arrives.

    python scripts/inspect_inputs.py
    python scripts/inspect_inputs.py --full      # dump more rows per sheet
"""
import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / 'data'
CODE_RE = re.compile(r'^\d{3,4}$')


def norm_code(v):
    """IOIG code as 4-char text. ABS files store 0101 as the number 101 in places,
    which is exactly how leading-zero codes get lost. Always normalise on read."""
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s.zfill(4) if CODE_RE.match(s) else None



def looks_like_code(v):
    return norm_code(v) is not None


def scan_sheet(ws, name, full=False):
    rows = []
    limit = 400 if full else 200
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limit, max_col=200, values_only=True), 1):
        rows.append(row)
        if i >= limit:
            break
    if not rows:
        print(f"    [{name}] empty")
        return

    # where do the 4-digit IOIG codes live?
    col_hits, row_hits = {}, {}
    for ri, row in enumerate(rows, 1):
        for ci, v in enumerate(row, 1):
            if looks_like_code(v):
                col_hits[ci] = col_hits.get(ci, 0) + 1
                row_hits[ri] = row_hits.get(ri, 0) + 1
    code_col = max(col_hits, key=col_hits.get) if col_hits else None
    code_row = max(row_hits, key=row_hits.get) if row_hits else None

    print(f"    [{name}] dims={ws.calculate_dimension()}")
    if code_col:
        n = col_hits[code_col]
        first = next((ri for ri, r in enumerate(rows, 1)
                      if len(r) >= code_col and looks_like_code(r[code_col - 1])), None)
        print(f"       codes DOWN column {code_col} ({n} found, first at row {first})")
        if n and n not in (115, 116):
            print(f"       NOTE: expected 115 (or 116 with re-exports). Check for codes stored as numbers.")
    if code_row and row_hits[code_row] > 20:
        print(f"       codes ACROSS row {code_row} ({row_hits[code_row]} found)")

    # region labels?
    regions = set()
    for row in rows:
        for v in row[:6]:
            s = str(v).strip() if v is not None else ''
            if s in ('Aus', 'Australia', 'NSW', 'Vic', 'VIC', 'QLD', 'Qld', 'SA', 'WA',
                     'Tas', 'TAS', 'NT', 'ACT'):
                regions.add(s)
    if regions:
        print(f"       region labels seen: {sorted(regions)}")

    # text in an otherwise numeric grid (n.a., n.p., np) - these break formulas
    texty = {}
    for row in rows:
        for v in row:
            if isinstance(v, str) and v.strip().lower() in ('n.a.', 'na', 'n.p.', 'np', '-', '..'):
                texty[v.strip()] = texty.get(v.strip(), 0) + 1
    if texty:
        print(f"       non-numeric placeholders: {texty}  <- must be handled in MAP, not stripped from RAW")

    head = 12 if not full else 30
    print(f"       first {head} non-empty rows:")
    shown = 0
    for ri, row in enumerate(rows, 1):
        vals = [('' if v is None else str(v)[:18]) for v in row[:10]]
        if any(vals):
            print(f"         r{ri:<4} " + ' | '.join(vals))
            shown += 1
            if shown >= head:
                break


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--full', action='store_true')
    args = ap.parse_args()

    files = sorted(p for p in DATA.rglob('*') if p.suffix.lower() in ('.xlsx', '.xlsm', '.xls'))
    if not files:
        sys.exit(f"No spreadsheets found under {DATA}. Put the ABS files in data/abs/ "
                 f"and the supplied files in data/supplied/.")

    print(f"{len(files)} file(s) under {DATA}\n")
    for p in files:
        size = p.stat().st_size / 1e6
        print("=" * 100)
        print(f"{p.relative_to(ROOT)}   ({size:.1f} MB)")
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception as e:
            print(f"    could not open: {e}")
            continue
        print(f"  {len(wb.sheetnames)} sheet(s): {wb.sheetnames}")
        for sname in wb.sheetnames:
            try:
                scan_sheet(wb[sname], sname, args.full)
            except Exception as e:
                print(f"    [{sname}] scan failed: {e}")
        wb.close()
        print()


if __name__ == '__main__':
    main()
